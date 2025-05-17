import logging
import json
import pandas as pd
from typing import Dict, Any, List, Optional
import sqlite3

logger = logging.getLogger(__name__)

def export_to_excel(db_conn_pool, output_path: str, filter_criteria: Dict = None):
    """Export database to Excel with optimized formatting"""
    try:
        # Kiểm tra thư viện
        try:
            import pandas as pd
            from openpyxl import styles
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise ImportError("Thư viện 'openpyxl' chưa được cài đặt.")
                
        conn = db_conn_pool.get_connection()
        # Base query với các cột được sắp xếp hợp lý nhưng loại bỏ cột do_mat
        query = '''
            SELECT 
                d.id as "ID",
                d.file_name as "Tên File",
                d.created_at as "Ngày Tạo",
                v.cqbh_tren as "CQBH Trên",
                v.cqbh_duoi as "CQBH Dưới",
                v.so_ki_hieu as "Số Ký Hiệu",
                v.ngay_bh as "Ngày Ban Hành",
                v.do_khan as "Độ Khẩn",
                v.loai_vb as "Loại Văn Bản",
                v.nd_chinh as "Nội Dung Chính",
                v.noi_nhan as "Nơi Nhận",
                v.chuc_vu as "Chức Vụ",
                v.chu_ky as "Chữ Ký"
            FROM documents d
            LEFT JOIN document_versions v ON d.id = v.document_id
            WHERE v.version_number = (
                SELECT MAX(version_number) 
                FROM document_versions 
                WHERE document_id = d.id
            )
        '''

        # Xử lý filter criteria nếu có
        params = []
        if filter_criteria:
            conditions = []
            for field, value in filter_criteria.items():
                if value:
                    if field == 'id':
                        conditions.append("d.id = ?")
                        params.append(value)
                    elif field == 'file_name':
                        conditions.append("d.file_name LIKE ?")
                        params.append(f'%{value}%')
                    elif field == 'date_from':
                        conditions.append("d.created_at >= ?")
                        params.append(value)
                    elif field == 'date_to':
                        conditions.append("d.created_at <= ?")
                        params.append(value)
                    elif field in ['cqbh_tren', 'cqbh_duoi', 'so_ki_hieu', 'loai_vb',
                                'do_khan', 'ngay_bh', 'chuc_vu']:  # Đã loại bỏ do_mat
                        conditions.append(f"v.{field} LIKE ?")
                        params.append(f'%{value}%')
                    elif field == 'nd_chinh':
                        conditions.append("v.nd_chinh LIKE ?")
                        params.append(f'%{value}%')
            
            if conditions:
                query += " AND " + " AND ".join(conditions)
        
        query += " ORDER BY d.created_at DESC"
        
        # Thực thi query
        if params:
            df = pd.read_sql_query(query, conn, params=params)
        else:
            df = pd.read_sql_query(query, conn)
        
        # Xử lý datetime columns
        datetime_columns = ['Ngày Tạo', 'Ngày Ban Hành']
        for col in datetime_columns:
            if col in df.columns:
                # Áp dụng hàm parse cho từng giá trị trong cột - giả định _parse_vietnamese_date
                df[col] = df[col].apply(lambda x: x)  # Placeholder, implement parsing if needed

        # Export to Excel với formatting tối ưu
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Documents')
            
            workbook = writer.book
            worksheet = writer.sheets['Documents']
            
            # Định nghĩa styles
            header_style = styles.NamedStyle(name='header_style')
            header_style.font = styles.Font(bold=True, size=11)
            header_style.fill = styles.PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
            header_style.alignment = styles.Alignment(horizontal='center', vertical='center', wrap_text=True)
            header_style.border = styles.Border(
                left=styles.Side(style='thin'),
                right=styles.Side(style='thin'),
                top=styles.Side(style='thin'),
                bottom=styles.Side(style='thin')
            )

            # Style cho dữ liệu
            data_style = styles.NamedStyle(name='data_style')
            data_style.font = styles.Font(size=10)
            data_style.alignment = styles.Alignment(vertical='center', wrap_text=True)
            data_style.border = styles.Border(
                left=styles.Side(style='thin'),
                right=styles.Side(style='thin'),
                top=styles.Side(style='thin'),
                bottom=styles.Side(style='thin')
            )
            
            # Áp dụng style cho header
            for row in worksheet.iter_rows(min_row=1, max_row=1):
                for cell in row:
                    cell.style = header_style
            
            # Áp dụng style cho dữ liệu
            for row in worksheet.iter_rows(min_row=2):
                for cell in row:
                    cell.style = data_style
            
            # Auto-width for columns
            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        max_length = max(max_length, cell_length)
                
                # Cap width to reasonable size
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Đặt freeze panes để cố định header
            worksheet.freeze_panes = 'A2'
        
        logger.info(f"Successfully exported data to Excel: {output_path}")
        return True
    
    except Exception as e:
        logger.error(f"Error exporting to Excel: {str(e)}")
        raise

def export_to_json(db_conn_pool, output_path: str, filter_criteria: Dict = None):
    """Export database to JSON"""
    try:
        conn = db_conn_pool.get_connection()
        # Base query với các cột được sắp xếp hợp lý nhưng loại bỏ cột do_mat
        query = '''
            SELECT 
                d.id as "ID",
                d.file_name as "Tên File",
                d.created_at as "Ngày Tạo",
                v.cqbh_tren as "CQBH Trên",
                v.cqbh_duoi as "CQBH Dưới",
                v.so_ki_hieu as "Số Ký Hiệu",
                v.ngay_bh as "Ngày Ban Hành",
                v.do_khan as "Độ Khẩn",
                v.loai_vb as "Loại Văn Bản",
                v.nd_chinh as "Nội Dung Chính",
                v.noi_nhan as "Nơi Nhận",
                v.chuc_vu as "Chức Vụ",
                v.chu_ky as "Chữ Ký"
            FROM documents d
            LEFT JOIN document_versions v ON d.id = v.document_id
            WHERE v.version_number = (
                SELECT MAX(version_number) 
                FROM document_versions 
                WHERE document_id = d.id
            )
        '''

        # Xử lý filter criteria nếu có
        params = []
        if filter_criteria:
            conditions = []
            for field, value in filter_criteria.items():
                if value:
                    if field == 'id':
                        conditions.append("d.id = ?")
                        params.append(value)
                    elif field == 'file_name':
                        conditions.append("d.file_name LIKE ?")
                        params.append(f'%{value}%')
                    elif field == 'date_from':
                        conditions.append("d.created_at >= ?")
                        params.append(value)
                    elif field == 'date_to':
                        conditions.append("d.created_at <= ?")
                        params.append(value)
                    elif field in ['cqbh_tren', 'cqbh_duoi', 'so_ki_hieu', 'loai_vb',
                                'do_khan', 'ngay_bh', 'chuc_vu']:  # Đã loại bỏ do_mat
                        conditions.append(f"v.{field} LIKE ?")
                        params.append(f'%{value}%')
                    elif field == 'nd_chinh':
                        conditions.append("v.nd_chinh LIKE ?")
                        params.append(f'%{value}%')
            
            if conditions:
                query += " AND " + " AND ".join(conditions)
        
        query += " ORDER BY d.created_at DESC"
        
        # Thực thi query
        if params:
            df = pd.read_sql_query(query, conn, params=params)
        else:
            df = pd.read_sql_query(query, conn)
        
        # Chuyển đổi DataFrame thành JSON
        result_json = df.to_json(orient='records', force_ascii=False, date_format='iso')
        
        # Ghi ra file với encoding UTF-8
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(result_json)

        logger.info(f"Successfully exported data to JSON: {output_path}")
        return True
            
    except Exception as e:
        logger.error(f"Error exporting to JSON: {str(e)}")
        raise
