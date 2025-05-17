"""
Hệ thống OCR Văn Bản
Phần mềm quản lý và trích xuất thông tin từ văn bản hành chính
"""
from pdf2image import convert_from_path
from typing import List, Dict, Any, Optional, Union, Tuple
import os
import sys
from pathlib import Path
import fitz
import numpy as np
import cv2
from PIL import Image
import google.generativeai as genai
import json
import sqlite3
from ultralytics import YOLO
import time
from tqdm import tqdm
import logging
from datetime import datetime
import pandas as pd
import multiprocessing as mp
import threading
from queue import Empty, Queue
import qdarkstyle
from typing import List, Dict, Any, Optional, Tuple
import io
import shutil
import re
import traceback
from statistics_dialog import StatisticsDialog
from PyQt5.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, 
                           QPushButton, QLabel, QFileDialog, QTableWidget, QTableWidgetItem,
                           QLineEdit, QTextEdit, QScrollArea, QFrame, QSplitter, QMessageBox,
                           QComboBox, QMenu, QAction, QProgressBar, QDialog, QShortcut,
                           QCompleter, QRadioButton, QButtonGroup, QGroupBox, QTabWidget,
                           QHeaderView, QSpacerItem, QSizePolicy, QStatusBar, QToolBar, 
                           QToolButton, QStyle, QStyleFactory, QCalendarWidget, QDateEdit, 
                           QCheckBox, QStyledItemDelegate, QGraphicsDropShadowEffect, QGridLayout, QFormLayout, QInputDialog)
from PyQt5.QtCore import (Qt, QThread, pyqtSignal, QSize, QRect, QPoint, QTimer, QStringListModel,
                         QDate, QDateTime, QEvent, QPropertyAnimation, QEasingCurve, QSettings,
                         QModelIndex, QSortFilterProxyModel, QAbstractTableModel, QRegExp, QUrl)
from PyQt5.QtGui import (QImage, QPixmap, QPainter, QPen, QKeySequence, QFont, QIcon, QColor,
                       QBrush, QLinearGradient, QPalette, QFontDatabase, QCursor, QRegExpValidator,
                       QDesktopServices, QPainterPath, QStandardItemModel, QStandardItem)
import pytesseract
import easyocr
from PIL import Image, ImageEnhance, ImageDraw

# Cấu hình cơ bản
APP_VERSION = "2.0.0"
BASE_DIR = Path(__file__).parent
OUTPUT_DIR = BASE_DIR / 'output'
DATABASE_DIR = BASE_DIR / 'database'
LOGS_DIR = BASE_DIR / 'logs'
IMAGES_DIR = OUTPUT_DIR / 'images'
RESULTS_DIR = OUTPUT_DIR / 'results'
BACKUP_DIR = BASE_DIR / 'backup'
CONFIG_DIR = BASE_DIR / 'config'
TEMP_DIR = BASE_DIR / 'temp'
ICON_DIR = BASE_DIR / 'icons'

# Tạo thư mục
for dir_path in [OUTPUT_DIR, DATABASE_DIR, LOGS_DIR, IMAGES_DIR, 
                RESULTS_DIR, BACKUP_DIR, CONFIG_DIR, TEMP_DIR, ICON_DIR]:
    dir_path.mkdir(exist_ok=True, parents=True)

# Cấu hình logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(LOGS_DIR / 'ocr_app.log', encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger("OCRApp")

# Configuration Constants
DEFAULT_CONFIDENCE_THRESHOLD = 0.5
DATABASE_TIMEOUT = 30.0
MAX_RETRY_ATTEMPTS = 5
DEFAULT_USER = "OCR System"
DEFAULT_WAIT_CURSOR = True
AUTOSAVE_INTERVAL = 60000  # ms (1 minute)
MAX_RECENT_FILES = 10

#############################
# Database Connection Pool  #
#############################
class DBConnectionPool:
    """Thread-safe database connection pool for SQLite"""
    
    def __init__(self, db_path, max_connections=10, timeout=30.0):
        self.db_path = db_path
        self.max_connections = max_connections
        self.timeout = timeout
        self.connections = {}
        self.lock = threading.RLock()
        
    def get_connection(self):
        """Get a connection for the current thread"""
        thread_id = threading.get_ident()
        
        with self.lock:
            if thread_id not in self.connections:
                if len(self.connections) >= self.max_connections:
                    # Find and close the oldest connection
                    oldest_thread = min(self.connections.keys(), 
                                       key=lambda k: self.connections[k]['last_used'])
                    self.connections[oldest_thread]['conn'].close()
                    del self.connections[oldest_thread]
                
                # Create new connection
                conn = sqlite3.connect(self.db_path, timeout=self.timeout)
                conn.row_factory = sqlite3.Row
                self._optimize_connection(conn)
                
                self.connections[thread_id] = {
                    'conn': conn,
                    'last_used': time.time()
                }
            else:
                # Update last used time
                self.connections[thread_id]['last_used'] = time.time()
                
            return self.connections[thread_id]['conn']
    
    def _optimize_connection(self, conn):
        """Optimize SQLite connection settings"""
        conn.execute("PRAGMA journal_mode = WAL")  # Write-Ahead Logging
        conn.execute("PRAGMA synchronous = NORMAL")  # Faster writes, still safe
        conn.execute("PRAGMA busy_timeout = 30000")  # 30 second timeout
        conn.execute("PRAGMA temp_store = MEMORY")  # Store temp data in memory
        conn.execute("PRAGMA foreign_keys = ON")    # Enable foreign key constraints
        return conn
    
    def close_all(self):
        """Close all connections in the pool"""
        with self.lock:
            for conn_data in self.connections.values():
                try:
                    conn_data['conn'].close()
                except:
                    pass
            self.connections.clear()
    
    def execute_with_retry(self, query, params=None, max_retries=5):
        """Execute a query with retry for locked database"""
        conn = self.get_connection()
        
        for attempt in range(max_retries):
            try:
                cursor = conn.cursor()
                if params:
                    result = cursor.execute(query, params)
                else:
                    result = cursor.execute(query)
                conn.commit()
                return result
            except sqlite3.OperationalError as e:
                if "database is locked" in str(e) and attempt < max_retries - 1:
                    # Exponential backoff
                    sleep_time = 0.1 * (2 ** attempt)
                    logger.warning(f"Database locked, retrying in {sleep_time:.2f}s (attempt {attempt+1}/{max_retries})")
                    time.sleep(sleep_time)
                else:
                    conn.rollback()
                    raise
            except Exception as e:
                conn.rollback()
                raise

#############################
# Theme Manager & Styling   #
#############################
class ThemeManager:
    """Manages application theming and styling"""
    
    # Define color schemes
    COLOR_SCHEMES = {
        "light": {
            "primary": "#2c3e50",
            "secondary": "#3498db",
            "accent": "#e74c3c",
            "background": "#ecf0f1",
            "text": "#2c3e50",
            "border": "#bdc3c7",
            "hover": "#d6dbdf",
            "button_text": "#ffffff",
            "success": "#2ecc71",
            "warning": "#f39c12",
            "danger": "#e74c3c",
            "info": "#3498db"
        },
        "dark": {
            "primary": "#34495e",
            "secondary": "#2980b9",
            "accent": "#e74c3c",
            "background": "#2c3e50",
            "text": "#ecf0f1",
            "border": "#7f8c8d",
            "hover": "#34495e",
            "button_text": "#ffffff",
            "success": "#27ae60",
            "warning": "#f39c12",
            "danger": "#c0392b",
            "info": "#2980b9"
        }
    }
    
    def __init__(self):
        self.dark_mode = False
        self.custom_colors = {}
        self.font_size = 10
        self.font_family = "Segoe UI"
        self.load_theme_settings()

    def load_theme_settings(self):
        """Load theme settings from file"""
        try:
            settings_path = CONFIG_DIR / 'theme_settings.json'
            if settings_path.exists():
                with open(settings_path, 'r', encoding='utf-8') as f:
                    settings = json.load(f)
                    self.dark_mode = settings.get('dark_mode', False)
                    self.custom_colors = settings.get('custom_colors', {})
                    self.font_size = settings.get('font_size', 10)
                    self.font_family = settings.get('font_family', "Segoe UI")
            else:
                self.dark_mode = False
        except Exception as e:
            logger.error(f"Error loading theme settings: {str(e)}")
            self.dark_mode = False

    def save_theme_settings(self):
        """Save theme settings to file"""
        try:
            settings_path = CONFIG_DIR / 'theme_settings.json'
            with open(settings_path, 'w', encoding='utf-8') as f:
                settings = {
                    'dark_mode': self.dark_mode,
                    'custom_colors': self.custom_colors,
                    'font_size': self.font_size,
                    'font_family': self.font_family
                }
                json.dump(settings, f, indent=4)
        except Exception as e:
            logger.error(f"Error saving theme settings: {str(e)}")

    def toggle_theme(self, app: QApplication):
        """Toggle between light and dark mode"""
        try:
            self.dark_mode = not self.dark_mode
            self.apply_theme(app)
            self.save_theme_settings()
            return self.dark_mode
        except Exception as e:
            logger.error(f"Error toggling theme: {str(e)}")
            return self.dark_mode
    
    def apply_theme(self, app: QApplication):
        """Apply current theme to application"""
        if self.dark_mode:
            # Apply dark theme (qdarkstyle)
            stylesheet = qdarkstyle.load_stylesheet_pyqt5()
            app.setStyleSheet(stylesheet)
            
            # Set dark palette
            palette = QPalette()
            palette.setColor(QPalette.Window, QColor(53, 53, 53))
            palette.setColor(QPalette.WindowText, Qt.white)
            palette.setColor(QPalette.Base, QColor(25, 25, 25))
            palette.setColor(QPalette.AlternateBase, QColor(53, 53, 53))
            palette.setColor(QPalette.ToolTipBase, Qt.white)
            palette.setColor(QPalette.ToolTipText, Qt.white)
            palette.setColor(QPalette.Text, Qt.white)
            palette.setColor(QPalette.Button, QColor(53, 53, 53))
            palette.setColor(QPalette.ButtonText, Qt.white)
            palette.setColor(QPalette.BrightText, Qt.red)
            palette.setColor(QPalette.Link, QColor(42, 130, 218))
            palette.setColor(QPalette.Highlight, QColor(42, 130, 218))
            palette.setColor(QPalette.HighlightedText, Qt.black)
            app.setPalette(palette)
            
        else:
            # Apply light theme
            app.setStyleSheet("")
            app.setPalette(app.style().standardPalette())
            
            # Apply custom light stylesheet
            stylesheet = """
            QMainWindow {
                background-color: #f5f5f5;
            }
            QLabel {
                color: #333333;
            }
            QPushButton {
                background-color: #3498db;
                color: white;
                border: none;
                border-radius: 4px;
                padding: 6px 12px;
                min-height: 30px;
            }
            QPushButton:hover {
                background-color: #2980b9;
            }
            QPushButton:pressed {
                background-color: #1c6ea4;
            }
            QPushButton:disabled {
                background-color: #cccccc;
                color: #999999;
            }
            QLineEdit, QTextEdit, QComboBox {
                background-color: white;
                border: 1px solid #dddddd;
                border-radius: 4px;
                padding: 4px;
                selection-background-color: #3498db;
            }
            QTableWidget {
                background-color: white;
                alternate-background-color: #f9f9f9;
                selection-background-color: #3498db;
            }
            QHeaderView::section {
                background-color: #e6e6e6;
                padding: 4px;
                border: 1px solid #dddddd;
                font-weight: bold;
            }
            QTabWidget::pane {
                border: 1px solid #dddddd;
                background-color: white;
            }
            QTabBar::tab {
                background-color: #e6e6e6;
                border: 1px solid #dddddd;
                border-bottom: none;
                padding: 6px 12px;
                border-top-left-radius: 4px;
                border-top-right-radius: 4px;
            }
            QTabBar::tab:selected {
                background-color: white;
                border-bottom: 1px solid white;
            }
            QScrollBar {
                background-color: #f5f5f5;
            }
            QScrollBar::handle {
                background-color: #cccccc;
                border-radius: 4px;
            }
            QScrollBar::handle:hover {
                background-color: #bbbbbb;
            }
            QGroupBox {
                border: 1px solid #dddddd;
                border-radius: 4px;
                margin-top: 1ex;
                padding-top: 1ex;
                background-color: white;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                subcontrol-position: top center;
                padding: 0 5px;
                background-color: white;
                color: #3498db;
                font-weight: bold;
            }
            """
            app.setStyleSheet(stylesheet)
        
        # Apply font settings
        font = QFont(self.font_family, self.font_size)
        app.setFont(font)
    
    def get_color(self, color_name):
        """Get color from current theme"""
        scheme = "dark" if self.dark_mode else "light"
        
        # Check for custom color override
        if color_name in self.custom_colors:
            return QColor(self.custom_colors[color_name])
            
        # Use default theme color
        if color_name in self.COLOR_SCHEMES[scheme]:
            return QColor(self.COLOR_SCHEMES[scheme][color_name])
            
        # Fallback
        return QColor(self.dark_mode and "#ecf0f1" or "#2c3e50")
    
    def set_custom_color(self, color_name, color_value):
        """Set custom color override"""
        self.custom_colors[color_name] = color_value
        self.save_theme_settings()
    
    def set_font(self, family, size):
        """Set application font"""
        self.font_family = family
        self.font_size = size
        self.save_theme_settings()

#############################
#   PDF Viewer Component    #
#############################
class PDFViewer(QWidget):
    """PDF viewing widget with page navigation and detection visualization"""
    
    pageChanged = pyqtSignal(int)
    zoomChanged = pyqtSignal(float)
    customBoxCreated = pyqtSignal(QRect, int)  # Thêm signal cho box tự vẽ
    
    # Zoom levels in percentage
    ZOOM_LEVELS = [25, 50, 75, 100, 125, 150, 175, 200, 250, 300]

    def __init__(self, parent=None):
        super().__init__(parent)
        self.current_page = 0
        self.pages = []
        self.detection_boxes = []
        self.custom_boxes_by_page = {}  # Lưu custom box theo trang
        self.pdf_path = None
        self.is_updating = False
        self.zoom_level = 100  # percentage
        self.zoom_idx = 3      # index in ZOOM_LEVELS (100%)
        self.rotation = 0      # degrees (0, 90, 180, 270)
        self.highlight_text = ""
        self.dpi = 150         # Base DPI for rendering
        
        # Thêm biến cho việc vẽ box
        self.is_drawing = False
        self.draw_start = None
        self.draw_end = None
        self.current_drawn_rect = None
        
        self.setup_ui()

    def setup_ui(self):
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(0, 0, 0, 0)
        
        # Toolbar
        toolbar = QHBoxLayout()
        toolbar.setSpacing(5)
        
        # Navigation controls
        self.first_btn = QPushButton()
        self.first_btn.setIcon(QIcon(str(ICON_DIR / "first.png")))  
        self.first_btn.setToolTip("First Page")
        self.first_btn.clicked.connect(self.goto_first_page)
        self.first_btn.setStyleSheet("""
            QPushButton {
                background-color: #f0f0f0;
                border: 1px solid #ccc;
                border-radius: 4px;
                padding: 4px;
                min-width: 28px;
            }
            QPushButton:hover {
                background-color: #e0e0e0;
            }
        """)

        self.prev_btn = QPushButton()
        self.prev_btn.setIcon(QIcon(str(ICON_DIR / "previous.png")))
        self.prev_btn.setToolTip("Previous Page")
        self.prev_btn.clicked.connect(self.previous_page)
        self.prev_btn.setStyleSheet("""
            QPushButton {
                background-color: #f0f0f0;
                border: 1px solid #ccc;
                border-radius: 4px;
                padding: 4px;
                min-width: 28px;
            }
            QPushButton:hover {
                background-color: #e0e0e0;
            }
        """)

        # Page label
        self.page_label = QLabel("Page: 0/0")
        self.page_label.setAlignment(Qt.AlignCenter)
        self.page_label.setMinimumWidth(100)
        self.page_label.setStyleSheet("""
            QLabel {
                color: #333;
                font-weight: bold;
                padding: 2px 8px;
            }
        """)

        # Page jump controls
        self.page_edit = QLineEdit()
        self.page_edit.setMaximumWidth(50)
        self.page_edit.setPlaceholderText("#")
        self.page_edit.returnPressed.connect(self.jump_to_page)
        self.page_edit.setStyleSheet("""
            QLineEdit {
                background-color: white;
                border: 1px solid #ccc;
                border-radius: 4px;
                padding: 4px;
            }
        """)

        self.jump_btn = QPushButton("Go")
        self.jump_btn.setMaximumWidth(40)
        self.jump_btn.clicked.connect(self.jump_to_page)
        self.jump_btn.setStyleSheet("""
            QPushButton {
                background-color: #4CAF50;
                color: white;
                border: none;
                border-radius: 4px;
                padding: 4px 8px;
            }
            QPushButton:hover {
                background-color: #45a049;
            }
        """)

        # Next/Last buttons
        self.next_btn = QPushButton()
        self.next_btn.setIcon(QIcon(str(ICON_DIR / "next.png")))
        self.next_btn.setToolTip("Next Page")
        self.next_btn.clicked.connect(self.next_page)
        self.next_btn.setStyleSheet("""
            QPushButton {
                background-color: #f0f0f0;
                border: 1px solid #ccc;
                border-radius: 4px;
                padding: 4px;
                min-width: 28px;
            }
            QPushButton:hover {
                background-color: #e0e0e0;
            }
        """)

        self.last_btn = QPushButton()
        self.last_btn.setIcon(QIcon(str(ICON_DIR / "last.png")))
        self.last_btn.setToolTip("Last Page")
        self.last_btn.clicked.connect(self.goto_last_page)
        self.last_btn.setStyleSheet("""
            QPushButton {
                background-color: #f0f0f0;
                border: 1px solid #ccc;
                border-radius: 4px;
                padding: 4px;
                min-width: 28px;
            }
            QPushButton:hover {
                background-color: #e0e0e0;
            }
        """)

        # Zoom controls
        self.zoom_out_btn = QPushButton()
        self.zoom_out_btn.setIcon(QIcon(str(ICON_DIR / "zoom-out.png")))
        self.zoom_out_btn.setToolTip("Zoom Out")
        self.zoom_out_btn.clicked.connect(self.zoom_out)
        self.zoom_out_btn.setStyleSheet("""
            QPushButton {
                background-color: #e8e8e8;
                border: 1px solid #ccc;
                border-radius: 4px;
                padding: 4px;
                min-width: 28px;
            }
            QPushButton:hover {
                background-color: #d8d8d8;
            }
        """)

        self.zoom_in_btn = QPushButton()
        self.zoom_in_btn.setIcon(QIcon(str(ICON_DIR / "zoom-in.png")))
        self.zoom_in_btn.setToolTip("Zoom In")
        self.zoom_in_btn.clicked.connect(self.zoom_in)
        self.zoom_in_btn.setStyleSheet("""
            QPushButton {
                background-color: #e8e8e8;
                border: 1px solid #ccc;
                border-radius: 4px;
                padding: 4px;
                min-width: 28px;
            }
            QPushButton:hover {
                background-color: #d8d8d8;
            }
        """)

        # Rotate controls
        self.rotate_left_btn = QPushButton()
        self.rotate_left_btn.setIcon(QIcon(str(ICON_DIR / "rotate-left.png")))
        self.rotate_left_btn.setToolTip("Rotate Left")
        self.rotate_left_btn.clicked.connect(self.rotate_left)
        self.rotate_left_btn.setStyleSheet("""
            QPushButton {
                background-color: #e0e0e0;
                border: 1px solid #ccc;
                border-radius: 4px;
                padding: 4px;
                min-width: 28px;
            }
            QPushButton:hover {
                background-color: #d0d0d0;
            }
        """)

        self.rotate_right_btn = QPushButton()
        self.rotate_right_btn.setIcon(QIcon(str(ICON_DIR / "rotate-right.png")))
        self.rotate_right_btn.setToolTip("Rotate Right")
        self.rotate_right_btn.clicked.connect(self.rotate_right)
        self.rotate_right_btn.setStyleSheet("""
            QPushButton {
                background-color: #e0e0e0;
                border: 1px solid #ccc;
                border-radius: 4px;
                padding: 4px;
                min-width: 28px;
            }
            QPushButton:hover {
                background-color: #d0d0d0;
            }
        """)

        # Zoom label
        self.zoom_label = QLabel("100%")
        self.zoom_label.setAlignment(Qt.AlignCenter)
        self.zoom_label.setMinimumWidth(60)
        self.zoom_label.setStyleSheet("""
            QLabel {
                color: #333;
                font-weight: bold;
                padding: 2px 8px;
                background-color: #f5f5f5;
                border: 1px solid #ddd;
                border-radius: 4px;
            }
        """)
        # Add controls to toolbar
        toolbar.addWidget(self.first_btn)
        toolbar.addWidget(self.prev_btn)
        toolbar.addWidget(self.page_label)
        toolbar.addWidget(self.next_btn)
        toolbar.addWidget(self.last_btn)
        toolbar.addSpacing(10)
        toolbar.addWidget(self.page_edit)
        toolbar.addWidget(self.jump_btn)
        toolbar.addStretch()
        toolbar.addWidget(self.zoom_out_btn)
        toolbar.addWidget(self.zoom_label)
        toolbar.addWidget(self.zoom_in_btn)
        toolbar.addSpacing(10)
        toolbar.addWidget(self.rotate_left_btn)
        toolbar.addWidget(self.rotate_right_btn)
        
        # Image display in scrollable area
        self.scroll_area = QScrollArea()
        self.scroll_area.setWidgetResizable(True)
        self.scroll_area.setAlignment(Qt.AlignCenter)
        
        self.image_label = QLabel()
        self.image_label.setAlignment(Qt.AlignCenter)
        self.image_label.setBackgroundRole(QPalette.Base)
        
        # Quan trọng: Sử dụng đúng policy để không bị méo hình
        self.image_label.setSizePolicy(QSizePolicy.MinimumExpanding, QSizePolicy.MinimumExpanding)
        self.image_label.setScaledContents(False)  # Tắt scaled contents
        
        self.scroll_area.setWidget(self.image_label)
        
        # Add to main layout
        main_layout.addLayout(toolbar)
        main_layout.addWidget(self.scroll_area)
        
        # Set initial state
        self.update_controls()

    def load_pdf(self, pdf_path):
        """Load a PDF file using pdf2image"""
        try:
            from pdf2image import convert_from_path
            
            self.pdf_path = pdf_path
            if not os.path.exists(pdf_path):
                logger.error(f"PDF file not found: {pdf_path}")
                self.pages = []
                self.current_page = 0
                self.update_controls()
                return False
                
            # Convert PDF to images
            self.pages = convert_from_path(pdf_path, dpi=self.dpi)
            self.current_page = 0
            self.detection_boxes = []
            self.update_page_display()
            self.update_controls()
            return True
            
        except Exception as e:
            logger.error(f"Error loading PDF: {str(e)}")
            self.pages = []
            self.current_page = 0
            self.update_controls()
            return False

    def set_detection_boxes(self, boxes):
        """Set detection boxes to visualize"""
        self.detection_boxes = boxes
        # Không cần gọi update vì chúng ta không vẽ detection boxes nữa

    def set_highlight_text(self, text):
        """Set text to highlight in the document"""
        self.highlight_text = text
        # Note: Text highlighting may require different implementation with pdf2image
        self.update_page_display()

    def update_controls(self):
        """Update navigation controls based on document state"""
        has_doc = len(self.pages) > 0
        
        # Update page display
        if has_doc:
            self.page_label.setText(f"Page: {self.current_page + 1}/{len(self.pages)}")
        else:
            self.page_label.setText("Page: 0/0")
            
        # Enable/disable controls
        self.first_btn.setEnabled(has_doc and self.current_page > 0)
        self.prev_btn.setEnabled(has_doc and self.current_page > 0)
        self.next_btn.setEnabled(has_doc and self.current_page < len(self.pages) - 1)
        self.last_btn.setEnabled(has_doc and self.current_page < len(self.pages) - 1)
        self.page_edit.setEnabled(has_doc)
        self.jump_btn.setEnabled(has_doc)
        self.zoom_in_btn.setEnabled(has_doc and self.zoom_idx < len(self.ZOOM_LEVELS) - 1)
        self.zoom_out_btn.setEnabled(has_doc and self.zoom_idx > 0)
        self.rotate_left_btn.setEnabled(has_doc)
        self.rotate_right_btn.setEnabled(has_doc)
        
        # Update zoom display
        self.zoom_label.setText(f"{self.zoom_level}%")

    def update_page_display(self):
        """Render current page"""
        # Ngăn đệ quy
        if self.is_updating:
            return
            
        self.is_updating = True
        
        try:
            if not self.pages or self.current_page >= len(self.pages):
                self.image_label.clear()
                self.is_updating = False
                return

            # Reset drawing state khi chuyển trang
            self.draw_start = None
            self.draw_end = None
            self.is_drawing = False

            # Get the PIL Image for current page
            current_pil_image = self.pages[self.current_page]
            
            # Apply rotation if needed
            if self.rotation != 0:
                current_pil_image = current_pil_image.rotate(-self.rotation, expand=True)
            
            # Convert PIL Image to QImage for display
            img_data = current_pil_image.convert('RGB')
            width, height = img_data.size
            
            # Convert to QImage
            bytes_per_line = 3 * width
            q_image = QImage(img_data.tobytes('raw', 'RGB'), width, height, bytes_per_line, QImage.Format_RGB888)
            
            # Tạo QPixmap từ QImage
            pixmap = QPixmap.fromImage(q_image)
            
            # Tính toán tỷ lệ zoom
            zoom_factor = self.zoom_level / 100.0
            
            # Áp dụng zoom nếu cần
            if self.zoom_level != 100:
                display_width = int(pixmap.width() * zoom_factor)
                display_height = int(pixmap.height() * zoom_factor)
                scaled_pixmap = pixmap.scaled(
                    display_width, 
                    display_height,
                    Qt.KeepAspectRatio,
                    Qt.SmoothTransformation
                )
                self.pixmap = scaled_pixmap
            else:
                self.pixmap = pixmap
                
            # Vẽ các box trên pixmap
            self.draw_boxes_on_pixmap()
                
            # Tắt scaleContents để tránh méo hình
            self.image_label.setScaledContents(False)
            
            # Emit page changed signal
            self.pageChanged.emit(self.current_page)
                
        except Exception as e:
            print(f"Error updating page display: {str(e)}")
            # Create an empty white image as fallback
            fallback_img = QImage(600, 800, QImage.Format_RGB888)
            fallback_img.fill(Qt.white)
            self.pixmap = QPixmap.fromImage(fallback_img)
            self.draw_boxes_on_pixmap()
        
        finally:
            # Luôn đặt lại cờ này vào cuối hàm
            self.is_updating = False

    def draw_boxes_on_pixmap(self):
        """Vẽ tất cả các box trên pixmap"""
        if not hasattr(self, 'pixmap') or self.pixmap.isNull():
            return
            
        # Tạo bản sao để vẽ
        result_pixmap = QPixmap(self.pixmap)
        painter = QPainter(result_pixmap)
        
        # Không vẽ detection boxes từ YOLO
        # Chỉ vẽ custom boxes cho trang hiện tại
        if self.current_page in self.custom_boxes_by_page:
            painter.setPen(QPen(Qt.red, 2, Qt.DashLine))
            for box in self.custom_boxes_by_page[self.current_page]:
                painter.drawRect(box)
        
        # Vẽ box đang vẽ nếu có
        if self.is_drawing and self.draw_start and self.draw_end:
            painter.setPen(QPen(Qt.blue, 2, Qt.DashLine))
            rect = QRect(self.draw_start, self.draw_end).normalized()
            painter.drawRect(rect)
        
        painter.end()
        
        # Hiển thị lên label
        self.image_label.setPixmap(result_pixmap)

    def previous_page(self):
        """Go to previous page"""
        if self.pages and self.current_page > 0:
            self.current_page -= 1
            self.update_page_display()
            self.update_controls()

    def next_page(self):
        """Go to next page"""
        if self.pages and self.current_page < len(self.pages) - 1:
            self.current_page += 1
            self.update_page_display()
            self.update_controls()
            
    def goto_first_page(self):
        """Go to first page"""
        if self.pages and self.current_page > 0:
            self.current_page = 0
            self.update_page_display()
            self.update_controls()
            
    def goto_last_page(self):
        """Go to last page"""
        if self.pages and self.current_page < len(self.pages) - 1:
            self.current_page = len(self.pages) - 1
            self.update_page_display()
            self.update_controls()
            
    def jump_to_page(self):
        """Jump to specific page number"""
        if not self.pages:
            return
            
        try:
            page_num = int(self.page_edit.text())
            if page_num < 1:
                page_num = 1
            elif page_num > len(self.pages):
                page_num = len(self.pages)
                
            self.current_page = page_num - 1
            self.update_page_display()
            self.update_controls()
            self.page_edit.clear()
        except ValueError:
            # Invalid input, do nothing
            self.page_edit.clear()
            
    def zoom_in(self):
        """Increase zoom level"""
        if not self.pages:
            return
            
        if self.zoom_idx < len(self.ZOOM_LEVELS) - 1:
            self.zoom_idx += 1
            self.zoom_level = self.ZOOM_LEVELS[self.zoom_idx]
            self.update_page_display()
            self.update_controls()
            self.zoomChanged.emit(self.zoom_level / 100.0)
            
    def zoom_out(self):
        """Decrease zoom level"""
        if not self.pages:
            return
            
        if self.zoom_idx > 0:
            self.zoom_idx -= 1
            self.zoom_level = self.ZOOM_LEVELS[self.zoom_idx]
            self.update_page_display()
            self.update_controls()
            self.zoomChanged.emit(self.zoom_level / 100.0)
    
    def set_zoom(self, zoom_percentage):
        """Set zoom to specific percentage"""
        if not self.pages:
            return
            
        # Find closest zoom level
        self.zoom_level = min(self.ZOOM_LEVELS, key=lambda x: abs(x - zoom_percentage))
        self.zoom_idx = self.ZOOM_LEVELS.index(self.zoom_level)
        self.update_page_display()
        self.update_controls()
        self.zoomChanged.emit(self.zoom_level / 100.0)
        
    def rotate_left(self):
        """Rotate view 90 degrees counterclockwise"""
        if not self.pages:
            return
            
        self.rotation = (self.rotation - 90) % 360
        self.update_page_display()
        
    def rotate_right(self):
        """Rotate view 90 degrees clockwise"""
        if not self.pages:
            return
            
        self.rotation = (self.rotation + 90) % 360
        self.update_page_display()
        
    def reset_view(self):
        """Reset zoom and rotation to defaults"""
        if not self.pages:
            return
            
        self.zoom_level = 100
        self.zoom_idx = self.ZOOM_LEVELS.index(self.zoom_level)
        self.rotation = 0
        self.update_page_display()
        self.update_controls()
        self.zoomChanged.emit(self.zoom_level / 100.0)

    def resizeEvent(self, event):
        """Handle resize events"""
        super().resizeEvent(event)
        if self.pages:
            self.update_page_display()

    def clear(self):
        """Clear current display"""
        self.pages = []
        self.current_page = 0
        self.detection_boxes = []
        self.pdf_path = None
        self.zoom_level = 100
        self.zoom_idx = self.ZOOM_LEVELS.index(self.zoom_level)
        self.rotation = 0
        self.highlight_text = ""
        self.image_label.clear()
        self.update_controls()
        
    def get_current_page_image(self):
        """Get current page as QImage"""
        if not self.pages or self.current_page >= len(self.pages):
            return None
            
        try:
            pil_image = self.pages[self.current_page]
            img_data = pil_image.convert('RGB')
            width, height = img_data.size
            bytes_per_line = 3 * width
            return QImage(img_data.tobytes('raw', 'RGB'), width, height, bytes_per_line, QImage.Format_RGB888)
        except Exception as e:
            logger.error(f"Error getting page image: {str(e)}")
            return None

    def extract_text_from_current_page(self):
        """Extract text from current page"""
        # Note: pdf2image doesn't provide text extraction
        # You would need to use another library like PyPDF2 or pdfplumber for this
        # For now, return empty string 
        return ""

    # Thêm các hàm xử lý sự kiện chuột
    def mousePressEvent(self, event):
        """Xử lý sự kiện nhấn chuột"""
        if event.button() == Qt.LeftButton and self.pages:
            # Chỉ xử lý khi click vào image_label
            pos = self.mapToGlobal(event.pos())
            label_pos = self.image_label.mapToGlobal(QPoint(0, 0))
            
            # Kiểm tra nếu click vào vùng của image_label
            if self.image_label.rect().contains(self.image_label.mapFromGlobal(pos)):
                # Tính toán vị trí tương đối trong image_label
                rel_pos = QPoint(
                    pos.x() - label_pos.x(),
                    pos.y() - label_pos.y()
                )
                
                self.is_drawing = True
                self.draw_start = rel_pos
                self.draw_end = rel_pos
                self.update_drawing()

    def mouseMoveEvent(self, event):
        """Xử lý sự kiện di chuyển chuột"""
        if self.is_drawing and self.draw_start:
            pos = self.mapToGlobal(event.pos())
            label_pos = self.image_label.mapToGlobal(QPoint(0, 0))
            
            # Tính toán vị trí tương đối trong image_label
            rel_pos = QPoint(
                pos.x() - label_pos.x(),
                pos.y() - label_pos.y()
            )
            
            self.draw_end = rel_pos
            self.update_drawing()

    def mouseReleaseEvent(self, event):
        """Xử lý sự kiện thả chuột"""
        if self.is_drawing and event.button() == Qt.LeftButton:
            pos = self.mapToGlobal(event.pos())
            label_pos = self.image_label.mapToGlobal(QPoint(0, 0))
            
            # Tính toán vị trí tương đối trong image_label
            rel_pos = QPoint(
                pos.x() - label_pos.x(),
                pos.y() - label_pos.y()
            )
            
            self.draw_end = rel_pos
            self.is_drawing = False
            self.show_class_selection()

    def update_drawing(self):
        """Cập nhật hiển thị box đang vẽ"""
        if hasattr(self, 'pixmap'):
            self.draw_boxes_on_pixmap()

    def show_class_selection(self):
        """Hiển thị dialog chọn class sau khi vẽ box"""
        if not self.draw_start or not self.draw_end:
            return
            
        rect = QRect(self.draw_start, self.draw_end).normalized()
        if rect.width() < 10 or rect.height() < 10:
            # Box quá nhỏ, bỏ qua
            self.draw_start = None
            self.draw_end = None
            self.update_drawing()
            return
            
        # Tạo menu hiển thị các class để chọn
        menu = QMenu(self)
        class_actions = {}
        
        class_names = [
            'CQBH', 'Chu_Ky', 'Chuc_Vu', 'Do_Khan',
            'Loai_VB', 'ND_Chinh', 'Ngay_BH', 'Noi_Nhan', 'So_Ki_Hieu'
        ]
        
        for class_name in class_names:
            action = menu.addAction(class_name)
            class_actions[action] = class_name
            
        # Thêm tùy chọn hủy
        menu.addSeparator()
        cancel_action = menu.addAction("Hủy")
        
        # Hiển thị menu tại vị trí chuột
        chosen_action = menu.exec_(QCursor.pos())
        
        if chosen_action and chosen_action != cancel_action:
            class_name = class_actions[chosen_action]
            class_id = {'CQBH': 0, 'Chu_Ky': 1, 'Chuc_Vu': 2, 'Do_Khan': 3,
                      'Loai_VB': 4, 'ND_Chinh': 5, 'Ngay_BH': 6, 'Noi_Nhan': 7, 'So_Ki_Hieu': 8}[class_name]
            
            # Lưu box vào danh sách theo trang hiện tại với tỷ lệ hiển thị
            if self.current_page not in self.custom_boxes_by_page:
                self.custom_boxes_by_page[self.current_page] = []
            self.custom_boxes_by_page[self.current_page].append(rect)
            
            # Tính tọa độ trên ảnh gốc
            original_rect = self.convert_viewport_to_image_rect(rect)
            
            # Emit signal với rect và class_id
            self.customBoxCreated.emit(original_rect, class_id)
            
        # Reset vẽ
        self.draw_start = None
        self.draw_end = None
        self.update_drawing()
        
    def convert_viewport_to_image_rect(self, viewport_rect):
        """Chuyển đổi tọa độ từ viewport sang tọa độ ảnh gốc"""
        if not self.pixmap or self.pixmap.isNull():
            return viewport_rect
            
        zoom_factor = self.zoom_level / 100.0
        
        # Kích thước label hiển thị và pixmap
        display_size = self.image_label.size()
        pixmap_size = self.pixmap.size()
        
        # Tính offset dựa vào canh giữa ảnh trong label
        offset_x = max(0, (display_size.width() - pixmap_size.width()) / 2)
        offset_y = max(0, (display_size.height() - pixmap_size.height()) / 2)
        
        # Lấy giá trị scroll hiện tại
        h_scroll = self.scroll_area.horizontalScrollBar().value()
        v_scroll = self.scroll_area.verticalScrollBar().value()
        
        # Tính toán tọa độ thực trên ảnh gốc, xét đến zoom, offset và scroll
        x1 = int((viewport_rect.x() - offset_x + h_scroll) / zoom_factor)
        y1 = int((viewport_rect.y() - offset_y + v_scroll) / zoom_factor)
        width = int(viewport_rect.width() / zoom_factor)
        height = int(viewport_rect.height() / zoom_factor)
        
        # Đảm bảo tọa độ nằm trong phạm vi ảnh
        x1 = max(0, x1)
        y1 = max(0, y1)
        
        # Tạo và trả về rect mới đã chuẩn hóa
        return QRect(x1, y1, width, height)

    def add_custom_box(self, rect, class_id=None, page=None):
        """Thêm box tùy chỉnh từ bên ngoài"""
        page_num = page if page is not None else self.current_page
        
        zoom_factor = self.zoom_level / 100.0
        scaled_rect = QRect(
            int(rect.x() * zoom_factor),
            int(rect.y() * zoom_factor),
            int(rect.width() * zoom_factor),
            int(rect.height() * zoom_factor)
        )
        
        if page_num not in self.custom_boxes_by_page:
            self.custom_boxes_by_page[page_num] = []
        self.custom_boxes_by_page[page_num].append(scaled_rect)
        
        if page_num == self.current_page:
            self.update_drawing()

    def clear_custom_boxes(self, page=None):
        """Xóa tất cả custom boxes"""
        if page is not None:
            if page in self.custom_boxes_by_page:
                del self.custom_boxes_by_page[page]
        else:
            self.custom_boxes_by_page = {}
        self.draw_boxes_on_pixmap()
        
    def set_detection_boxes(self, boxes):
        """Set detection boxes to visualize"""
        self.detection_boxes = boxes
        # Không cần gọi update vì chúng ta không vẽ detection boxes nữa

#############################
#  Document Database Class  #
#############################
class DocumentDatabase:
    """Database manager for documents and OCR results"""
    
    def __init__(self, db_path=DATABASE_DIR / "documents.db"):
        self.db_path = db_path
        self.suggestions_cache = {}
        self.conn_pool = DBConnectionPool(db_path, max_connections=10, timeout=DATABASE_TIMEOUT)
        self.init_db()
        self.load_suggestions()

    def init_db(self):
        """Initialize database with tables"""
        with sqlite3.connect(self.db_path) as conn:
            cursor = conn.cursor()
            
            # Create documents table first with all required columns
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS documents (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    file_path TEXT NOT NULL,
                    file_name TEXT NOT NULL,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    last_modified TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    file_hash TEXT,
                    file_size INTEGER,
                    page_count INTEGER
                )
            ''')
            
            # Document versions table
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS document_versions (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    document_id INTEGER,
                    version_number INTEGER,
                    cqbh_tren TEXT,
                    cqbh_duoi TEXT,
                    so_ki_hieu TEXT,
                    loai_vb TEXT,
                    nd_chinh TEXT,
                    ngay_bh TEXT,
                    noi_nhan TEXT,
                    chuc_vu TEXT,
                    chu_ky TEXT,
                    do_khan TEXT,
                    modified_by TEXT,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (document_id) REFERENCES documents(id)
                )
            ''')
            
            # Page detections table
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS page_detections (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    document_id INTEGER,
                    page_number INTEGER,
                    detection_data TEXT,
                    page_text TEXT,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (document_id) REFERENCES documents(id)
                )
            ''')
            
            # Field suggestions table
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS field_suggestions (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    field_name TEXT NOT NULL,
                    value TEXT NOT NULL,
                    frequency INTEGER DEFAULT 1,
                    last_used TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    UNIQUE(field_name, value)
                )
            ''')
            
            # Document backup table
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS document_backups (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    document_id INTEGER,
                    backup_path TEXT NOT NULL,
                    reason TEXT,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (document_id) REFERENCES documents(id)
                )
            ''')
            
            # Document tags table
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS document_tags (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    document_id INTEGER,
                    tag_name TEXT NOT NULL,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (document_id) REFERENCES documents(id),
                    UNIQUE(document_id, tag_name)
                )
            ''')
            
            # Add indexes for performance
            cursor.execute('CREATE INDEX IF NOT EXISTS idx_doc_filename ON documents(file_name)')
            cursor.execute('CREATE INDEX IF NOT EXISTS idx_doc_lastmod ON documents(last_modified)')
            cursor.execute('CREATE INDEX IF NOT EXISTS idx_versions_docid ON document_versions(document_id)')
            cursor.execute('CREATE INDEX IF NOT EXISTS idx_detections_docid ON page_detections(document_id, page_number)')
            cursor.execute('CREATE INDEX IF NOT EXISTS idx_suggestions_field ON field_suggestions(field_name, frequency)')
            cursor.execute('CREATE INDEX IF NOT EXISTS idx_tags_docid ON document_tags(document_id)')
            
            conn.commit()

    def _convert_vn_date_to_standard(self, date_string):
        """
        Chuyển đổi chuỗi ngày tháng tiếng Việt sang định dạng chuẩn yyyy-mm-dd
        
        Args:
            date_string: Chuỗi ngày tháng cần chuyển đổi
            
        Returns:
            str: Chuỗi ngày tháng định dạng yyyy-mm-dd hoặc chuỗi gốc nếu không chuyển đổi được
        """
        if not isinstance(date_string, str):
            return date_string
            
        try:
            # Loại bỏ khoảng trắng thừa và chuyển về chữ thường
            date_string = date_string.lower().strip()
            
            # Xử lý format "ngày dd tháng mm năm yyyy"
            if "ngày" in date_string and "tháng" in date_string and "năm" in date_string:
                # Loại bỏ các từ không cần thiết
                date_string = date_string.replace("ngày", "").replace("tháng", "").replace("năm", "")
                
                # Tách và lấy các phần ngày, tháng, năm
                parts = [part.strip() for part in date_string.split() if part.strip()]
                if len(parts) >= 3:
                    day = int(parts[0])
                    month = int(parts[1])
                    year = int(parts[2])
                    
                    # Kiểm tra tính hợp lệ của ngày tháng
                    if 1 <= month <= 12 and 1 <= day <= 31 and 1900 <= year <= 9999:
                        return f"{year:04d}-{month:02d}-{day:02d}"
            
            # Xử lý format "dd/mm/yyyy"
            if "/" in date_string:
                parts = date_string.split("/")
                if len(parts) == 3:
                    day = int(parts[0])
                    month = int(parts[1])
                    year = int(parts[2])
                    
                    if 1 <= month <= 12 and 1 <= day <= 31 and 1900 <= year <= 9999:
                        return f"{year:04d}-{month:02d}-{day:02d}"
            
            # Xử lý format "dd-mm-yyyy"
            if "-" in date_string:
                parts = date_string.split("-")
                if len(parts) == 3:
                    day = int(parts[0])
                    month = int(parts[1])
                    year = int(parts[2])
                    
                    if 1 <= month <= 12 and 1 <= day <= 31 and 1900 <= year <= 9999:
                        return f"{year:04d}-{month:02d}-{day:02d}"
                        
            # Trả về chuỗi gốc nếu không match format nào
            return date_string
            
        except Exception as e:
            logger.error(f"Error converting date string '{date_string}': {str(e)}")
            return date_string

    def get_statistics(self):
        """Lấy thống kê từ database"""
        try:
            conn = self.conn_pool.get_connection()
            cursor = conn.cursor()
            stats = {}

            # Tổng số văn bản
            cursor.execute("SELECT COUNT(*) FROM documents")
            stats['total_documents'] = cursor.fetchone()[0]

            # Thống kê theo loại văn bản
            cursor.execute("""
                SELECT loai_vb, COUNT(*) as count 
                FROM document_versions 
                WHERE version_number = (
                    SELECT MAX(version_number) 
                    FROM document_versions v2 
                    WHERE v2.document_id = document_versions.document_id
                )
                GROUP BY loai_vb 
                ORDER BY count DESC
            """)
            stats['by_type'] = cursor.fetchall()

            # Thống kê theo độ khẩn
            cursor.execute("""
                SELECT do_khan, COUNT(*) as count 
                FROM document_versions 
                WHERE version_number = (
                    SELECT MAX(version_number) 
                    FROM document_versions v2 
                    WHERE v2.document_id = document_versions.document_id
                )
                GROUP BY do_khan 
                ORDER BY count DESC
            """)
            stats['by_urgency'] = cursor.fetchall()

            # Thống kê theo thời gian
            cursor.execute("""
                SELECT 
                    strftime('%Y-%m', created_at) as month,
                    COUNT(*) as count
                FROM documents
                GROUP BY month
                ORDER BY month DESC
                LIMIT 12
            """)
            stats['by_month'] = cursor.fetchall()

            # Văn bản mới nhất
            cursor.execute("""
                SELECT d.id, d.file_name, d.created_at, v.so_ki_hieu
                FROM documents d
                LEFT JOIN document_versions v ON d.id = v.document_id
                WHERE v.version_number = (
                    SELECT MAX(version_number) 
                    FROM document_versions 
                    WHERE document_id = d.id
                )
                ORDER BY d.created_at DESC
                LIMIT 5
            """)
            stats['recent_docs'] = cursor.fetchall()

            return stats

        except Exception as e:
            logger.error(f"Error getting statistics: {str(e)}")
            return None

    def create_backup(self, doc_id, file_path, reason="Manual backup"):
        """Create a backup of the document file"""
        try:
            if not os.path.exists(file_path):
                logger.warning(f"Cannot backup file that doesn't exist: {file_path}")
                return None
                
            backup_filename = f"{Path(file_path).stem}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
            backup_path = BACKUP_DIR / backup_filename
            
            # Copy file to backup
            shutil.copy2(file_path, backup_path)
            
            # Register backup in database
            conn = self.conn_pool.get_connection()
            cursor = conn.cursor()
            cursor.execute(
                "INSERT INTO document_backups (document_id, backup_path, reason) VALUES (?, ?, ?)",
                (doc_id, str(backup_path), reason)
            )
            conn.commit()
                
            logger.info(f"Created backup of document {doc_id} at {backup_path}")
            return str(backup_path)
            
        except Exception as e:
            logger.error(f"Error creating document backup: {str(e)}")
            return None

    def load_suggestions(self):
        """Load suggestions from database into cache"""
        try:
            conn = self.conn_pool.get_connection()
            cursor = conn.cursor()
            cursor.execute('''
                SELECT field_name, value, frequency
                FROM field_suggestions
                ORDER BY frequency DESC, last_used DESC
            ''')
            
            for field, value, freq in cursor.fetchall():
                if field not in self.suggestions_cache:
                    self.suggestions_cache[field] = []
                self.suggestions_cache[field].append({
                    'value': value,
                    'frequency': freq
                })
                
        except Exception as e:
            logger.error(f"Error loading suggestions: {str(e)}")

    def get_suggestions(self, field_name: str, prefix: str = "") -> List[str]:
        """Get suggestions for a field, optionally filtered by prefix"""
        if field_name not in self.suggestions_cache:
            return []
            
        suggestions = self.suggestions_cache[field_name]
        if prefix:
            suggestions = [s for s in suggestions 
                         if s['value'].lower().startswith(prefix.lower())]
            
        return [s['value'] for s in sorted(
            suggestions, 
            key=lambda x: x['frequency'], 
            reverse=True
        )]

    def add_suggestion(self, field_name: str, value: str):
        """Add or update a suggestion in the database"""
        if not value or not value.strip():
            return
            
        value = value.strip()
            
        try:
            conn = self.conn_pool.get_connection()
            with conn:  # Auto commit/rollback
                cursor = conn.cursor()
                
                # Check if suggestion exists
                cursor.execute('''
                    SELECT id, frequency FROM field_suggestions
                    WHERE field_name = ? AND value = ?
                ''', (field_name, value))
                
                result = cursor.fetchone()
                if result:
                    cursor.execute('''
                        UPDATE field_suggestions
                        SET frequency = frequency + 1,
                            last_used = CURRENT_TIMESTAMP
                        WHERE id = ?
                    ''', (result[0],))
                else:
                    cursor.execute('''
                        INSERT INTO field_suggestions (field_name, value)
                        VALUES (?, ?)
                    ''', (field_name, value))
                    
            # Update cache
            if field_name not in self.suggestions_cache:
                self.suggestions_cache[field_name] = []
                
            # Find suggestion in cache
            found = False
            for suggestion in self.suggestions_cache[field_name]:
                if suggestion['value'] == value:
                    suggestion['frequency'] += 1
                    found = True
                    break
                    
            # Add to cache if not found
            if not found:
                self.suggestions_cache[field_name].append({
                    'value': value,
                    'frequency': 1
                })
                
        except Exception as e:
            logger.error(f"Error adding suggestion: {str(e)}")

    def add_document(self, file_path: str, ocr_results: Dict[str, Any], page_count: int = None) -> int:
        """Create a new document in the database"""
        try:
            if not os.path.exists(file_path):
                raise FileNotFoundError(f"File not found: {file_path}")
                
            # Calculate file hash for duplicate detection
            file_hash = self._calculate_file_hash(file_path)
            file_size = os.path.getsize(file_path)
            
            # Check if this document already exists
            existing_id = self._check_duplicate_document(file_path, file_hash)
            if existing_id:
                logger.info(f"Document already exists with ID {existing_id}")
                return existing_id
            
            if page_count is None and Path(file_path).suffix.lower() == '.pdf':
                try:
                    # Sử dụng pdf2image để đếm số trang
                    from pdf2image.pdf2image import pdfinfo_from_path
                    info = pdfinfo_from_path(file_path)
                    page_count = info["Pages"]
                except:
                    page_count = None
                    
            # Insert document
            conn = self.conn_pool.get_connection()
            with conn:
                cursor = conn.cursor()
                cursor.execute('''
                    INSERT INTO documents (file_path, file_name, file_hash, file_size, page_count)
                    VALUES (?, ?, ?, ?, ?)
                ''', (file_path, Path(file_path).name, file_hash, file_size, page_count))
                
                doc_id = cursor.lastrowid
                
                # Insert first version
                cursor.execute('''
                    INSERT INTO document_versions (
                        document_id, version_number, cqbh_tren, cqbh_duoi,
                        so_ki_hieu, loai_vb, nd_chinh, ngay_bh,
                        noi_nhan, chuc_vu, chu_ky, do_khan, modified_by
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (
                    doc_id, 1,
                    ocr_results.get('CQBH_tren', ''),
                    ocr_results.get('CQBH_duoi', ''),
                    ocr_results.get('So_Ki_Hieu', ''),
                    ocr_results.get('Loai_VB', ''),
                    ocr_results.get('ND_Chinh', ''),
                    ocr_results.get('Ngay_BH', ''),
                    ocr_results.get('Noi_Nhan', ''),
                    ocr_results.get('Chuc_Vu', ''),
                    ocr_results.get('Chu_Ky', ''),
                    ocr_results.get('Do_Khan', 'Không'),
                    'OCR System'
                ))
            
            # Add suggestions for all fields
            for field, value in {
                'so_ki_hieu': ocr_results.get('So_Ki_Hieu', ''),
                'loai_vb': ocr_results.get('Loai_VB', ''),
                'chuc_vu': ocr_results.get('Chuc_Vu', ''),
                'cqbh_tren': ocr_results.get('CQBH_tren', ''),
                'cqbh_duoi': ocr_results.get('CQBH_duoi', ''),
                'do_khan': ocr_results.get('Do_Khan', 'Không')
            }.items():
                if value:
                    self.add_suggestion(field, value)
            return doc_id
            
        except Exception as e:
            logger.error(f"Error adding document: {str(e)}")
            raise
    
    def _calculate_file_hash(self, file_path: str) -> str:
        """Calculate SHA-256 hash of a file for deduplication"""
        import hashlib
        
        try:
            sha256_hash = hashlib.sha256()
            with open(file_path, "rb") as f:
                # Read the file in chunks to handle large files efficiently
                for byte_block in iter(lambda: f.read(4096), b""):
                    sha256_hash.update(byte_block)
            return sha256_hash.hexdigest()
        except Exception as e:
            logger.error(f"Error calculating file hash: {str(e)}")
            return ""
            
    def _check_duplicate_document(self, file_path: str, file_hash: str) -> Optional[int]:
        """Check if document already exists by path or hash"""
        if not file_hash:
            return None
            
        try:
            conn = self.conn_pool.get_connection()
            cursor = conn.cursor()
            
            # First check by file path
            cursor.execute(
                "SELECT id FROM documents WHERE file_path = ? LIMIT 1",
                (file_path,)
            )
            result = cursor.fetchone()
            if result:
                return result[0]
                
            # Then check by hash if provided
            cursor.execute(
                "SELECT id FROM documents WHERE file_hash = ? LIMIT 1",
                (file_hash,)
            )
            result = cursor.fetchone()
            if result:
                return result[0]
                
            return None
        except Exception as e:
            logger.error(f"Error checking for duplicate document: {str(e)}")
            return None

    def add_page_detections(self, doc_id: int, page_number: int, detections: List[Dict], page_text: str = None):
        """Save detections for a specific page"""
        try:
            conn = self.conn_pool.get_connection()
            cursor = conn.cursor()
            
            # Check if detections already exist for this page
            cursor.execute(
                "SELECT id FROM page_detections WHERE document_id = ? AND page_number = ?",
                (doc_id, page_number)
            )
            
            existing = cursor.fetchone()
            if existing:
                cursor.execute(
                    "UPDATE page_detections SET detection_data = ?, page_text = ? WHERE id = ?",
                    (json.dumps(detections), page_text, existing[0])
                )
            else:
                cursor.execute('''
                    INSERT INTO page_detections (document_id, page_number, detection_data, page_text)
                    VALUES (?, ?, ?, ?)
                ''', (doc_id, page_number, json.dumps(detections), page_text))
                
            conn.commit()
        except Exception as e:
            logger.error(f"Error adding page detections: {str(e)}")
            raise

    def get_document_detections(self, doc_id: int, page_number: int = None) -> Union[List[Dict], Dict[int, List[Dict]]]:
        """Get detections for a document page or all pages"""
        try:
            conn = self.conn_pool.get_connection()
            cursor = conn.cursor()
            
            if page_number is not None:
                cursor.execute('''
                    SELECT detection_data
                    FROM page_detections
                    WHERE document_id = ? AND page_number = ?
                ''', (doc_id, page_number))
                
                result = cursor.fetchone()
                if result and result[0]:
                    try:
                        # Set higher recursion limit temporarily
                        current_limit = sys.getrecursionlimit()
                        sys.setrecursionlimit(10000)
                        decoded = json.loads(result[0])
                        sys.setrecursionlimit(current_limit)
                        return decoded
                    except json.JSONDecodeError:
                        return []
                return []
            else:
                cursor.execute('''
                    SELECT page_number, detection_data
                    FROM page_detections
                    WHERE document_id = ?
                    ORDER BY page_number
                ''', (doc_id,))
                
                results = {}
                for page_num, detection_data in cursor.fetchall():
                    if detection_data:
                        try:
                            results[page_num] = json.loads(detection_data)
                        except json.JSONDecodeError:
                            results[page_num] = []
                return results
                
        except Exception as e:
            print(f"Error getting document detections: {str(e)}")  # Use print instead of logger
            return [] if page_number is not None else {}

    def delete_document(self, doc_id: int, keep_file: bool = False):
        """Delete document and all related data"""
        try:
            conn = self.conn_pool.get_connection()
            with conn:
                cursor = conn.cursor()
                
                # Backup file path before deleting
                cursor.execute('SELECT file_path FROM documents WHERE id = ?', (doc_id,))
                file_path_result = cursor.fetchone()
                file_path = file_path_result[0] if file_path_result else None
                
                # Create backup if file exists
                if file_path and os.path.exists(file_path):
                    self.create_backup(doc_id, file_path, reason="Pre-deletion backup")
                
                # Delete related records in correct order
                # 1. Delete page detections
                cursor.execute('DELETE FROM page_detections WHERE document_id = ?', (doc_id,))
                
                # 2. Delete document tags
                cursor.execute('DELETE FROM document_tags WHERE document_id = ?', (doc_id,))
                
                # 3. Delete document versions
                cursor.execute('DELETE FROM document_versions WHERE document_id = ?', (doc_id,))
                
                # 4. Delete from document backups
                cursor.execute('DELETE FROM document_backups WHERE document_id = ?', (doc_id,))
                
                # 5. Finally delete the document
                cursor.execute('DELETE FROM documents WHERE id = ?', (doc_id,))
                
                # Commit the transaction
                conn.commit()
                
                # Delete file if it exists and not keeping
                if not keep_file and file_path and os.path.exists(file_path):
                    try:
                        os.remove(file_path)
                        logger.info(f"Deleted file: {file_path}")
                    except Exception as e:
                        logger.warning(f"Cannot delete file: {file_path}. Error: {str(e)}")
                        
                return True
                
        except Exception as e:
            logger.error(f"Error deleting document: {str(e)}")
            if conn:
                conn.rollback()
            raise

    def _parse_vietnamese_date(self, date_string):
        """
        Parse Vietnamese date string to standard format
        """
        if not isinstance(date_string, str):
            return date_string
            
        try:
            # Chuẩn hóa chuỗi đầu vào
            date_string = date_string.lower().strip()
            
            # Xử lý format chuẩn từ database (yyyy-mm-dd)
            if '-' in date_string and len(date_string.split('-')[0]) == 4:
                return date_string
                
            # Xử lý format "ngày dd tháng mm năm yyyy"
            if "ngày" in date_string and "tháng" in date_string and "năm" in date_string:
                date_string = date_string.replace("ngày", "").replace("tháng", "").replace("năm", "")
                parts = [x.strip() for x in date_string.split() if x.strip()]
                if len(parts) >= 3:
                    day = int(parts[0])
                    month = int(parts[1])
                    year = int(parts[2])
                    return f"{day:02d}/{month:02d}/{year:04d}"
                    
            # Xử lý format dd/mm/yyyy
            if "/" in date_string:
                parts = date_string.split("/")
                if len(parts) == 3:
                    day = int(parts[0])
                    month = int(parts[1])
                    year = int(parts[2])
                    return f"{day:02d}/{month:02d}/{year:04d}"
                    
            return date_string
            
        except Exception as e:
            logger.error(f"Error parsing date string '{date_string}': {str(e)}")
            return date_string

    def export_to_excel(self, output_path: str, filter_criteria: Dict = None):
        """Export database to Excel with optimized formatting"""
        try:
            # Kiểm tra thư viện
            try:
                import pandas as pd
                from openpyxl import styles
                from openpyxl.utils import get_column_letter
            except ImportError:
                raise ImportError("Thư viện 'openpyxl' chưa được cài đặt.")
                    
            conn = self.conn_pool.get_connection()
            # Base query với các cột được sắp xếp hợp lý
            query = '''
                SELECT 
                    d.id as "ID",
                    d.file_name as "Tên File",
                    d.created_at as "Ngày Tạo",
                    v.cqbh_tren as "CQBH Trên",
                    v.cqbh_duoi as "CQBH Dưới",
                    v.so_ki_hieu as "Số Ký Hiệu",
                    v.loai_vb as "Loại Văn Bản",
                    v.nd_chinh as "Nội Dung Chính",
                    v.ngay_bh as "Ngày Ban Hành",
                    v.noi_nhan as "Nơi Nhận",
                    v.chuc_vu as "Chức Vụ",
                    v.chu_ky as "Chữ Ký",
                    v.do_khan as "Độ Khẩn"
                FROM documents d
                LEFT JOIN document_versions v ON d.id = v.document_id
                WHERE v.version_number = (
                    SELECT MAX(version_number)
                    FROM document_versions
                    WHERE document_id = d.id
                )
            '''

            # Xử lý filter criteria nếu có
            params = []
            if filter_criteria:
                conditions = []
                for field, value in filter_criteria.items():
                    if value:
                        if field == 'id':
                            conditions.append("d.id = ?")
                            params.append(value)
                        elif field == 'file_name':
                            conditions.append("d.file_name LIKE ?")
                            params.append(f'%{value}%')
                        elif field == 'date_from':
                            conditions.append("d.created_at >= ?")
                            params.append(value)
                        elif field == 'date_to':
                            conditions.append("d.created_at <= ?")
                            params.append(value)
                        elif field in ['cqbh_tren', 'cqbh_duoi', 'so_ki_hieu', 'loai_vb',
                                    'do_khan', 'ngay_bh', 'chuc_vu']:
                            conditions.append(f"v.{field} LIKE ?")
                            params.append(f'%{value}%')
                        elif field == 'nd_chinh':
                            conditions.append("v.nd_chinh LIKE ?")
                            params.append(f'%{value}%')
                
                if conditions:
                    query += " AND " + " AND ".join(conditions)
            
            query += " ORDER BY d.created_at DESC"
            
            # Thực thi query
            if params:
                df = pd.read_sql_query(query, conn, params=params)
            else:
                df = pd.read_sql_query(query, conn)
            
            # Xử lý datetime columns
            datetime_columns = ['Ngày Tạo', 'Ngày Ban Hành']
            for col in datetime_columns:
                if col in df.columns:
                    # Áp dụng hàm parse cho từng giá trị trong cột
                    df[col] = df[col].apply(self._parse_vietnamese_date)

            # Export to Excel với formatting tối ưu
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Documents')
                
                workbook = writer.book
                worksheet = writer.sheets['Documents']
                
                # Định nghĩa styles
                header_style = styles.NamedStyle(name='header_style')
                header_style.font = styles.Font(bold=True, size=11)
                header_style.fill = styles.PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
                header_style.alignment = styles.Alignment(horizontal='center', vertical='center', wrap_text=True)
                header_style.border = styles.Border(
                    left=styles.Side(style='thin'),
                    right=styles.Side(style='thin'),
                    top=styles.Side(style='thin'),
                    bottom=styles.Side(style='thin')
                )

                # Style cho dữ liệu
                data_style = styles.NamedStyle(name='data_style')
                data_style.font = styles.Font(size=10)
                data_style.alignment = styles.Alignment(vertical='center', wrap_text=True)
                data_style.border = styles.Border(
                    left=styles.Side(style='thin'),
                    right=styles.Side(style='thin'),
                    top=styles.Side(style='thin'),
                    bottom=styles.Side(style='thin')
                )

                # Cấu hình độ rộng và style cho từng cột
                column_widths = {
                    'ID': 8,
                    'Tên File': 25,
                    'Ngày Tạo': 12,
                    'CQBH Trên': 25,
                    'CQBH Dưới': 25,
                    'Số Ký Hiệu': 20,
                    'Loại Văn Bản': 15,
                    'Nội Dung Chính': 40,
                    'Ngày Ban Hành': 15,
                    'Nơi Nhận': 50,
                    'Chức Vụ': 20,
                    'Chữ Ký': 20,
                    'Độ Khẩn': 12
                }

                # Áp dụng style và độ rộng cho các cột
                for idx, column in enumerate(df.columns, 1):
                    col_letter = get_column_letter(idx)
                    
                    # Áp dụng style cho header
                    cell = worksheet.cell(row=1, column=idx)
                    cell.style = header_style
                    
                    # Set độ rộng cột
                    width = column_widths.get(column, 15)  # Default width là 15 nếu không được định nghĩa
                    worksheet.column_dimensions[col_letter].width = width
                    
                    # Áp dụng style cho tất cả cells trong cột
                    for row in range(2, worksheet.max_row + 1):
                        cell = worksheet.cell(row=row, column=idx)
                        cell.style = data_style
                        
                        # Căn giữa cho một số cột cụ thể
                        if column in ['ID', 'Ngày Tạo', 'Ngày Ban Hành', 'Độ Khẩn']:
                            cell.alignment = styles.Alignment(horizontal='center', vertical='center', wrap_text=True)
                        
                        # Căn trái cho các cột còn lại
                        else:
                            cell.alignment = styles.Alignment(horizontal='left', vertical='center', wrap_text=True)

                # Set độ cao cho header
                worksheet.row_dimensions[1].height = 35

                # Set độ cao cho data rows
                for row in range(2, worksheet.max_row + 1):
                    # Tính toán độ cao dựa trên nội dung
                    max_length = 0
                    for cell in worksheet[row]:
                        if cell.value:
                            lines = str(cell.value).count('\n') + 1
                            max_length = max(max_length, lines)
                    
                    # Set độ cao tối thiểu 20, và thêm 15 cho mỗi dòng nếu có nhiều dòng
                    row_height = max(20, min(15 * max_length, 100))  # giới hạn độ cao tối đa là 100
                    worksheet.row_dimensions[row].height = row_height

                # Freeze panes
                worksheet.freeze_panes = 'A2'
                
                # Auto-filter
                worksheet.auto_filter.ref = worksheet.dimensions

            logger.info(f"Successfully exported data to {output_path}")
            return True
                
        except Exception as e:
            logger.error(f"Error exporting to Excel: {str(e)}")
            raise

    def get_all_documents(self, filter_criteria: Dict = None, sort_by: str = 'created_at', sort_desc: bool = True):
        """Get all documents with optional filtering and sorting"""
        try:
            conn = self.conn_pool.get_connection()
            query = '''
                SELECT 
                    d.id, 
                    d.file_path,
                    d.file_name,
                    d.created_at,
                    d.page_count,
                    (SELECT COUNT(v.id) FROM document_versions v WHERE v.document_id = d.id) as version_count,
                    (SELECT cqbh_tren FROM document_versions 
                     WHERE document_id = d.id 
                     ORDER BY version_number DESC LIMIT 1) as cqbh_tren,
                    (SELECT cqbh_duoi FROM document_versions 
                     WHERE document_id = d.id 
                     ORDER BY version_number DESC LIMIT 1) as cqbh_duoi,
                    (SELECT so_ki_hieu FROM document_versions 
                     WHERE document_id = d.id 
                     ORDER BY version_number DESC LIMIT 1) as so_ki_hieu,
                    (SELECT loai_vb FROM document_versions 
                     WHERE document_id = d.id 
                     ORDER BY version_number DESC LIMIT 1) as loai_vb,
                    (SELECT do_khan FROM document_versions 
                     WHERE document_id = d.id 
                     ORDER BY version_number DESC LIMIT 1) as do_khan,
                    (SELECT modified_by FROM document_versions 
                     WHERE document_id = d.id 
                     ORDER BY version_number DESC LIMIT 1) as modified_by,
                    (SELECT created_at FROM document_versions 
                     WHERE document_id = d.id 
                     ORDER BY version_number DESC LIMIT 1) as last_modified
                FROM documents d
            '''
            
            # Add filter conditions if provided
            params = []
            if filter_criteria:
                conditions = []
                for field, value in filter_criteria.items():
                    if value:
                        if field == 'file_name':
                            conditions.append(f"d.file_name LIKE ?")
                            params.append(f'%{value}%')
                        elif field == 'text':
                            text_condition = '''
                                EXISTS (
                                    SELECT 1 FROM document_versions v 
                                    WHERE v.document_id = d.id 
                                    AND (
                                        v.cqbh_tren LIKE ? OR
                                        v.cqbh_duoi LIKE ? OR
                                        v.so_ki_hieu LIKE ? OR
                                        v.loai_vb LIKE ? OR
                                        v.nd_chinh LIKE ? OR
                                        v.ngay_bh LIKE ? OR
                                        v.noi_nhan LIKE ? OR 
                                        v.chuc_vu LIKE ? OR
                                        v.chu_ky LIKE ? OR
                                        v.do_khan LIKE ?
                                    )
                                )
                            '''
                            conditions.append(text_condition)
                            params.extend([f'%{value}%'] * 10)  # One for each field
                        elif field == 'date_from':
                            conditions.append(f"d.created_at >= ?")
                            params.append(value)
                        elif field == 'date_to':
                            conditions.append(f"d.created_at <= ?")
                            params.append(value)
                        elif field == 'do_khan':
                            conditions.append('''
                                EXISTS (
                                    SELECT 1 FROM document_versions v
                                    WHERE v.document_id = d.id
                                    AND v.version_number = (
                                        SELECT MAX(version_number) FROM document_versions
                                        WHERE document_id = d.id
                                    )
                                    AND v.do_khan = ?
                                )
                            ''')
                            params.append(value)
                
                if conditions:
                    query += " WHERE " + " AND ".join(conditions)
            
            # Add sorting
            sort_column = sort_by
            if sort_by == 'so_ki_hieu':
                sort_column = "(SELECT so_ki_hieu FROM document_versions WHERE document_id = d.id ORDER BY version_number DESC LIMIT 1)"
            elif sort_by == 'do_khan':
                sort_column = "(SELECT do_khan FROM document_versions WHERE document_id = d.id ORDER BY version_number DESC LIMIT 1)"
                
            query += f" ORDER BY {sort_column} {'DESC' if sort_desc else 'ASC'}"
            
            # Execute query
            cursor = conn.cursor()
            if params:
                cursor.execute(query, params)
            else:
                cursor.execute(query)
                
            return cursor.fetchall()
                
        except Exception as e:
            logger.error(f"Error fetching documents: {str(e)}")
            return []

    def get_document_version(self, doc_id, version_number):
        """Get a specific version of a document"""
        try:
            conn = self.conn_pool.get_connection()
            cursor = conn.cursor()
            cursor.execute('''
                SELECT * FROM document_versions
                WHERE document_id = ? AND version_number = ?
            ''', (doc_id, version_number))
            return cursor.fetchone()
        except Exception as e:
            logger.error(f"Error getting document version: {str(e)}")
            return None

    def get_latest_version(self, doc_id):
        """Get the most recent version of a document"""
        try:
            conn = self.conn_pool.get_connection()
            cursor = conn.cursor()
            cursor.execute('''
                SELECT * FROM document_versions
                WHERE document_id = ?
                ORDER BY version_number DESC
                LIMIT 1
            ''', (doc_id,))
            return cursor.fetchone()
        except Exception as e:
            logger.error(f"Error getting latest version: {str(e)}")
            return None
        
    def get_document_versions(self, doc_id):
        """Get all versions of a document"""
        try:
            conn = self.conn_pool.get_connection()
            cursor = conn.cursor()
            cursor.execute('''
                SELECT * FROM document_versions
                WHERE document_id = ?
                ORDER BY version_number DESC
            ''', (doc_id,))
            return cursor.fetchall()
        except Exception as e:
            logger.error(f"Error getting document versions: {str(e)}")
            return []
        
    def get_document_info(self, doc_id):
        """Get basic document information"""
        try:
            conn = self.conn_pool.get_connection()
            cursor = conn.cursor()
            cursor.execute('''
                SELECT * FROM documents
                WHERE id = ?
            ''', (doc_id,))
            return cursor.fetchone()
        except Exception as e:
            logger.error(f"Error getting document info: {str(e)}")
            return None

    def create_new_version(self, doc_id: int, updates: Dict[str, str], modified_by: str = "User"):
        """Create a new version of a document"""
        try:
            # Đầu tiên kiểm tra xem document có tồn tại không
            conn = self.conn_pool.get_connection()
            cursor = conn.cursor()
            cursor.execute('SELECT id FROM documents WHERE id = ?', (doc_id,))
            if not cursor.fetchone():
                raise ValueError(f"Document with ID {doc_id} does not exist")

            with conn:
                # Get current version number
                cursor.execute('''
                    SELECT MAX(version_number)
                    FROM document_versions
                    WHERE document_id = ?
                ''', (doc_id,))
                
                current_version = cursor.fetchone()[0] or 0
                new_version = current_version + 1

                # Chuẩn bị dữ liệu cho version mới
                version_data = (
                    doc_id,
                    new_version,
                    updates.get('cqbh_tren', ''),
                    updates.get('cqbh_duoi', ''),
                    updates.get('so_ki_hieu', ''),
                    updates.get('loai_vb', ''),
                    updates.get('nd_chinh', ''),
                    updates.get('ngay_bh', ''),
                    updates.get('noi_nhan', ''),
                    updates.get('chuc_vu', ''),
                    updates.get('chu_ky', ''),
                    updates.get('do_khan', 'Không'),
                    modified_by
                )
                
                # Insert new version with error handling
                try:
                    cursor.execute('''
                        INSERT INTO document_versions (
                            document_id, version_number, cqbh_tren, cqbh_duoi,
                            so_ki_hieu, loai_vb, nd_chinh, ngay_bh,
                            noi_nhan, chuc_vu, chu_ky, do_khan, modified_by
                        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    ''', version_data)

                    # Update last_modified in documents table
                    cursor.execute('''
                        UPDATE documents
                        SET last_modified = CURRENT_TIMESTAMP
                        WHERE id = ?
                    ''', (doc_id,))

                    conn.commit()
                    
                    # Add suggestions for fields
                    for field in ['so_ki_hieu', 'loai_vb', 'chuc_vu', 'cqbh_tren', 'cqbh_duoi', 'do_khan']:
                        if field in updates and updates[field].strip():
                            self.add_suggestion(field, updates[field].strip())
                    
                    return new_version

                except sqlite3.IntegrityError as e:
                    conn.rollback()
                    logger.error(f"Database integrity error: {str(e)}")
                    raise
                except Exception as e:
                    conn.rollback()
                    logger.error(f"Error creating new version: {str(e)}")
                    raise
                    
        except Exception as e:
            logger.error(f"Error creating new version: {str(e)}")
            raise

    def search_documents(self, query: str, search_type: str = "all"):
        """Search documents based on query and search type"""
        if not query:
            return self.get_all_documents()
            
        filter_criteria = {}
        if search_type == "file_name":
            filter_criteria['file_name'] = query
        elif search_type == "content":
            filter_criteria['text'] = query
        elif search_type == "so_ki_hieu":
            filter_criteria['text'] = query  # We'll filter in memory
        elif search_type == "do_khan":
            filter_criteria['do_khan'] = query
        else:  # "all"
            filter_criteria['text'] = query
            
        documents = self.get_all_documents(filter_criteria)
        
        # Additional filtering for so_ki_hieu if needed
        if search_type == "so_ki_hieu":
            documents = [doc for doc in documents if query.lower() in (doc[8] or '').lower()]
            
        return documents
        
    def verify_file_paths(self):
        """Check and identify missing file paths"""
        try:
            conn = self.conn_pool.get_connection()
            cursor = conn.cursor()
            cursor.execute('SELECT id, file_path FROM documents')
            documents = cursor.fetchall()
            
            missing_files = []
            for doc_id, file_path in documents:
                if not os.path.exists(file_path):
                    missing_files.append((doc_id, file_path))
                    
            return missing_files
        except Exception as e:
            logger.error(f"Error verifying file paths: {str(e)}")
            return []

    def update_file_path(self, doc_id: int, new_path: str):
        """Update file path for a document"""
        try:
            if not os.path.exists(new_path):
                raise FileNotFoundError(f"New file path does not exist: {new_path}")
                
            # Calculate new hash and size
            file_hash = self._calculate_file_hash(new_path)
            file_size = os.path.getsize(new_path)
            
            # Update page count if PDF
            page_count = None
            if Path(new_path).suffix.lower() == '.pdf':
                try:
                    with fitz.open(new_path) as doc:
                        page_count = len(doc)
                except:
                    pass
                    
            conn = self.conn_pool.get_connection()
            with conn:
                cursor = conn.cursor()
                if page_count is not None:
                    cursor.execute(
                        'UPDATE documents SET file_path = ?, file_hash = ?, file_size = ?, page_count = ? WHERE id = ?',
                        (new_path, file_hash, file_size, page_count, doc_id)
                    )
                else:
                    cursor.execute(
                        'UPDATE documents SET file_path = ?, file_hash = ?, file_size = ? WHERE id = ?',
                        (new_path, file_hash, file_size, doc_id)
                    )
                
                return cursor.rowcount > 0
        except Exception as e:
            logger.error(f"Error updating file path: {str(e)}")
            raise

    def add_tag(self, doc_id: int, tag_name: str):
        """Add a tag to a document"""
        try:
            conn = self.conn_pool.get_connection()
            with conn:
                cursor = conn.cursor()
                cursor.execute(
                    'INSERT OR IGNORE INTO document_tags (document_id, tag_name) VALUES (?, ?)',
                    (doc_id, tag_name)
                )
                return cursor.rowcount > 0
        except Exception as e:
            logger.error(f"Error adding tag: {str(e)}")
            return False
            
    def remove_tag(self, doc_id: int, tag_name: str):
        """Remove a tag from a document"""
        try:
            conn = self.conn_pool.get_connection()
            with conn:
                cursor = conn.cursor()
                cursor.execute(
                    'DELETE FROM document_tags WHERE document_id = ? AND tag_name = ?',
                    (doc_id, tag_name)
                )
                return cursor.rowcount > 0
        except Exception as e:
            logger.error(f"Error removing tag: {str(e)}")
            return False
            
    def get_document_tags(self, doc_id: int):
        """Get all tags for a document"""
        try:
            conn = self.conn_pool.get_connection()
            cursor = conn.cursor()
            cursor.execute(
                'SELECT tag_name FROM document_tags WHERE document_id = ? ORDER BY tag_name',
                (doc_id,)
            )
            return [row[0] for row in cursor.fetchall()]
        except Exception as e:
            logger.error(f"Error getting document tags: {str(e)}")
            return []
            
    def get_all_tags(self):
        """Get all unique tags in the system"""
        try:
            conn = self.conn_pool.get_connection()
            cursor = conn.cursor()
            cursor.execute(
                'SELECT DISTINCT tag_name FROM document_tags ORDER BY tag_name'
            )
            return [row[0] for row in cursor.fetchall()]
        except Exception as e:
            logger.error(f"Error getting all tags: {str(e)}")
            return []

    def get_documents_by_tag(self, tag_name: str):
        """Get all documents with a specific tag"""
        try:
            conn = self.conn_pool.get_connection()
            cursor = conn.cursor()
            query = '''
                SELECT 
                    d.id, 
                    d.file_path,
                    d.file_name,
                    d.created_at,
                    d.page_count,
                    (SELECT COUNT(v.id) FROM document_versions v WHERE v.document_id = d.id) as version_count,
                    (SELECT cqbh_tren FROM document_versions 
                     WHERE document_id = d.id 
                     ORDER BY version_number DESC LIMIT 1) as cqbh_tren,
                    (SELECT cqbh_duoi FROM document_versions 
                     WHERE document_id = d.id 
                     ORDER BY version_number DESC LIMIT 1) as cqbh_duoi,
                    (SELECT so_ki_hieu FROM document_versions 
                     WHERE document_id = d.id 
                     ORDER BY version_number DESC LIMIT 1) as so_ki_hieu,
                    (SELECT loai_vb FROM document_versions 
                     WHERE document_id = d.id 
                     ORDER BY version_number DESC LIMIT 1) as loai_vb,
                    (SELECT do_khan FROM document_versions 
                     WHERE document_id = d.id 
                     ORDER BY version_number DESC LIMIT 1) as do_khan,
                    (SELECT modified_by FROM document_versions 
                     WHERE document_id = d.id 
                     ORDER BY version_number DESC LIMIT 1) as modified_by,
                    (SELECT created_at FROM document_versions 
                     WHERE document_id = d.id 
                     ORDER BY version_number DESC LIMIT 1) as last_modified
                FROM documents d
                JOIN document_tags t ON d.id = t.document_id
                WHERE t.tag_name = ?
                ORDER BY d.created_at DESC
            '''
            cursor.execute(query, (tag_name,))
            return cursor.fetchall()
        except Exception as e:
            logger.error(f"Error getting documents by tag: {str(e)}")
            return []
            
    def get_document_count(self):
        """Get total number of documents"""
        try:
            conn = self.conn_pool.get_connection()
            cursor = conn.cursor()
            cursor.execute('SELECT COUNT(*) FROM documents')
            return cursor.fetchone()[0]
        except Exception as e:
            logger.error(f"Error getting document count: {str(e)}")
            return 0
            
    def close(self):
        """Close all database connections"""
        self.conn_pool.close_all()

# #############################
# #
# #############################
# class StatisticsDialog(QDialog):
#     """Dialog hiển thị thống kê"""

#     def __init__(self, db, parent=None):
#         super().__init__(parent)
#         self.db = db
#         self.setup_ui()
#         self.load_statistics()

#     def setup_ui(self):
#         self.setWindowTitle("Thống kê hệ thống")
#         self.setMinimumWidth(800)
#         self.setMinimumHeight(600)

#         layout = QVBoxLayout(self)

#         # Tạo tab widget
#         tab_widget = QTabWidget()
        
#         # Tab tổng quan
#         overview_tab = QWidget()
#         overview_layout = QVBoxLayout(overview_tab)
        
#         # Box thông tin tổng quan
#         overview_group = QGroupBox("Thông tin tổng quan")
#         overview_group_layout = QVBoxLayout()
#         self.total_label = QLabel()
#         self.total_label.setStyleSheet("font-size: 14px; font-weight: bold;")
#         overview_group_layout.addWidget(self.total_label)
#         overview_group.setLayout(overview_group_layout)
#         overview_layout.addWidget(overview_group)
        
#         # Thống kê theo loại
#         type_group = QGroupBox("Thống kê theo loại văn bản")
#         type_layout = QVBoxLayout()
#         self.type_table = QTableWidget()
#         self.type_table.setColumnCount(2)
#         self.type_table.setHorizontalHeaderLabels(["Loại văn bản", "Số lượng"])
#         self.type_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
#         self.type_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeToContents)
#         type_layout.addWidget(self.type_table)
#         type_group.setLayout(type_layout)
#         overview_layout.addWidget(type_group)
        
#         # Thống kê độ khẩn
#         urgency_group = QGroupBox("Thống kê theo độ khẩn")
#         urgency_layout = QVBoxLayout()
#         self.urgency_table = QTableWidget()
#         self.urgency_table.setColumnCount(2)
#         self.urgency_table.setHorizontalHeaderLabels(["Độ khẩn", "Số lượng"])
#         self.urgency_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
#         self.urgency_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeToContents)
#         urgency_layout.addWidget(self.urgency_table)
#         urgency_group.setLayout(urgency_layout)
#         overview_layout.addWidget(urgency_group)
        
#         tab_widget.addTab(overview_tab, "Tổng quan")
        
#         # Tab văn bản mới nhất
#         recent_tab = QWidget()
#         recent_layout = QVBoxLayout(recent_tab)
#         self.recent_table = QTableWidget()
#         self.recent_table.setColumnCount(4)
#         self.recent_table.setHorizontalHeaderLabels(["ID", "Tên file", "Ngày tạo", "Số ký hiệu"])
#         self.recent_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
#         recent_layout.addWidget(self.recent_table)
#         tab_widget.addTab(recent_tab, "Văn bản mới nhất")
        
#         # Tab thống kê theo thời gian
#         time_tab = QWidget()
#         time_layout = QVBoxLayout(time_tab)
#         self.time_table = QTableWidget()
#         self.time_table.setColumnCount(2)
#         self.time_table.setHorizontalHeaderLabels(["Tháng", "Số lượng"])
#         self.time_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
#         time_layout.addWidget(self.time_table)
#         tab_widget.addTab(time_tab, "Thống kê theo thời gian")
        
#         layout.addWidget(tab_widget)
        
#         # Buttons
#         button_box = QHBoxLayout()
#         close_btn = QPushButton("Đóng")
#         close_btn.clicked.connect(self.close)
#         export_btn = QPushButton("Xuất báo cáo")
#         export_btn.clicked.connect(self.export_statistics)
#         button_box.addStretch()
#         button_box.addWidget(export_btn)
#         button_box.addWidget(close_btn)
#         layout.addLayout(button_box)

#     def load_statistics(self):
#         """Load thống kê từ database"""
#         stats = self.db.get_statistics()
#         if not stats:
#             return

#         # Cập nhật tổng quan
#         self.total_label.setText(f"Tổng số văn bản: {stats['total_documents']}")

#         # Cập nhật bảng loại văn bản
#         self.type_table.setRowCount(len(stats['by_type']))
#         for i, (type_name, count) in enumerate(stats['by_type']):
#             self.type_table.setItem(i, 0, QTableWidgetItem(type_name or "Không xác định"))
#             self.type_table.setItem(i, 1, QTableWidgetItem(str(count)))

#         # Cập nhật bảng độ khẩn
#         self.urgency_table.setRowCount(len(stats['by_urgency']))
#         for i, (urgency, count) in enumerate(stats['by_urgency']):
#             item = QTableWidgetItem(urgency or "Không")
#             if urgency == "Độ Mật":
#                 item.setBackground(QColor(255, 200, 200))
#             elif urgency == "Hỏa Tốc":
#                 item.setBackground(QColor(255, 140, 0))
#                 item.setForeground(QColor(255, 255, 255))
#             self.urgency_table.setItem(i, 0, item)
#             self.urgency_table.setItem(i, 1, QTableWidgetItem(str(count)))

#         # Cập nhật văn bản mới nhất
#         self.recent_table.setRowCount(len(stats['recent_docs']))
#         for i, (doc_id, filename, created_at, so_kh) in enumerate(stats['recent_docs']):
#             self.recent_table.setItem(i, 0, QTableWidgetItem(str(doc_id)))
#             self.recent_table.setItem(i, 1, QTableWidgetItem(filename))
#             self.recent_table.setItem(i, 2, QTableWidgetItem(str(created_at)))
#             self.recent_table.setItem(i, 3, QTableWidgetItem(so_kh))

#         # Cập nhật thống kê theo thời gian
#         self.time_table.setRowCount(len(stats['by_month']))
#         for i, (month, count) in enumerate(stats['by_month']):
#             self.time_table.setItem(i, 0, QTableWidgetItem(month))
#             self.time_table.setItem(i, 1, QTableWidgetItem(str(count)))

#     def export_statistics(self):
#         """Xuất thống kê ra file Excel"""
#         try:
#             file_path, _ = QFileDialog.getSaveFileName(
#                 self, "Xuất báo cáo", "", "Excel Files (*.xlsx)"
#             )
#             if not file_path:
#                 return

#             if not file_path.endswith('.xlsx'):
#                 file_path += '.xlsx'

#             stats = self.db.get_statistics()
#             if not stats:
#                 return

#             with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
#                 # Tổng quan
#                 overview_data = {
#                     'Thông tin': ['Tổng số văn bản'],
#                     'Số lượng': [stats['total_documents']]
#                 }
#                 pd.DataFrame(overview_data).to_excel(writer, sheet_name='Tổng quan', index=False)

#                 # Thống kê theo loại
#                 type_data = pd.DataFrame(stats['by_type'], columns=['Loại văn bản', 'Số lượng'])
#                 type_data.to_excel(writer, sheet_name='Theo loại', index=False)

#                 # Thống kê theo độ khẩn
#                 urgency_data = pd.DataFrame(stats['by_urgency'], columns=['Độ khẩn', 'Số lượng'])
#                 urgency_data.to_excel(writer, sheet_name='Theo độ khẩn', index=False)

#                 # Thống kê theo thời gian
#                 time_data = pd.DataFrame(stats['by_month'], columns=['Tháng', 'Số lượng'])
#                 time_data.to_excel(writer, sheet_name='Theo thời gian', index=False)

#                 # Văn bản mới nhất
#                 recent_data = pd.DataFrame(
#                     stats['recent_docs'],
#                     columns=['ID', 'Tên file', 'Ngày tạo', 'Số ký hiệu']
#                 )
#                 recent_data.to_excel(writer, sheet_name='Văn bản mới', index=False)

#             QMessageBox.information(self, "Thành công", f"Đã xuất báo cáo thống kê tới {file_path}")

#         except Exception as e:
#             QMessageBox.critical(self, "Lỗi", f"Lỗi khi xuất báo cáo: {str(e)}")

#############################
#    Document OCR Class     #
#############################
class DocumentOCR:
    """OCR engine for extracting text from documents"""
    
    def __init__(self, model_path):
        """Initialize OCR system with YOLO model and EasyOCR"""
        self.model_path = model_path
        self.confidence_threshold = DEFAULT_CONFIDENCE_THRESHOLD
        self.num_processes = max(1, mp.cpu_count() - 1)  # Leave one core free
        
        # Khởi tạo EasyOCR reader (sẽ được khởi tạo lại trong mỗi worker process)
        self.ocr_reader = None
        
        # Thư mục lưu ảnh
        self.image_save_dir = Path(OUTPUT_DIR) / 'ocr_images'
        self.image_save_dir.mkdir(exist_ok=True, parents=True)
        (self.image_save_dir / 'original').mkdir(exist_ok=True)
        (self.image_save_dir / 'processed').mkdir(exist_ok=True)
        
        # Classes for YOLO model
        self.classes = {
            'CQBH': 0, 'Chu_Ky': 1, 'Chuc_Vu': 2, 'Do_Khan': 3,
            'Loai_VB': 4, 'ND_Chinh': 5, 'Ngay_BH': 6, 'Noi_Nhan': 7, 'So_Ki_Hieu': 8
        }
        
        # Xác minh file model
        if not os.path.exists(model_path):
            logger.error(f"Model file not found: {model_path}")
            raise FileNotFoundError(f"Model file not found: {model_path}")

    @staticmethod
    def preprocess_image_for_document(image, class_id=None):
        """Tiền xử lý ảnh tối ưu cho văn bản tiếng Việt với dấu"""
        # Chuyển đổi sang array nếu là đối tượng PIL
        if isinstance(image, Image.Image):
            img_array = np.array(image)
        else:
            img_array = image
        
        # Chuyển sang ảnh xám nếu là ảnh màu
        if len(img_array.shape) == 3:
            gray = cv2.cvtColor(img_array, cv2.COLOR_RGB2GRAY)
        else:
            gray = img_array
        
        # Tăng kích thước ảnh lên 2x để cải thiện nhận dạng dấu
        height, width = gray.shape
        gray = cv2.resize(gray, (width*2, height*2), interpolation=cv2.INTER_CUBIC)
        
        # Xử lý dựa vào loại văn bản (nếu được cung cấp)
        if class_id is not None:
            # Chữ ký, Chức vụ, Nơi nhận -> cần độ tương phản cao
            if class_id in [1, 2, 7]:  
                # Loại bỏ nhiễu
                gray = cv2.fastNlMeansDenoising(gray, None, 10, 7, 21)
                # Cải thiện độ tương phản
                clahe = cv2.createCLAHE(clipLimit=3.0, tileGridSize=(8,8))
                gray = clahe.apply(gray)
                # Ngưỡng thích ứng
                binary = cv2.adaptiveThreshold(
                    gray, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, 
                    cv2.THRESH_BINARY, 11, 2
                )
            # Số ký hiệu, Ngày BH -> cần rõ nét
            elif class_id in [6, 8]:
                # Làm mờ nhẹ
                gray = cv2.GaussianBlur(gray, (3, 3), 0)
                # Tăng độ tương phản
                clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8,8))
                gray = clahe.apply(gray)
                # Tăng độ sắc nét
                kernel = np.array([[-1,-1,-1], [-1,9,-1], [-1,-1,-1]])
                gray = cv2.filter2D(gray, -1, kernel)
                # Ngưỡng hóa Otsu
                _, binary = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
            # Nội dung chính -> đảm bảo giữ dấu
            elif class_id == 5:
                # Loại bỏ nhiễu nhẹ
                gray = cv2.GaussianBlur(gray, (3, 3), 0)
                # Cải thiện độ tương phản
                clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(16,16))
                gray = clahe.apply(gray)
                # Ngưỡng hóa Otsu
                _, binary = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
            # Loại văn bản - cần rõ text
            elif class_id == 4:  
                # Sắc nét và nhiễu thấp cho văn bản
                gray = cv2.GaussianBlur(gray, (3, 3), 0)
                clahe = cv2.createCLAHE(clipLimit=2.5, tileGridSize=(8,8))
                gray = clahe.apply(gray)
                _, binary = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
            # CQBH - Cơ quan ban hành
            elif class_id == 0:  
                # Cải thiện văn bản in đậm, logo
                gray = cv2.GaussianBlur(gray, (3, 3), 0)
                clahe = cv2.createCLAHE(clipLimit=3.0, tileGridSize=(8,8))
                gray = clahe.apply(gray)
                kernel = np.array([[-1,-1,-1], [-1,9,-1], [-1,-1,-1]])
                gray = cv2.filter2D(gray, -1, kernel)
                _, binary = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
            # Độ khẩn
            elif class_id == 3:  
                # Xử lý đặc biệt cho text màu đỏ nổi bật
                gray = cv2.fastNlMeansDenoising(gray, None, 10, 7, 21)
                clahe = cv2.createCLAHE(clipLimit=4.0, tileGridSize=(4,4))
                gray = clahe.apply(gray)
                binary = cv2.adaptiveThreshold(gray, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, 
                                  cv2.THRESH_BINARY, 11, 2)
            # Mặc định với các loại còn lại
            else:
                # Loại bỏ nhiễu nhẹ
                gray = cv2.GaussianBlur(gray, (3, 3), 0)
                # Cải thiện độ tương phản
                clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8,8))
                gray = clahe.apply(gray)
                # Ngưỡng hóa Otsu
                _, binary = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
        else:
            # Tiền xử lý mặc định khi không biết loại văn bản
            # Loại bỏ nhiễu nhẹ
            gray = cv2.GaussianBlur(gray, (3, 3), 0)
            # Cải thiện độ tương phản
            clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8,8))
            gray = clahe.apply(gray)
            # Ngưỡng hóa Otsu
            _, binary = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
        
        # Thêm biên trắng xung quanh (giúp cải thiện OCR)
        binary = cv2.copyMakeBorder(binary, 10, 10, 10, 10, cv2.BORDER_CONSTANT, value=255)
        
        return binary

    @staticmethod
    def _ocr_region(image: Image, config_params=None) -> str:
        """Extract text from image region using EasyOCR"""
        try:
            # Xác định loại class từ tham số (nếu có)
            class_id = None
            if config_params and 'class_id' in config_params:
                class_id = config_params['class_id']
            
            # Chuyển đổi thành đối tượng PIL Image nếu cần
            if not isinstance(image, Image.Image):
                image = Image.fromarray(image)
            
            # Tiền xử lý ảnh
            processed_img = DocumentOCR.preprocess_image_for_document(image, class_id)
            
            # Chọn ngôn ngữ OCR cho từng loại class
            languages = ['vi']
            if class_id in [7, 8, 4]:  # Noi_Nhan, So_Ki_Hieu, Loai_VB - thêm English để nhận dạng tốt hơn các ký tự đặc biệt
                languages = ['vi', 'en']
            
            # Tắt cảnh báo GPU
            import logging
            logging.getLogger('easyocr.easyocr').setLevel(logging.ERROR)
            
            # Khởi tạo EasyOCR reader với tùy chọn tắt cảnh báo GPU
            reader = easyocr.Reader(languages, gpu=False, verbose=False)
            
            # Chuyển đổi sang numpy array
            if isinstance(processed_img, Image.Image):
                img_array = np.array(processed_img)
            else:
                img_array = processed_img
                
            # Thực hiện OCR
            results = reader.readtext(img_array)
            
            # Áp dụng xử lý theo dòng cho mọi loại class để cải thiện kết quả
            # Sắp xếp kết quả theo tọa độ y (từ trên xuống dưới)
            sorted_results = sorted(results, key=lambda x: (x[0][0][1] + x[0][2][1])/2)
            
            # Nhóm các kết quả theo dòng
            line_height = 20  # Độ cao dòng ước tính
            lines = []
            current_line = []
            
            for detection in sorted_results:
                if not current_line:
                    current_line.append(detection)
                else:
                    y_current = (detection[0][0][1] + detection[0][2][1])/2
                    y_prev = (current_line[-1][0][0][1] + current_line[-1][0][2][1])/2
                    
                    if abs(y_current - y_prev) < line_height:
                        current_line.append(detection)
                    else:
                        # Sắp xếp từ trái sang phải trong một dòng
                        current_line = sorted(current_line, key=lambda x: x[0][0][0])
                        lines.append(current_line)
                        current_line = [detection]
            
            if current_line:
                current_line = sorted(current_line, key=lambda x: x[0][0][0])
                lines.append(current_line)
            
            # Xử lý đặc biệt cho từng loại class
            if class_id == 7:  # Noi_Nhan
                # Tạo văn bản theo định dạng Nơi nhận
                text_lines = []
                for line in lines:
                    line_text = " ".join([detection[1] for detection in line])
                    text_lines.append(line_text)
                
                text = "\n".join(text_lines)
                
                # Định dạng lại để dễ đọc
                if text and not text.startswith("Nơi nhận:") and not text.startswith("Nơi nhận") and not text.startswith("-"):
                    text = "Nơi nhận:\n" + text
                    
                # Thêm dấu gạch đầu dòng nếu cần
                lines = text.split('\n')
                formatted_lines = []
                for i, line in enumerate(lines):
                    if i == 0 and (line.startswith("Nơi nhận:") or line.startswith("Nơi nhận")):
                        formatted_lines.append(line)
                    elif not line.strip().startswith("-") and not line.strip().startswith("•") and i > 0 and line.strip():
                        formatted_lines.append("- " + line.strip())
                    else:
                        formatted_lines.append(line)
                
                text = "\n".join(formatted_lines)
            elif class_id == 5:  # ND_Chinh - Nội dung chính
                # Tạo văn bản theo định dạng đoạn văn
                text_lines = []
                for line in lines:
                    line_text = " ".join([detection[1] for detection in line])
                    text_lines.append(line_text)
                
                text = "\n".join(text_lines)
            elif class_id == 0:  # CQBH - Cơ quan ban hành
                # CQBH thường có 1-2 dòng
                text_lines = []
                for line in lines:
                    line_text = " ".join([detection[1] for detection in line])
                    text_lines.append(line_text)
                
                text = "\n".join(text_lines)
            elif class_id == 6:  # Ngày BH
                # Ghép tất cả các phát hiện thành một dòng
                all_text = []
                for line in lines:
                    line_text = " ".join([detection[1] for detection in line])
                    all_text.append(line_text)
                
                text = " ".join(all_text)
                
                # Chuẩn hóa dấu ngày tháng
                text = text.replace('/', '-').replace('.', '-')
                # Sửa các số hay nhận nhầm
                text = text.replace('l', '1').replace('O', '0').replace('o', '0')
                
                # Định dạng lại ngày tháng nếu có thể
                date_pattern = r'(\d{1,2})[-./](\d{1,2})[-./](\d{2,4})'
                match = re.search(date_pattern, text)
                if match:
                    day, month, year = match.groups()
                    # Đảm bảo định dạng DD-MM-YYYY
                    if len(year) == 2:
                        year = '20' + year  # Giả sử năm hiện tại là thế kỷ 21
                    text = f"ngày {day} tháng {month} năm {year}"
            elif class_id == 8:  # Số ký hiệu
                # Ghép tất cả các phát hiện từ các dòng
                text_parts = []
                for line in lines:
                    line_text = " ".join([detection[1] for detection in line])
                    text_parts.append(line_text)
                
                text = " ".join(text_parts)
                
                # Chuẩn hóa dấu gạch ngang
                text = text.replace('—', '-').replace('–', '-').replace('_', '-')
                # Loại bỏ các ký tự không cần thiết và giữ lại những ký tự quan trọng
                text = ''.join(c for c in text if c.isalnum() or c in "/-_.,: ")
                # Sửa các số hay nhận nhầm
                text = text.replace('l', '1').replace('O', '0').replace('o', '0')
            else:
                # Các trường hợp khác - ghép các dòng lại
                all_lines = []
                for line in lines:
                    line_text = " ".join([detection[1] for detection in line])
                    all_lines.append(line_text)
                
                text = "\n".join(all_lines)
            
            # Hậu xử lý cho text tiếng Việt
            text = text.strip()
            
            return text
            
        except Exception as e:
            logger.error(f"OCR error: {str(e)}")
            return ""

    @staticmethod
    def _process_page_wrapper(args):
        """Wrapper function for multiprocessing"""
        try:
            img, page_num, model_path, confidence_threshold, classes, save_dir = args
            
            # Initialize YOLO model in worker process
            model = YOLO(model_path)
            
            # Khởi tạo EasyOCR reader cho worker process
            reader = easyocr.Reader(['vi'], gpu=False)
            
            # Lưu ảnh trang gốc nếu có thư mục lưu
            if save_dir:
                original_dir = Path(save_dir) / 'original'
                img_filename = f"page_{page_num}_original.png"
                img.save(original_dir / img_filename)
            
            # Detect regions
            predictions = model(img)[0]
            detections = predictions.boxes.data.cpu().numpy()
            results = {}
            page_detections = []
            
            for det in detections:
                conf = det[4]
                class_id = int(det[5])
                
                if conf > confidence_threshold:
                    class_name = [k for k, v in classes.items() if v == class_id][0]
                    box = det[:4].tolist()
                    
                    # Mở rộng vùng box để đảm bảo lấy được đầy đủ dấu
                    x1, y1, x2, y2 = box
                    width = x2 - x1
                    height = y2 - y1
                    
                    # Thêm padding xung quanh, đặc biệt ở phía trên để lấy được dấu
                    padding_top = int(height * 0.15)  # 15% chiều cao ở trên để lấy dấu
                    padding_side = int(width * 0.05)  # 5% chiều rộng mỗi bên
                    padding_bottom = int(height * 0.05)  # 5% chiều cao ở dưới
                    
                    # Đảm bảo tọa độ không âm và không vượt quá kích thước ảnh
                    img_width, img_height = img.size
                    x1_padded = max(0, x1 - padding_side)
                    y1_padded = max(0, y1 - padding_top)
                    x2_padded = min(img_width, x2 + padding_side)
                    y2_padded = min(img_height, y2 + padding_bottom)
                    
                    # Tạo region mới với padding
                    region = img.crop((x1_padded, y1_padded, x2_padded, y2_padded))
                    
                    # Lưu ảnh vùng cắt gốc
                    if save_dir:
                        region_filename = f"page_{page_num}_{class_name}_{conf:.2f}_original.png"
                        region.save(Path(save_dir) / 'original' / region_filename)
                    
                    # Cấu hình OCR với class_id để tối ưu hóa tiền xử lý
                    config_params = {
                        'class_id': class_id
                    }
                    
                    # Tiền xử lý ảnh với tối ưu cho loại class
                    processed_region = DocumentOCR.preprocess_image_for_document(region, class_id)
                    processed_region_img = Image.fromarray(processed_region)
                    
                    # Lưu ảnh đã xử lý
                    if save_dir:
                        processed_filename = f"page_{page_num}_{class_name}_{conf:.2f}_processed.png"
                        processed_region_img.save(Path(save_dir) / 'processed' / processed_filename)
                    
                    # Thực hiện OCR với cấu hình tùy chỉnh tùy thuộc vào loại vùng
                    custom_config = f'--oem 1 --dpi 300 '
                    
                    # Chọn PSM phù hợp với loại vùng
                    if class_id in [1, 2, 3, 6, 8]:  # Chữ ký, Chức vụ, Độ khẩn, Ngày BH, Số KH
                        custom_config += '--psm 7'  # Treat the image as a single text line
                    elif class_id == 7:  # Nơi nhận
                        custom_config += '--psm 4'  # Assume a single column of text
                    else:
                        custom_config += '--psm 6'  # Assume a single uniform block of text
                    
                    # Extract text from the region using OCR
                    text = DocumentOCR._ocr_region(processed_region_img, {
                        'lang': 'vie',
                        'class_id': class_id,
                        'config': custom_config
                    })
                    
                    # Save detection info
                    detection_info = {
                        'box': [x1_padded, y1_padded, x2_padded, y2_padded],
                        'original_box': box,
                        'confidence': float(conf),
                        'class': class_name,
                        'text': text
                    }
                    page_detections.append(detection_info)
                    
                    if text:
                        # Đặc biệt xử lý cho CQBH (có thể có 2 phần trên/dưới)
                        if class_name == 'CQBH':
                            parts = text.split('\n')
                            if len(parts) > 1:
                                results['CQBH_tren'] = parts[0].strip()
                                results['CQBH_duoi'] = ' '.join(parts[1:]).strip()
                            else:
                                results['CQBH_tren'] = text.strip()
                        # Xử lý các vùng khác
                        else:
                            results[class_name] = text.strip()
            
            return page_num, results, page_detections
            
        except Exception as e:
            logger.error(f"Error processing document: {str(e)}")
            traceback.print_exc()
            return results, []
            
    def extract_text(self, image):
        """Extract text from a single image"""
        try:
            # Khởi tạo EasyOCR reader nếu chưa có
            if not hasattr(self, 'ocr_reader') or self.ocr_reader is None:
                self.ocr_reader = easyocr.Reader(['vi'], gpu=False)
            
            # Tiền xử lý ảnh
            if not isinstance(image, Image.Image):
                if isinstance(image, np.ndarray):
                    image = Image.fromarray(image)
                else:
                    image = Image.open(image)
            
            # Chuyển sang ảnh xám và tiền xử lý
            processed_img = self.preprocess_image_for_document(image)
            
            # Chuyển sang numpy array để dùng với EasyOCR
            img_array = np.array(processed_img)
            
            # Thực hiện OCR với EasyOCR
            results = self.ocr_reader.readtext(img_array)
            
            # Tổng hợp kết quả
            lines = []
            for (bbox, text, prob) in results:
                lines.append(text)
            
            # Kết hợp văn bản từ tất cả các dòng
            text = " ".join(lines)
            
            # Hậu xử lý
            text = text.strip()
            
            return text
        except Exception as e:
            logger.error(f"Error extracting text: {str(e)}")
            return ""
            
    def process_document(self, pdf_path, progress_callback=None):
        """Process a PDF document and extract text from detected regions"""
        try:
            if not os.path.exists(pdf_path):
                logger.error(f"PDF file not found: {pdf_path}")
                raise FileNotFoundError(f"PDF file not found: {pdf_path}")
                
            doc = fitz.open(pdf_path)
            total_pages = len(doc)
            all_page_detections = []
            
            if progress_callback:
                progress_callback(0, 100, "Khởi tạo...")
                
            results = {
                'ND_Chinh': '',
                'Ngay_BH': '',
                'CQBH_tren': '',
                'CQBH_duoi': '',
                'So_Ki_Hieu': '',
                'Loai_VB': '',
                'Noi_Nhan': '',
                'Chuc_Vu': '',
                'Chu_Ky': '',
                'Do_Khan': 'Không'
            }
            
            # Tắt các cảnh báo từ logging
            easyocr_logger = logging.getLogger('easyocr')
            original_level = easyocr_logger.level
            easyocr_logger.setLevel(logging.ERROR)
            
            # Tắt cảnh báo PyTorch
            import warnings
            warnings.filterwarnings("ignore", message="'pin_memory' argument is set as true but no accelerator is found")
            
            try:
                # Chuyển đổi các trang PDF thành ảnh
                images = []
                for i in range(total_pages):
                    # Cập nhật progress: 40% đầu tiên cho việc load PDF
                    if progress_callback:
                        progress_callback(int((i/total_pages) * 40), 100, f"Đang nạp trang {i+1}/{total_pages}...")
                        
                    page = doc[i]
                    pix = page.get_pixmap(matrix=fitz.Matrix(1.5, 1.5))
                    img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
                    images.append(img)
                
                if progress_callback:
                    progress_callback(40, 100, "Đang xử lý các trang...")
                
                # Xử lý OCR
                if self.num_processes > 1 and total_pages > 1:
                    # Tạo args cho xử lý song song
                    process_args = [
                        (images[i], i, self.model_path, self.confidence_threshold, self.classes, 
                         str(self.image_save_dir) if self.image_save_dir else None)
                        for i in range(total_pages)
                    ]
                    
                    with mp.Pool(processes=self.num_processes) as pool:
                        # Xử lý các trang song song với imap để cập nhật progress
                        process_results = []
                        for i, result in enumerate(pool.imap(self._process_page_wrapper, process_args)):
                            # Cập nhật progress: 40% - 90% cho việc OCR
                            if progress_callback:
                                progress = 40 + int((i+1)/total_pages * 50)
                                progress_callback(progress, 100, 
                                                f"Đang OCR trang {i+1}/{total_pages}...")
                            
                            process_results.append(result)
                        
                        # Xử lý kết quả từ các trang
                        for page_num, page_results, page_detections in process_results:
                            # Ghép kết quả
                            for key, value in page_results.items():
                                if key == 'CQBH_tren' and not results['CQBH_tren']:
                                    results['CQBH_tren'] = value
                                elif key == 'CQBH_duoi' and not results['CQBH_duoi']:
                                    results['CQBH_duoi'] = value
                                elif key in results and not results[key]:
                                    results[key] = value
                            
                            all_page_detections.append((page_num, page_detections))
                else:
                    # Xử lý tuần tự
                    for i, img in enumerate(images):
                        # Cập nhật progress: 40% - 90% cho việc OCR
                        if progress_callback:
                            progress = 40 + int((i+1)/total_pages * 50)
                            progress_callback(progress, 100, 
                                            f"Đang OCR trang {i+1}/{total_pages}...")
                        
                        page_num, page_results, page_detections = self._process_page_wrapper(
                            (img, i, self.model_path, self.confidence_threshold, self.classes, 
                             str(self.image_save_dir) if self.image_save_dir else None)
                        )
                        
                        # Ghép kết quả
                        for key, value in page_results.items():
                            if key == 'CQBH_tren' and not results['CQBH_tren']:
                                results['CQBH_tren'] = value
                            elif key == 'CQBH_duoi' and not results['CQBH_duoi']:
                                results['CQBH_duoi'] = value
                            elif key in results and not results[key]:
                                results[key] = value
                        
                        all_page_detections.append((page_num, page_detections))
                
                # Cập nhật progress phần cuối
                if progress_callback:
                    progress_callback(95, 100, "Đang hoàn thiện kết quả...")
            
            finally:
                # Khôi phục mức độ logging ban đầu
                easyocr_logger.setLevel(original_level)
            
            # Cập nhật progress khi hoàn thành
            if progress_callback:
                progress_callback(100, 100, "Hoàn thành!")
                
            return results, all_page_detections
            
        except Exception as e:
            logger.error(f"Error processing document: {str(e)}")
            traceback.print_exc()
            return {}, []

#############################
#    OCR Result Editor      #
#############################
class OCRResultEditor(QWidget):
    """Editor for OCR results with autosuggestions"""
    
    ocr_updated = pyqtSignal(int, dict)
    field_changed = pyqtSignal(str, str)  # field_name, new_value

    def __init__(self, db, parent=None):  # Thêm tham số db
        super().__init__(parent)
        self.db = db  # Lưu database instance
        self.doc_id = None
        self.full_cqbh_text = ""
        self.current_doc_data = None 
        self.is_loading = False
        self.setup_ui()

    def setup_ui(self):
        # Main layout
        main_layout = QVBoxLayout(self)
        main_layout.setSpacing(10)
        
        # Create scroll area for form fields
        self.scroll_area = QScrollArea()
        self.scroll_area.setWidgetResizable(True)
        self.scroll_area.setFrameShape(QFrame.NoFrame)
        
        # Create form container
        form_widget = QWidget()
        self.form_layout = QVBoxLayout(form_widget)
        self.form_layout.setSpacing(15)
        
        # Create fields
        self.fields = {}

        # # Right panel - OCR Editor
        # self.ocr_editor = OCRResultEditor(self.db)  # Truyền self.db vào
        # self.ocr_editor.ocr_updated.connect(self.update_document)
        # self.ocr_editor.field_changed.connect(self.field_value_changed)


        # Header with document title
        self.doc_title = QLabel("Thông tin văn bản")
        self.doc_title.setStyleSheet("font-size: 14pt; font-weight: bold; color: #2980b9;")
        self.form_layout.addWidget(self.doc_title)
        
        # CQBH with selection
        cqbh_group = QGroupBox("Cơ quan ban hành")
        cqbh_layout = QVBoxLayout(cqbh_group)
        
        cqbh_select_layout = QHBoxLayout()
        cqbh_select_layout.setSpacing(10)
        
        self.fields['cqbh'] = QLineEdit()
        completer = QCompleter([])
        completer.setCaseSensitivity(Qt.CaseInsensitive)
        self.fields['cqbh'].setCompleter(completer)
        self.fields['cqbh'].setPlaceholderText("Nhập cơ quan ban hành...")
        
        self.cqbh_selector = QComboBox()
        self.cqbh_selector.addItems(["Tất cả", "Trên", "Dưới"])
        self.cqbh_selector.currentTextChanged.connect(self.cqbh_selection_changed)
        self.cqbh_selector.setFixedWidth(120)
        
        cqbh_select_layout.addWidget(self.fields['cqbh'])
        cqbh_select_layout.addWidget(self.cqbh_selector)
        cqbh_layout.addLayout(cqbh_select_layout)
        
        self.form_layout.addWidget(cqbh_group)

        # Document fields
        # 1. Số ký hiệu & Loại văn bản
        doc_info_group = QGroupBox("Thông tin văn bản")
        doc_info_layout = QGridLayout(doc_info_group)
        doc_info_layout.setColumnStretch(0, 1)
        doc_info_layout.setColumnStretch(1, 1)
        
        # Số ký hiệu
        so_kh_label = QLabel("Số ký hiệu:")
        self.fields['so_ki_hieu'] = QLineEdit()
        self.fields['so_ki_hieu'].setPlaceholderText("Nhập số ký hiệu...")
        completer = QCompleter([])
        completer.setCaseSensitivity(Qt.CaseInsensitive)
        self.fields['so_ki_hieu'].setCompleter(completer)
        self.fields['so_ki_hieu'].textChanged.connect(lambda: self.field_changed.emit('so_ki_hieu', self.fields['so_ki_hieu'].text()))
        
        # Loại văn bản
        loai_vb_label = QLabel("Loại văn bản:")
        self.fields['loai_vb'] = QLineEdit()
        self.fields['loai_vb'].setPlaceholderText("Nhập loại văn bản...")
        completer = QCompleter([])
        completer.setCaseSensitivity(Qt.CaseInsensitive)
        self.fields['loai_vb'].setCompleter(completer)
        
        # Độ khẩn with icons
        do_khan_label = QLabel("Độ khẩn:")
        self.fields['do_khan'] = QComboBox()
        
        # Create độ khẩn items with icons
        self.fields['do_khan'].addItem(QIcon(), "Không")
        
        urgent_icon = QIcon()
        urgent_pixmap = QPixmap(16, 16)
        urgent_pixmap.fill(QColor(255, 0, 0))
        urgent_icon.addPixmap(urgent_pixmap)
        self.fields['do_khan'].addItem(urgent_icon, "Hỏa Tốc")
        
        secret_icon = QIcon()
        secret_pixmap = QPixmap(16, 16)
        secret_pixmap.fill(QColor(128, 0, 128))
        secret_icon.addPixmap(secret_pixmap)
        self.fields['do_khan'].addItem(secret_icon, "Độ Mật")
        
        self.fields['do_khan'].currentTextChanged.connect(lambda: self.field_changed.emit('do_khan', self.fields['do_khan'].currentText()))
        
        # Ngày ban hành with date picker
        ngay_bh_label = QLabel("Ngày ban hành:")
        self.fields['ngay_bh'] = QLineEdit()
        self.fields['ngay_bh'].setPlaceholderText("Nhập ngày ban hành...")
        
        # Add calendar button
        ngay_bh_layout = QHBoxLayout()
        ngay_bh_layout.setSpacing(5)
        ngay_bh_layout.addWidget(self.fields['ngay_bh'])
        
        calendar_btn = QPushButton(QIcon.fromTheme("office-calendar"), "")
        calendar_btn.setFixedWidth(30)
        calendar_btn.clicked.connect(self.show_date_picker)
        ngay_bh_layout.addWidget(calendar_btn)
        
        # Add fields to layout
        doc_info_layout.addWidget(so_kh_label, 0, 0)
        doc_info_layout.addWidget(self.fields['so_ki_hieu'], 0, 1)
        doc_info_layout.addWidget(loai_vb_label, 1, 0)
        doc_info_layout.addWidget(self.fields['loai_vb'], 1, 1)
        doc_info_layout.addWidget(do_khan_label, 2, 0)
        doc_info_layout.addWidget(self.fields['do_khan'], 2, 1)
        doc_info_layout.addWidget(ngay_bh_label, 3, 0)
        doc_info_layout.addLayout(ngay_bh_layout, 3, 1)
        
        self.form_layout.addWidget(doc_info_group)
        
        # 2. Nội dung chính
        nd_chinh_group = QGroupBox("Nội dung chính")
        nd_chinh_layout = QVBoxLayout(nd_chinh_group)
        
        self.fields['nd_chinh'] = QTextEdit()
        self.fields['nd_chinh'].setPlaceholderText("Nhập nội dung chính của văn bản...")
        self.fields['nd_chinh'].setMinimumHeight(100)
        
        nd_chinh_layout.addWidget(self.fields['nd_chinh'])
        self.form_layout.addWidget(nd_chinh_group)
        
        # 3. Nơi nhận
        noi_nhan_group = QGroupBox("Nơi nhận")
        noi_nhan_layout = QVBoxLayout(noi_nhan_group)
        
        self.fields['noi_nhan'] = QTextEdit()
        self.fields['noi_nhan'].setPlaceholderText("Nhập danh sách nơi nhận...")
        self.fields['noi_nhan'].setMinimumHeight(80)
        
        noi_nhan_layout.addWidget(self.fields['noi_nhan'])
        self.form_layout.addWidget(noi_nhan_group)
        
        # 4. Chức vụ & chữ ký
        sign_group = QGroupBox("Thông tin ký duyệt")
        sign_layout = QVBoxLayout(sign_group)
        
        # Chức vụ with selector
        chuc_vu_layout = QHBoxLayout()
        self.fields['chuc_vu'] = QLineEdit()
        self.fields['chuc_vu'].setPlaceholderText("Nhập chức vụ người ký...")
        completer = QCompleter([])
        completer.setCaseSensitivity(Qt.CaseInsensitive)
        self.fields['chuc_vu'].setCompleter(completer)
        
        self.chuc_vu_selector = QComboBox()
        self.chuc_vu_selector.addItems(["Tất cả", "Chức danh"])
        self.chuc_vu_selector.currentTextChanged.connect(self.chuc_vu_selection_changed)
        self.chuc_vu_selector.setFixedWidth(120)
        
        chuc_vu_layout.addWidget(self.fields['chuc_vu'])
        chuc_vu_layout.addWidget(self.chuc_vu_selector)
        
        # Chữ ký
        self.fields['chu_ky'] = QLineEdit()
        self.fields['chu_ky'].setPlaceholderText("Nhập tên người ký...")
        
        sign_layout.addWidget(QLabel("Chức vụ:"))
        sign_layout.addLayout(chuc_vu_layout)
        sign_layout.addWidget(QLabel("Người ký:"))
        sign_layout.addWidget(self.fields['chu_ky'])
        
        self.form_layout.addWidget(sign_group)
        
        # Add form to scroll area
        self.scroll_area.setWidget(form_widget)
        main_layout.addWidget(self.scroll_area)
        
        # Add form fields to change monitoring
        for field_id, widget in self.fields.items():
            if field_id != 'cqbh' and field_id != 'do_khan':  # These have special handling
                if isinstance(widget, QLineEdit):
                    widget.textChanged.connect(self.content_changed)
                elif isinstance(widget, QTextEdit):
                    widget.textChanged.connect(self.content_changed)
                elif isinstance(widget, QComboBox):
                    widget.currentTextChanged.connect(self.content_changed)

        # Buttons
        buttons_layout = QHBoxLayout()
        buttons_layout.setSpacing(10)
        
        self.save_btn = QPushButton("Lưu thay đổi")
        self.save_btn.setIcon(QIcon.fromTheme("document-save"))
        self.save_btn.clicked.connect(self.save_changes)
        self.save_btn.setEnabled(False)
        
        self.reset_btn = QPushButton("Đặt lại")
        self.reset_btn.setIcon(QIcon.fromTheme("edit-undo"))
        self.reset_btn.clicked.connect(self.reset_fields)
        
        buttons_layout.addStretch()
        buttons_layout.addWidget(self.save_btn)
        buttons_layout.addWidget(self.reset_btn)
        
        main_layout.addLayout(buttons_layout)
        
        # Apply stylesheets
        for field_id, widget in self.fields.items():
            if isinstance(widget, QLineEdit):
                widget.setMinimumHeight(30)
            elif isinstance(widget, QComboBox):
                widget.setMinimumHeight(30)
    
    def show_date_picker(self):
        """Show date picker for ngay_bh field"""
        date_dialog = QDialog(self)
        date_dialog.setWindowTitle("Chọn ngày")
        date_dialog.setMinimumSize(300, 300)
        
        layout = QVBoxLayout(date_dialog)
        
        calendar = QCalendarWidget()
        layout.addWidget(calendar)
        
        # Try to set initial date from current value
        current_date = self.fields['ngay_bh'].text()
        if current_date:
            try:
                # Try common Vietnamese date formats
                for fmt in ['%d/%m/%Y', '%d-%m-%Y', '%d.%m.%Y', '%d/%m/%y']:
                    try:
                        date = datetime.strptime(current_date, fmt)
                        calendar.setSelectedDate(QDate(date.year, date.month, date.day))
                        break
                    except:
                        continue
            except:
                pass
        
        # Buttons
        buttons = QHBoxLayout()
        ok_btn = QPushButton("Chọn")
        ok_btn.clicked.connect(date_dialog.accept)
        cancel_btn = QPushButton("Hủy")
        cancel_btn.clicked.connect(date_dialog.reject)
        
        buttons.addStretch()
        buttons.addWidget(ok_btn)
        buttons.addWidget(cancel_btn)
        layout.addLayout(buttons)
        
        if date_dialog.exec_() == QDialog.Accepted:
            selected_date = calendar.selectedDate()
            # Format date as dd/mm/yyyy
            formatted_date = f"{selected_date.day():02d}/{selected_date.month():02d}/{selected_date.year():04d}"
            self.fields['ngay_bh'].setText(formatted_date)
            self.content_changed()

    def chuc_vu_selection_changed(self, selection):
        """Handle change in chuc_vu selection dropdown"""
        if self.is_loading or not self.current_doc_data:
            return
            
        chuc_vu_text = self.current_doc_data[10] or ''
        lines = chuc_vu_text.strip().split('\n')
        
        if selection == "Tất cả":
            self.fields['chuc_vu'].setText(chuc_vu_text.strip())
        elif selection == "Chức danh" and lines:
            self.fields['chuc_vu'].setText(lines[-1].strip())
        
        self.content_changed()

    def cqbh_selection_changed(self, selection):
        """Handle change in cqbh selection dropdown"""
        if self.is_loading:
            return
            
        if self.full_cqbh_text:
            lines = self.full_cqbh_text.strip().split('\n')
            if selection == "Tất cả":
                self.fields['cqbh'].setText(self.full_cqbh_text.strip())
            elif selection == "Trên" and lines:
                self.fields['cqbh'].setText(lines[0].strip())
            elif selection == "Dưới" and len(lines) > 1:
                self.fields['cqbh'].setText(lines[1].strip())
            self.content_changed()

    def content_changed(self):
        """Enable save button when content changes"""
        if not self.is_loading and self.doc_id is not None:
            self.save_btn.setEnabled(True)

    def load_data(self, doc_data):
        """Load document data into the editor fields"""
        if not doc_data:
            return
                
        try:
            self.is_loading = True
            
            # Store document ID from input data
            self.doc_id = doc_data[0]  # Lưu ID từ dữ liệu đầu vào
            self.current_doc_data = doc_data
            
            # Set title with correct document ID
            self.doc_title.setText(f"Văn bản: {self.doc_id} - {doc_data[2] if len(doc_data) > 2 else ''}")
            
            # Handle CQBH
            self.full_cqbh_text = f"{doc_data[3] or ''}\n{doc_data[4] or ''}".strip()
            selection = self.cqbh_selector.currentText()
            
            if selection == "Tất cả":
                self.fields['cqbh'].setText(self.full_cqbh_text)
            else:
                lines = self.full_cqbh_text.split('\n')
                if selection == "Trên" and lines:
                    self.fields['cqbh'].setText(lines[0].strip())
                elif selection == "Dưới" and len(lines) > 1:
                    self.fields['cqbh'].setText(lines[1].strip())
            
            # Load other fields
            field_mapping = {
                'so_ki_hieu': 5,
                'loai_vb': 6,
                'nd_chinh': 7,
                'ngay_bh': 8,
                'noi_nhan': 9,
                'chuc_vu': 10,
                'chu_ky': 11,
                'do_khan': 12
            }
            
            for field_id, col_index in field_mapping.items():
                if field_id in self.fields and col_index < len(doc_data):
                    editor = self.fields[field_id]
                    value = doc_data[col_index] or ''
                    
                    if field_id == 'do_khan':
                        # For do_khan ComboBox
                        index = self.fields[field_id].findText(value)
                        if index >= 0:
                            self.fields[field_id].setCurrentIndex(index)
                        else:
                            self.fields[field_id].setCurrentIndex(0)  # Default to "Không"
                    elif isinstance(editor, QTextEdit):
                        editor.setPlainText(value.strip())
                    else:
                        editor.setText(value.strip())
            
            self.save_btn.setEnabled(False)
            
        except Exception as e:
            logger.error(f"Error loading document data: {str(e)}")
        finally:
            self.is_loading = False

    def get_current_data(self):
        """Get current data from all fields"""
        data = {}
        
        # Handle CQBH based on selection
        lines = self.full_cqbh_text.strip().split('\n')
        selection = self.cqbh_selector.currentText()
        
        if selection == "Trên":
            data['cqbh_tren'] = self.fields['cqbh'].text()
            data['cqbh_duoi'] = lines[1] if len(lines) > 1 else ''
        elif selection == "Dưới":
            data['cqbh_tren'] = lines[0] if lines else ''
            data['cqbh_duoi'] = self.fields['cqbh'].text()
        else:
            parts = self.fields['cqbh'].text().split('\n')
            data['cqbh_tren'] = parts[0] if parts else ''
            data['cqbh_duoi'] = parts[1] if len(parts) > 1 else ''
        
        # Get other fields
        for field_id, editor in self.fields.items():
            if field_id != 'cqbh':
                if field_id == 'do_khan':
                    # Get selected value from ComboBox
                    data[field_id] = self.fields[field_id].currentText()
                elif isinstance(editor, QTextEdit):
                    data[field_id] = editor.toPlainText().strip()
                else:
                    data[field_id] = editor.text().strip()
        
        return data

    def save_changes(self):
        """Save changes to current document"""
        if self.doc_id is None:
            return
            
        try:
            # Kiểm tra document tồn tại
            doc_info = self.db.get_document_info(self.doc_id)
            if not doc_info:
                raise ValueError(f"Document with ID {self.doc_id} not found")
                
            data = self.get_current_data()
            # Emit update signal
            self.ocr_updated.emit(self.doc_id, data)
            self.save_btn.setEnabled(False)
            
        except Exception as e:
            logger.error(f"Error saving changes: {str(e)}")
            QMessageBox.critical(None, "Error", 
                            f"Failed to save changes: {str(e)}\nPlease try again.")

    def reset_fields(self):
        """Reset fields to original values"""
        if self.doc_id is not None:
            self.load_data(self.current_doc_data)

    def update_suggestions(self, db):
        """Update suggestions for all fields from database"""
        try:
            for field_id, editor in self.fields.items():
                if isinstance(editor, QLineEdit):
                    suggestions = db.get_suggestions(field_id)
                    completer = QCompleter(suggestions)
                    completer.setCaseSensitivity(Qt.CaseInsensitive)
                    completer.setFilterMode(Qt.MatchContains)
                    editor.setCompleter(completer)
        except Exception as e:
            logger.error(f"Error updating suggestions: {str(e)}")

    def clear(self):
        """Clear all fields"""
        self.is_loading = True
        try:
            self.doc_id = None
            self.full_cqbh_text = ""
            self.current_doc_data = None
            self.doc_title.setText("Thông tin văn bản")
            
            for field_id, editor in self.fields.items():
                if field_id == 'do_khan':
                    editor.setCurrentIndex(0)  # Reset to "Không"
                elif isinstance(editor, QTextEdit):
                    editor.clear()
                else:
                    editor.setText("")
            
            self.save_btn.setEnabled(False)
        finally:
            self.is_loading = False
            
    def set_read_only(self, read_only=True):
        """Set all fields to read-only mode"""
        for field_id, editor in self.fields.items():
            if isinstance(editor, QLineEdit):
                editor.setReadOnly(read_only)
            elif isinstance(editor, QTextEdit):
                editor.setReadOnly(read_only)
            elif isinstance(editor, QComboBox):
                editor.setEnabled(not read_only)
                
        self.cqbh_selector.setEnabled(not read_only)
        self.chuc_vu_selector.setEnabled(not read_only)
        self.save_btn.setEnabled(not read_only and self.doc_id is not None)
        self.reset_btn.setEnabled(not read_only)
#############################
#       Dialog Classes      #
#############################
class ProgressDialog(QDialog):
    """Dialog for showing progress of long operations"""
    canceled = pyqtSignal()
    
    def __init__(self, total: int, title: str = "Processing", parent=None):
        super().__init__(parent)
        self.total = total
        self.was_canceled = False
        self.setup_ui(title)
        
    def setup_ui(self, title: str):
        self.setWindowTitle(title)
        self.setWindowModality(Qt.ApplicationModal)
        self.resize(400, 150)
        self.setMinimumWidth(400)
        
        layout = QVBoxLayout(self)
        layout.setSpacing(15)
        
        # Task label
        self.task_label = QLabel(title)
        self.task_label.setAlignment(Qt.AlignCenter)
        self.task_label.setStyleSheet("font-size: 12pt; font-weight: bold;")
        layout.addWidget(self.task_label)

        # Progress bar
        self.progress_bar = QProgressBar()
        self.progress_bar.setRange(0, self.total)
        self.progress_bar.setValue(0)
        self.progress_bar.setMinimumHeight(24)
        self.progress_bar.setTextVisible(True)
        layout.addWidget(self.progress_bar)

        # Status label
        self.status_label = QLabel("Preparing...")
        self.status_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(self.status_label)
        
        # Time estimate
        self.time_label = QLabel("")
        self.time_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(self.time_label)
        
        # Cancel button
        self.cancel_btn = QPushButton("Cancel")
        self.cancel_btn.setIcon(QIcon.fromTheme("process-stop"))
        self.cancel_btn.clicked.connect(self.cancel_operation)
        
        button_layout = QHBoxLayout()
        button_layout.addStretch()
        button_layout.addWidget(self.cancel_btn)
        button_layout.addStretch()
        layout.addLayout(button_layout)
        
        # Timer for estimates
        self.start_time = None
        self.timer = QTimer(self)
        self.timer.timeout.connect(self.update_time_estimate)
        
    def show(self):
        """Show dialog and start timer"""
        self.start_time = time.time()
        self.timer.start(1000)  # Update every second
        super().show()
        
    def cancel_operation(self):
        """Handle cancel button click"""
        self.was_canceled = True
        self.task_label.setText("Canceling...")
        self.canceled.emit()
        self.reject()
        
    def update_time_estimate(self):
        """Update time estimate display"""
        if self.start_time is None or self.progress_bar.value() == 0:
            return
            
        elapsed = time.time() - self.start_time
        progress = self.progress_bar.value() / max(1, self.total)
        
        if progress > 0:
            total_estimated = elapsed / progress
            remaining = total_estimated - elapsed
            
            if remaining > 60:
                mins = int(remaining // 60)
                secs = int(remaining % 60)
                self.time_label.setText(f"Estimated time remaining: {mins}m {secs}s")
            else:
                self.time_label.setText(f"Estimated time remaining: {int(remaining)}s")
        else:
            self.time_label.setText("Calculating time remaining...")

    def update_progress(self, current: int, total: int = None, message: str = None):
        if total is not None:
            self.progress_bar.setRange(0, total)
            self.total = total
        
        self.progress_bar.setValue(current)
        # Thêm kiểm tra để tránh chia cho 0
        percentage = (current / self.total) * 100 if self.total > 0 else 0
        
        # Cập nhật thông báo nếu được cung cấp
        if message:
            self.status_label.setText(message)
        else:
            self.status_label.setText(f"Processing... {percentage:.1f}%")
            
    def closeEvent(self, event):
        """Handle dialog close"""
        self.timer.stop()
        super().closeEvent(event)

class PreviewDialog(QDialog):
    """Dialog for previewing PDF before processing"""
    
    def __init__(self, pdf_path: str, parent=None):
        super().__init__(parent)
        self.pdf_path = pdf_path
        self.setup_ui()
        
    def setup_ui(self):
        self.setWindowTitle("PDF Preview")
        self.setWindowModality(Qt.ApplicationModal)
        self.resize(800, 600)
        
        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 10)
        
        # File info
        info_layout = QHBoxLayout()
        
        file_info = os.path.basename(self.pdf_path)
        if os.path.exists(self.pdf_path):
            size_mb = os.path.getsize(self.pdf_path) / (1024 * 1024)
            file_info += f" ({size_mb:.1f} MB)"
            
        info_label = QLabel(f"File: {file_info}")
        info_label.setStyleSheet("font-weight: bold; padding: 5px;")
        info_layout.addWidget(info_label)
        info_layout.addStretch()
        
        # PDF viewer
        self.viewer = PDFViewer()
        if not self.viewer.load_pdf(self.pdf_path):
            # Error loading PDF - show message
            error_widget = QWidget()
            error_layout = QVBoxLayout(error_widget)
            error_icon = QLabel()
            error_icon.setPixmap(QIcon.fromTheme("dialog-error").pixmap(64, 64))
            error_icon.setAlignment(Qt.AlignCenter)
            error_message = QLabel(f"Error loading PDF file:\n{self.pdf_path}")
            error_message.setAlignment(Qt.AlignCenter)
            error_message.setStyleSheet("color: red; font-weight: bold;")
            error_layout.addStretch()
            error_layout.addWidget(error_icon)
            error_layout.addWidget(error_message)
            error_layout.addStretch()
            layout.addWidget(error_widget)
        else:
            layout.addLayout(info_layout)
            layout.addWidget(self.viewer)

        # Buttons
        buttons = QHBoxLayout()
        self.process_btn = QPushButton("Process Document")
        self.process_btn.setIcon(QIcon.fromTheme("document-send"))
        self.process_btn.clicked.connect(self.accept)
        self.process_btn.setEnabled(os.path.exists(self.pdf_path))
        
        cancel_btn = QPushButton("Cancel")
        cancel_btn.setIcon(QIcon.fromTheme("dialog-cancel"))
        cancel_btn.clicked.connect(self.reject)
        
        buttons.addStretch()
        buttons.addWidget(self.process_btn)
        buttons.addWidget(cancel_btn)
        
        layout.addLayout(buttons)
#############################
#     Worker Threads        #
#############################
class OCRWorker(QThread):
    """Worker thread for OCR processing"""
    finished = pyqtSignal(dict, list)  # results, page_detections
    progress = pyqtSignal(int, int, str)  # current, total, message
    error = pyqtSignal(str)

    def __init__(self, ocr_system, file_path):
        super().__init__()
        self.ocr_system = ocr_system
        self.file_path = file_path
        self.canceled = False

    def run(self):
        try:
            if not os.path.exists(self.file_path):
                self.error.emit(f"File not found: {self.file_path}")
                return
            
            def progress_callback(current, total, message="Processing..."):
                self.progress.emit(current, total, message)
                # Check if canceled
                return not self.canceled
                
            results, all_page_detections = self.ocr_system.process_document(
                self.file_path,
                progress_callback=progress_callback
            )
            
            if not self.canceled:
                self.finished.emit(results, all_page_detections)
                
        except Exception as e:
            if not self.canceled:
                logger.error(f"OCR worker error: {str(e)}")
                self.error.emit(str(e))
    
    def cancel(self):
        """Cancel the OCR process"""
        self.canceled = True

class BatchProcessWorker(QThread):
    """Worker thread for batch OCR processing"""
    progress = pyqtSignal(int, int, str)
    finished = pyqtSignal(list)
    error = pyqtSignal(str)

    def __init__(self, ocr_system, file_paths):
        super().__init__()
        self.ocr_system = ocr_system
        self.file_paths = file_paths
        self.canceled = False

    def run(self):
        try:
            results = []
            total = len(self.file_paths)
            
            for i, file_path in enumerate(self.file_paths, 1):
                if self.canceled:
                    break
                    
                if not os.path.exists(file_path):
                    logger.warning(f"File not found: {file_path}")
                    self.progress.emit(i, total, f"Skipped: {os.path.basename(file_path)} (not found)")
                    continue
                    
                try:
                    self.progress.emit(i, total, f"Processing: {os.path.basename(file_path)}")
                    result, page_detections = self.ocr_system.process_document(
                        file_path,
                        progress_callback=lambda curr, tot: not self.canceled
                    )
                    results.append((file_path, result, page_detections))
                except Exception as e:
                    logger.error(f"Error processing {file_path}: {str(e)}")
                    
                if not self.canceled:
                    self.progress.emit(i, total, f"Completed: {os.path.basename(file_path)}")
            
            if not self.canceled:
                self.finished.emit(results)
                
        except Exception as e:
            if not self.canceled:
                logger.error(f"Batch processing error: {str(e)}")
                self.error.emit(str(e))
    
    def cancel(self):
        """Cancel the batch process"""
        self.canceled = True

class FileRepairWorker(QThread):
    """Worker thread for repairing broken file paths"""
    progress = pyqtSignal(int, int, str)
    file_path_request = pyqtSignal(str, str)  # doc_id, old_path
    file_path_response = None
    finished = pyqtSignal(list, list)  # fixed_files, remaining_issues
    error = pyqtSignal(str)

    def __init__(self, db, missing_files):
        super().__init__()
        self.db = db
        self.missing_files = missing_files
        self.canceled = False
        self.wait_condition = threading.Condition()
        
    def run(self):
        try:
            fixed_files = []
            remaining_issues = []
            total = len(self.missing_files)
            
            for i, (doc_id, old_path) in enumerate(self.missing_files, 1):
                if self.canceled:
                    break
                    
                self.progress.emit(i, total, f"Checking: {os.path.basename(old_path)}")
                
                # 1. Try to find file in original location
                if os.path.exists(old_path):
                    fixed_files.append((doc_id, old_path))
                    continue
                    
                # 2. Try to find file in backup directory
                backup_name = Path(old_path).name
                backup_path = BACKUP_DIR / backup_name
                if os.path.exists(backup_path):
                    if self.db.update_file_path(doc_id, str(backup_path)):
                        fixed_files.append((doc_id, str(backup_path)))
                        continue
                
                # 3. Ask user for file location
                filename = Path(old_path).name
                self.file_path_response = None
                
                with self.wait_condition:
                    self.file_path_request.emit(str(doc_id), old_path)
                    # Wait for response with timeout
                    self.wait_condition.wait(30)  # 30 second timeout
                
                if self.file_path_response:
                    new_path = self.file_path_response
                    if os.path.exists(new_path):
                        if self.db.update_file_path(doc_id, new_path):
                            fixed_files.append((doc_id, new_path))
                            continue
                
                # File not found or update failed
                remaining_issues.append((doc_id, old_path))
                
                if not self.canceled:
                    self.progress.emit(i, total, f"Processed: {i}/{total}")
            
            if not self.canceled:
                self.finished.emit(fixed_files, remaining_issues)
                
        except Exception as e:
            if not self.canceled:
                logger.error(f"File repair error: {str(e)}")
                self.error.emit(str(e))
    
    def set_file_path_response(self, file_path):
        """Set response from main thread with new file path"""
        with self.wait_condition:
            self.file_path_response = file_path
            self.wait_condition.notify()
    
    def cancel(self):
        """Cancel the repair process"""
        self.canceled = True
        with self.wait_condition:
            self.wait_condition.notify()  # Wake up waiting thread

#############################
#      Main Window Class    #
#############################
class MainWindow(QMainWindow):
    """Main application window"""
    
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Hệ thống OCR Văn Bản")
            
        # Get display information
        screen = QApplication.primaryScreen().geometry()
        window_width = int(min(1280, screen.width() * 0.8))
        window_height = int(min(900, screen.height() * 0.8))
        self.setGeometry(
            int((screen.width() - window_width) // 2),
            int((screen.height() - window_height) // 2),
            window_width,
            window_height
        )
        
        # Set minimum size
        self.setMinimumSize(1024, 768)

        # Initialize settings
        self.settings = QSettings("OCRApp", "DocumentManagement")
        self.load_window_settings()

        # Initialize managers
        self.theme_manager = ThemeManager()
        self.db = DocumentDatabase()
        
        # Document tracking
        self.current_doc_id = None
        self.detections_by_page = {}
        self.current_view_mode = "details" 
        
        # Worker threads
        self.ocr_worker = None
        self.batch_worker = None
        self.repair_worker = None
        
        # Auto-save timer
        self.autosave_timer = QTimer(self)
        self.autosave_timer.timeout.connect(self.auto_save)
        self.autosave_timer.setInterval(AUTOSAVE_INTERVAL)
        
        # Khởi tạo OCR system chỉ với model YOLO
        FIXED_MODEL_PATH = r"D:\OCR_VanBang_Local_New\Code\models\best.pt"  

        # Kiểm tra model
        if not os.path.exists(FIXED_MODEL_PATH):
            QMessageBox.critical(None, "Lỗi", f"Không tìm thấy model tại: {FIXED_MODEL_PATH}")
            sys.exit()

        # Khởi tạo OCR system với EasyOCR
        self.ocr_system = DocumentOCR(FIXED_MODEL_PATH)

        # self.db = DocumentDatabase()
        # Setup UI
        self.setup_ui()
        self.setup_menus()
        self.setup_shortcuts()
        self.setup_statusbar()
        
        # Apply theme
        self.theme_manager.apply_theme(QApplication.instance())
        
        # Verify file paths and repair broken links
        QTimer.singleShot(500, self.check_file_paths)
        
        # Load existing documents
        self.load_documents()

    def load_window_settings(self):
        """Load window position and size from settings"""
        geometry = self.settings.value("geometry")
        if geometry:
            self.restoreGeometry(geometry)
            
        state = self.settings.value("windowState")
        if state:
            self.restoreState(state)

    def save_window_settings(self):
        """Save window position and size to settings"""
        self.settings.setValue("geometry", self.saveGeometry())
        self.settings.setValue("windowState", self.saveState())

    def load_api_keys(self):
        """Load API keys from settings"""
        api_keys_str = self.settings.value("api_keys", "")
        if api_keys_str:
            return api_keys_str.split(',')
        return []

    def save_api_keys(self, keys):
        """Save API keys to settings"""
        if keys:
            self.settings.setValue("api_keys", ",".join(keys))
        else:
            self.settings.remove("api_keys")

    def check_file_paths(self):
        """Check and potentially repair missing file paths"""
        missing_files = self.db.verify_file_paths()
        if missing_files:
            reply = QMessageBox.question(
                self,
                "Tệp tin bị thiếu",
                f"Tìm thấy {len(missing_files)} tài liệu có đường dẫn tệp bị thiếu. Bạn có muốn sửa chữa không?",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.Yes
            )
            
            if reply == QMessageBox.Yes:
                self.repair_file_paths(missing_files)
    def repair_file_paths(self, missing_files):
        """Repair missing file paths"""
        if not missing_files:
            return
            
        # Create progress dialog
        progress = ProgressDialog(len(missing_files), "Repairing File Paths", self)
        
        # Create worker thread
        self.repair_worker = FileRepairWorker(self.db, missing_files)
        self.repair_worker.progress.connect(progress.update_progress)
        self.repair_worker.file_path_request.connect(self.request_file_path)
        self.repair_worker.finished.connect(self.repair_completed)
        self.repair_worker.error.connect(self.show_error)
        progress.canceled.connect(self.repair_worker.cancel)
        
        # Start worker
        self.repair_worker.start()
        progress.exec_()
        
    def request_file_path(self, doc_id, old_path):
        """Request file path from user"""
        filename = Path(old_path).name
        
        # Ask user for the file
        msg = f"Đang tìm tệp: {filename}\nĐường dẫn cũ: {old_path}"
        QMessageBox.information(self, "Find File", msg)
        
        new_path, _ = QFileDialog.getOpenFileName(
            self, f"Locate {filename}", "", "PDF Files (*.pdf)"
        )
        
        # Send response back to worker thread
        self.repair_worker.set_file_path_response(new_path)
        
    def repair_completed(self, fixed_files, remaining_issues):
        """Handle completion of file path repair"""
        # Update UI
        if fixed_files:
            self.load_documents()
            
        # Show summary
        if remaining_issues:
            msg = f"Đã sửa chữa {len(fixed_files)} tệp, nhưng vẫn còn {len(remaining_issues)} vấn đề.\n\n"
            msg += "Các tệp sau vẫn bị thiếu:\n"
            for doc_id, path in remaining_issues[:10]:  # Show first 10
                msg += f"- ID {doc_id}: {os.path.basename(path)}\n"
                
            if len(remaining_issues) > 10:
                msg += f"...và {len(remaining_issues) - 10} tệp khác."
                
            QMessageBox.warning(self, "Repair Summary", msg)
        else:
            QMessageBox.information(
                self, 
                "Repair Complete", 
                f"Đã sửa chữa thành công {len(fixed_files)} tệp."
            )

    def setup_ui(self):
        """Set up the main user interface"""
        # Base setup
        main_widget = QWidget()
        self.setCentralWidget(main_widget)
        main_layout = QVBoxLayout(main_widget)
        main_layout.setContentsMargins(2, 0, 2, 2)  # Minimal margins
        main_layout.setSpacing(0)  # No spacing between major sections

        # 1. Setup Header/Toolbar
        toolbar_widget = QWidget()
        toolbar_widget.setFixedHeight(32)  # Fixed height for header
        toolbar_layout = QHBoxLayout(toolbar_widget)
        toolbar_layout.setContentsMargins(4, 0, 4, 0)
        toolbar_layout.setSpacing(8)

        # Search controls
        search_label = QLabel("Tìm:")
        search_label.setFixedWidth(25)
        
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("Nhập từ khóa tìm kiếm...")
        self.search_input.setClearButtonEnabled(True)
        self.search_input.setMinimumWidth(200)
        self.search_input.textChanged.connect(lambda text: self.search_timer.start(300))

        self.search_type = QComboBox()
        self.search_type.addItems(["Tất cả", "Tên file", "Nội dung", "Số ký hiệu", "Độ khẩn"])
        self.search_type.setFixedWidth(100)
        self.search_type.currentTextChanged.connect(lambda: self.search_timer.start(300))

        self.advanced_search_btn = QPushButton("Nâng cao") 
        self.advanced_search_btn.setFixedHeight(30)
        self.advanced_search_btn.setFixedWidth(90)
        self.advanced_search_btn.clicked.connect(self.show_advanced_search)

        # View selector
        view_group = QFrame()
        view_layout = QHBoxLayout(view_group)
        view_layout.setContentsMargins(0, 0, 0, 0)
        view_layout.setSpacing(4)

        self.view_buttons = QButtonGroup(self)
        self.details_view_btn = QRadioButton("Chi tiết")
        self.gallery_view_btn = QRadioButton("Thư viện")
        
        self.details_view_btn.setChecked(True)
        self.details_view_btn.clicked.connect(lambda: self.switch_view("details"))
        self.gallery_view_btn.clicked.connect(lambda: self.switch_view("gallery"))
        
        self.view_buttons.addButton(self.details_view_btn)
        self.view_buttons.addButton(self.gallery_view_btn)
        
        view_layout.addWidget(self.details_view_btn)
        view_layout.addWidget(self.gallery_view_btn)

        # Add all controls to toolbar
        toolbar_layout.addWidget(search_label)
        toolbar_layout.addWidget(self.search_input)
        toolbar_layout.addWidget(self.search_type)
        toolbar_layout.addWidget(self.advanced_search_btn)
        toolbar_layout.addStretch()  # Flexible space
        toolbar_layout.addWidget(view_group)

        main_layout.addWidget(toolbar_widget)

        # 2. Main Content Area with Splitter
        content_widget = QWidget()
        content_layout = QHBoxLayout(content_widget)
        content_layout.setContentsMargins(0, 0, 0, 0)
        content_layout.setSpacing(0)

        splitter = QSplitter(Qt.Horizontal)
        splitter.setHandleWidth(1)  # Thin splitter handle
        splitter.setChildrenCollapsible(False)

        # LEFT PANEL - Document List and Versions
        left_panel = QWidget()
        left_layout = QVBoxLayout(left_panel)
        left_layout.setContentsMargins(0, 0, 0, 0)
        left_layout.setSpacing(1)

        # Document Tabs
        self.doc_tabs = QTabWidget()
        self.doc_tabs.setDocumentMode(True)
        self.doc_tabs.setStyleSheet("""
            QTabBar::tab {
                height: 24px;
                padding: 2px 8px;
            }
        """)

        # Document Table
        self.doc_table = QTableWidget()
        self.doc_table.setColumnCount(5)
        self.doc_table.setHorizontalHeaderLabels([
            "ID", "Tên File", "Ngày tạo", "Phiên bản", "Số ký hiệu"
        ])
        self.doc_table.horizontalHeader().setFixedHeight(25)
        self.doc_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.doc_table.setSelectionMode(QTableWidget.SingleSelection)
        self.doc_table.setSortingEnabled(True)
        self.doc_table.verticalHeader().setVisible(False)
        self.doc_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.doc_table.setAlternatingRowColors(True)
        self.doc_table.itemClicked.connect(self.document_selected)
        self.doc_table.setContextMenuPolicy(Qt.CustomContextMenu)
        self.doc_table.customContextMenuRequested.connect(self.show_document_context_menu)

        # Set column widths
        self.doc_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)  # ID
        self.doc_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)  # Filename
        self.doc_table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeToContents)  # Date
        self.doc_table.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeToContents)  # Versions
        self.doc_table.horizontalHeader().setSectionResizeMode(4, QHeaderView.ResizeToContents)  # Number

        # Document Gallery
        self.doc_gallery = QTableWidget()
        self.doc_gallery.setIconSize(QSize(150, 200))
        self.doc_gallery.setColumnCount(4)
        self.doc_gallery.horizontalHeader().setVisible(False)
        self.doc_gallery.verticalHeader().setVisible(False)
        self.doc_gallery.setShowGrid(False)
        for i in range(4):
            self.doc_gallery.setColumnWidth(i, 175)
        self.doc_gallery.itemClicked.connect(self.gallery_item_selected)
        self.doc_gallery.setContextMenuPolicy(Qt.CustomContextMenu)
        self.doc_gallery.customContextMenuRequested.connect(self.show_gallery_context_menu)

        # Add views to tabs
        self.doc_tabs.addTab(self.doc_table, "Danh sách")
        self.doc_tabs.addTab(self.doc_gallery, "Thư viện")

        # Version History
        version_group = QGroupBox("Lịch sử phiên bản")
        version_layout = QVBoxLayout(version_group)
        version_layout.setContentsMargins(4, 8, 4, 4)
        version_layout.setSpacing(2)

        self.version_list = QTableWidget()
        self.version_list.setColumnCount(3)
        self.version_list.setHorizontalHeaderLabels(["Phiên bản", "Người sửa", "Ngày sửa"])
        self.version_list.horizontalHeader().setFixedHeight(25)
        self.version_list.setSelectionBehavior(QTableWidget.SelectRows)
        self.version_list.setAlternatingRowColors(True)
        self.version_list.itemClicked.connect(self.version_selected)
        self.version_list.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)
        self.version_list.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeToContents)
        self.version_list.horizontalHeader().setSectionResizeMode(2, QHeaderView.Stretch)

        version_layout.addWidget(self.version_list)

        # Add to left panel
        left_layout.addWidget(self.doc_tabs, 7)  # 70% height
        left_layout.addWidget(version_group, 3)   # 30% height

        splitter.addWidget(left_panel)

    # CENTER PANEL - PDF Viewer and Metadata
        center_panel = QTabWidget()
        center_panel.setDocumentMode(True)
        center_panel.setStyleSheet("""
            QTabBar::tab {
                height: 24px;
                padding: 2px 8px;
            }
        """)

        # PDF Viewer Tab
        self.pdf_viewer = PDFViewer()
        self.pdf_viewer.pageChanged.connect(self.page_changed)
        # Kết nối signal custom box đến hàm xử lý
        self.pdf_viewer.customBoxCreated.connect(self.process_custom_box)
        center_panel.addTab(self.pdf_viewer, "PDF Viewer")

        # Metadata Tab
        metadata_scroll = QScrollArea()
        metadata_scroll.setWidgetResizable(True)
        metadata_scroll.setFrameShape(QFrame.NoFrame)
        
        metadata_content = QWidget()
        metadata_layout = QVBoxLayout(metadata_content)
        metadata_layout.setSpacing(8)
        metadata_layout.setContentsMargins(8, 8, 8, 8)

        # Create metadata form
        metadata_form = QWidget()
        form_layout = QFormLayout(metadata_form)
        form_layout.setVerticalSpacing(6)
        form_layout.setFieldGrowthPolicy(QFormLayout.AllNonFixedFieldsGrow)

        # Initialize metadata fields
        self.metadata_fields = {}
        metadata_field_labels = [
            ('id', 'ID'),
            ('filename', 'Tên file'),
            ('created', 'Ngày tạo'),
            ('modified', 'Sửa đổi'),
            ('pages', 'Số trang'),
            ('filesize', 'Kích thước'),
            ('versions', 'Phiên bản'),
            ('cqbh', 'Cơ quan BH'),
            ('so_kh', 'Số ký hiệu'),
            ('loai_vb', 'Loại văn bản'),
            ('do_khan', 'Độ khẩn')
        ]

        for field_id, label in metadata_field_labels:
            # Create label
            field_label = QLabel(f"<b>{label}:</b>")
            
            # Create value field
            field = QLabel()
            field.setFrameShape(QFrame.StyledPanel)
            field.setFrameShadow(QFrame.Sunken)
            field.setMinimumHeight(22)
            field.setTextInteractionFlags(Qt.TextSelectableByMouse)
            field.setStyleSheet("""
                QLabel {
                    padding: 2px 4px;
                    background: white;
                }
            """)
            
            self.metadata_fields[field_id] = field
            form_layout.addRow(field_label, field)

        metadata_layout.addWidget(metadata_form)
        metadata_layout.addStretch()
        metadata_scroll.setWidget(metadata_content)
        center_panel.addTab(metadata_scroll, "Metadata")

        splitter.addWidget(center_panel)

        # RIGHT PANEL - OCR Editor
        self.ocr_editor = OCRResultEditor(db=self.db)
        self.ocr_editor.ocr_updated.connect(self.update_document)
        self.ocr_editor.field_changed.connect(self.field_value_changed)
        splitter.addWidget(self.ocr_editor)

        # Set initial splitter sizes (ratio: 25:45:30)
        splitter.setSizes([
            int(self.width() * 0.25),  # Left panel
            int(self.width() * 0.45),  # Center panel
            int(self.width() * 0.30)   # Right panel
        ])

        # Add splitter to content layout
        content_layout.addWidget(splitter)
        main_layout.addWidget(content_widget, 1)  # Give content area stretch factor of 1

        # Initialize search timer
        self.search_timer = QTimer(self)
        self.search_timer.setSingleShot(True)
        self.search_timer.timeout.connect(self.search_documents)
        # Đảm bảo menubar nằm trên cùng
        self.setMenuBar(self.menuBar())
        
        # Set style cho window để menubar nằm ngang
        self.setStyleSheet("""
            QMainWindow {
                padding: 0px;
                margin: 0px;
            }
        """)
        # Global style adjustments
        self.setStyleSheet("""
            QGroupBox {
                border: 1px solid #cccccc;
                border-radius: 3px;
                margin-top: 6px;
                padding-top: 6px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 7px;
                padding: 0 3px;
            }
            QTableWidget {
                border: none;
            }
            QTabWidget::pane {
                border: 1px solid #cccccc;
            }
            QHeaderView::section {
                padding: 4px;
                border: none;
                border-bottom: 1px solid #cccccc;
                background-color: #f5f5f5;
            }
        """)

    def setup_menus(self):
        """Set up application menus"""
        menubar = self.menuBar()
        
        # File menu
        file_menu = menubar.addMenu("&File")
        
        add_file_action = QAction("Thêm File...", self)
        add_file_action.setShortcut("Ctrl+O")
        add_file_action.setStatusTip("Thêm file PDF mới")
        add_file_action.triggered.connect(self.add_file)
        file_menu.addAction(add_file_action)
        
        add_files_action = QAction("Thêm Nhiều File...", self)
        add_files_action.setShortcut("Ctrl+Shift+O")
        add_files_action.setStatusTip("Thêm nhiều file PDF cùng lúc")
        add_files_action.triggered.connect(self.add_files)
        file_menu.addAction(add_files_action)
        
        file_menu.addSeparator()
        
        export_excel_action = QAction("Xuất Excel...", self)
        export_excel_action.setShortcut("Ctrl+E")
        export_excel_action.setStatusTip("Xuất dữ liệu ra file Excel")
        export_excel_action.triggered.connect(self.export_excel)
        file_menu.addAction(export_excel_action)
        
        file_menu.addSeparator()
        
        backup_db_action = QAction("Sao lưu Database...", self)
        backup_db_action.setStatusTip("Tạo bản sao lưu của cơ sở dữ liệu")
        backup_db_action.triggered.connect(self.backup_database)
        file_menu.addAction(backup_db_action)
        
        file_menu.addSeparator()
        
        exit_action = QAction("E&xit", self)
        exit_action.setShortcut("Alt+F4")
        exit_action.setStatusTip("Thoát ứng dụng")
        exit_action.triggered.connect(self.close)
        file_menu.addAction(exit_action)
        
        # Edit menu
        edit_menu = menubar.addMenu("&Edit")
        
        save_action = QAction("Lưu thay đổi", self)
        save_action.setShortcut("Ctrl+S")
        save_action.setStatusTip("Lưu thay đổi cho tài liệu hiện tại")
        save_action.triggered.connect(self.save_current)
        edit_menu.addAction(save_action)
        
        reset_action = QAction("Đặt lại", self)
        reset_action.setShortcut("Ctrl+Z")
        reset_action.setStatusTip("Đặt lại thay đổi chưa lưu")
        reset_action.triggered.connect(self.ocr_editor.reset_fields)
        edit_menu.addAction(reset_action)
        
        edit_menu.addSeparator()
        
        delete_action = QAction("Xóa tài liệu", self)
        delete_action.setShortcut("Delete")
        delete_action.setStatusTip("Xóa tài liệu đã chọn")
        delete_action.triggered.connect(self.delete_current_document)
        edit_menu.addAction(delete_action)
        
        # View menu
        view_menu = menubar.addMenu("&View")
        
        toggle_preview_action = QAction("Xem trước", self)
        toggle_preview_action.setShortcut("Ctrl+P")
        toggle_preview_action.setStatusTip("Xem trước tài liệu")
        toggle_preview_action.triggered.connect(self.toggle_preview)
        view_menu.addAction(toggle_preview_action)
        
        view_menu.addSeparator()
        
        dark_mode_action = QAction("Dark Mode", self)
        dark_mode_action.setShortcut("Ctrl+D")
        dark_mode_action.setCheckable(True)
        dark_mode_action.setChecked(self.theme_manager.dark_mode)
        dark_mode_action.setStatusTip("Chuyển đổi chế độ tối")
        dark_mode_action.triggered.connect(self.toggle_dark_mode)
        view_menu.addAction(dark_mode_action)
        
        refresh_action = QAction("Refresh", self)
        refresh_action.setShortcut("F5")
        refresh_action.setStatusTip("Làm mới dữ liệu")
        refresh_action.triggered.connect(self.refresh_view)
        view_menu.addAction(refresh_action)
        
        # Tools menu
        tools_menu = menubar.addMenu("&Tools")
        
        settings_action = QAction("Cài đặt...", self)
        settings_action.setStatusTip("Mở cài đặt ứng dụng")
        settings_action.triggered.connect(self.show_settings)
        tools_menu.addAction(settings_action)
        
        check_files_action = QAction("Kiểm tra tệp tin...", self)
        check_files_action.setStatusTip("Kiểm tra và sửa chữa đường dẫn tệp tin")
        check_files_action.triggered.connect(self.check_file_paths)
        tools_menu.addAction(check_files_action)
        
        tools_menu.addSeparator()
        
        stats_action = QAction("Thống kê...", self)
        stats_action.setStatusTip("Xem thống kê hệ thống")
        # Kết nối trực tiếp với lambda để debug
        stats_action.triggered.connect(lambda: self.debug_statistics())
        tools_menu.addAction(stats_action)
        
        # Help menu
        help_menu = menubar.addMenu("&Help")
        
        about_action = QAction("Về ứng dụng", self)
        about_action.setStatusTip("Thông tin về ứng dụng")
        about_action.triggered.connect(self.show_about)
        help_menu.addAction(about_action)
        
        help_action = QAction("Hướng dẫn sử dụng", self)
        help_action.setShortcut("F1")
        help_action.setStatusTip("Mở hướng dẫn sử dụng")
        help_action.triggered.connect(self.show_help)
        help_menu.addAction(help_action)

    def debug_statistics(self):
        """Debug function for statistics"""
        print("Opening statistics dialog...")  # Debug print
        try:
            from statistics_dialog import StatisticsDialog
            dialog = StatisticsDialog(self.db, self)
            print("Dialog created")  # Debug print
            dialog.exec_()
            print("Dialog closed")  # Debug print
        except Exception as e:
            print(f"Error in statistics: {str(e)}")  # Debug print
            logging.error(f"Statistics error: {str(e)}")
            QMessageBox.critical(self, "Error", f"Error showing statistics: {str(e)}")

    def setup_shortcuts(self):
        """Set up keyboard shortcuts"""
        # Navigation
        QShortcut(QKeySequence("Ctrl+F"), self, lambda: self.search_input.setFocus())
        QShortcut(QKeySequence("Alt+Left"), self, self.pdf_viewer.previous_page)
        QShortcut(QKeySequence("Alt+Right"), self, self.pdf_viewer.next_page)
        QShortcut(QKeySequence("Escape"), self, self.clear_selection)
        
        # Zoom controls
        QShortcut(QKeySequence("Ctrl++"), self, self.pdf_viewer.zoom_in)
        QShortcut(QKeySequence("Ctrl+-"), self, self.pdf_viewer.zoom_out)
        QShortcut(QKeySequence("Ctrl+0"), self, self.pdf_viewer.reset_view)
        
        # Rotation
        QShortcut(QKeySequence("Ctrl+R"), self, self.pdf_viewer.rotate_right)
        QShortcut(QKeySequence("Ctrl+Shift+R"), self, self.pdf_viewer.rotate_left)

    def convert_word_to_pdf(self, word_path):
        """Convert Word document to PDF"""
        try:
            # Show progress
            self.statusBar().showMessage(f"Converting {os.path.basename(word_path)} to PDF...", 0)
            QApplication.processEvents()
            
            # Create output PDF path
            file_base = os.path.splitext(os.path.basename(word_path))[0]
            pdf_path = str(TEMP_DIR / f"{file_base}_{int(time.time())}.pdf")
            
            # Method 1: Using comtypes (Windows)
            if sys.platform == 'win32':
                try:
                    import comtypes.client
                    
                    word = comtypes.client.CreateObject('Word.Application')
                    word.Visible = False
                    
                    doc = word.Documents.Open(word_path)
                    doc.SaveAs(pdf_path, FileFormat=17)  # 17 = PDF format
                    doc.Close()
                    word.Quit()
                    
                    self.statusBar().showMessage(f"Conversion completed: {os.path.basename(pdf_path)}", 3000)
                    return pdf_path
                except ImportError:
                    logger.warning("comtypes not available, trying alternative method")
                except Exception as e:
                    logger.error(f"COM conversion failed: {str(e)}")
            
            # Method 2: Using LibreOffice (cross-platform)
            try:
                import subprocess
                # Check if LibreOffice is available
                if sys.platform == 'win32':
                    # Windows: Look for LibreOffice in common installation paths
                    libreoffice_paths = [
                        r"C:\Program Files\LibreOffice\program\soffice.exe",
                        r"C:\Program Files (x86)\LibreOffice\program\soffice.exe",
                    ]
                    soffice = next((path for path in libreoffice_paths if os.path.exists(path)), None)
                else:
                    # Linux/Mac: Use command available in PATH
                    soffice = "soffice"
                    
                if soffice:
                    cmd = [
                        soffice,
                        '--headless',
                        '--convert-to', 'pdf',
                        '--outdir', str(TEMP_DIR),
                        word_path
                    ]
                    process = subprocess.Popen(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
                    output, error = process.communicate()
                    
                    if process.returncode == 0:
                        # LibreOffice places output in the outdir with original name but .pdf extension
                        converted_path = os.path.join(TEMP_DIR, os.path.splitext(os.path.basename(word_path))[0] + '.pdf')
                        if os.path.exists(converted_path):
                            # Rename to our desired path to avoid conflicts
                            os.rename(converted_path, pdf_path)
                            self.statusBar().showMessage(f"Conversion completed: {os.path.basename(pdf_path)}", 3000)
                            return pdf_path
                    else:
                        logger.error(f"LibreOffice conversion failed: {error.decode('utf-8')}")
            except Exception as e:
                logger.error(f"LibreOffice conversion failed: {str(e)}")
                
            # Method 3: Using python-docx2pdf (if available)
            try:
                from docx2pdf import convert
                convert(word_path, pdf_path)
                if os.path.exists(pdf_path):
                    self.statusBar().showMessage(f"Conversion completed: {os.path.basename(pdf_path)}", 3000)
                    return pdf_path
            except ImportError:
                logger.warning("docx2pdf not available, trying next method")
            except Exception as e:
                logger.error(f"docx2pdf conversion failed: {str(e)}")
            
            # If all conversions failed
            QMessageBox.critical(
                self, 
                "Conversion Failed", 
                f"Failed to convert {os.path.basename(word_path)} to PDF.\n\n"
                "Please ensure you have either Microsoft Word, LibreOffice, "
                "or the docx2pdf Python package installed."
            )
            self.statusBar().showMessage("Conversion failed", 3000)
            return None
            
        except Exception as e:
            logger.error(f"Word to PDF conversion error: {str(e)}")
            QMessageBox.critical(self, "Conversion Error", str(e))
            return None

    def show_statistics(self):
        """Hiển thị dialog thống kê"""
        try:
            # Kiểm tra nếu có kết nối database
            if not self.db:
                QMessageBox.warning(
                    self,
                    "Lỗi",
                    "Không thể kết nối cơ sở dữ liệu. Vui lòng thử lại."
                )
                return

            # Tạo và hiển thị dialog thống kê
            dialog = StatisticsDialog(db=self.db, parent=self)
            dialog.exec_()

        except Exception as e:
            logger.error(f"Error showing statistics: {str(e)}")
            QMessageBox.critical(
                self,
                "Lỗi",
                f"Không thể hiển thị thống kê: {str(e)}"
            )

    def setup_statusbar(self):
        """Set up the status bar"""
        statusbar = self.statusBar()
        
        # Document count
        self.doc_count_label = QLabel("Documents: 0")
        statusbar.addPermanentWidget(self.doc_count_label)
        
        # Current document info
        self.doc_info_label = QLabel("")
        statusbar.addPermanentWidget(self.doc_info_label)
    def show_document_context_menu(self, position):
        """Show context menu for document table"""
        item = self.doc_table.itemAt(position)
        if item is None:
            return
            
        row = item.row()
        try:
            doc_id = int(self.doc_table.item(row, 0).text())
            
            context_menu = QMenu(self)
            
            # Actions for selected document
            open_action = context_menu.addAction("Mở PDF")
            open_action.setIcon(QIcon.fromTheme("document-open"))
            
            preview_action = context_menu.addAction("Xem trước")
            preview_action.setIcon(QIcon.fromTheme("document-preview"))
            
            export_action = context_menu.addAction("Xuất thông tin")
            export_action.setIcon(QIcon.fromTheme("document-export"))
            
            context_menu.addSeparator()
            
            # Tags submenu
            tags_menu = context_menu.addMenu("Tags")
            tags_menu.setIcon(QIcon.fromTheme("tag"))
            
            # Get current tags
            current_tags = self.db.get_document_tags(doc_id)
            all_tags = self.db.get_all_tags()
            
            # Add actions for existing tags
            for tag in all_tags:
                tag_action = tags_menu.addAction(tag)
                tag_action.setCheckable(True)
                tag_action.setChecked(tag in current_tags)
                
            # New tag option
            if all_tags:
                tags_menu.addSeparator()
            new_tag_action = tags_menu.addAction("Add New Tag...")
            
            context_menu.addSeparator()
            
            backup_action = context_menu.addAction("Tạo backup")
            backup_action.setIcon(QIcon.fromTheme("document-save-as"))
            
            context_menu.addSeparator()
            
            delete_action = context_menu.addAction("Xóa")
            delete_action.setIcon(QIcon.fromTheme("edit-delete"))
            
            # Show menu and handle action
            action = context_menu.exec_(self.doc_table.viewport().mapToGlobal(position))
            
            if action is None:
                return
                
            if action == open_action:
                self.open_document_pdf(doc_id)
            elif action == preview_action:
                self.preview_document(doc_id)
            elif action == export_action:
                self.export_document(doc_id)
            elif action == backup_action:
                self.backup_document(doc_id)
            elif action == delete_action:
                self.delete_document(doc_id)
            elif action == new_tag_action:
                self.add_new_tag(doc_id)
            elif action in [a for a in tags_menu.actions() if a != new_tag_action]:
                # Handle tag toggling
                tag_name = action.text()
                if action.isChecked():
                    self.db.add_tag(doc_id, tag_name)
                else:
                    self.db.remove_tag(doc_id, tag_name)
                    
        except Exception as e:
            logger.error(f"Error showing context menu: {str(e)}")

    def show_gallery_context_menu(self, position):
        """Show context menu for gallery view"""
        item = self.doc_gallery.itemAt(position)
        if item is None:
            return
            
        # Get document ID from item data
        doc_id = item.data(Qt.UserRole)
        if not doc_id:
            return
            
        # Create and show same context menu as document table
        try:
            context_menu = QMenu(self)
            
            # Actions for selected document
            open_action = context_menu.addAction("Mở PDF")
            open_action.setIcon(QIcon.fromTheme("document-open"))
            
            preview_action = context_menu.addAction("Xem trước")
            preview_action.setIcon(QIcon.fromTheme("document-preview"))
            
            export_action = context_menu.addAction("Xuất thông tin")
            export_action.setIcon(QIcon.fromTheme("document-export"))
            
            context_menu.addSeparator()
            
            # Tags submenu (same as in document context menu)
            # ...code omitted for brevity, identical to document context menu...
            
            context_menu.addSeparator()
            
            backup_action = context_menu.addAction("Tạo backup")
            backup_action.setIcon(QIcon.fromTheme("document-save-as"))
            
            context_menu.addSeparator()
            
            delete_action = context_menu.addAction("Xóa")
            delete_action.setIcon(QIcon.fromTheme("edit-delete"))
            
            # Show menu and handle action
            action = context_menu.exec_(self.doc_gallery.viewport().mapToGlobal(position))
            
            if action is None:
                return
                
            if action == open_action:
                self.open_document_pdf(doc_id)
            elif action == preview_action:
                self.preview_document(doc_id)
            elif action == export_action:
                self.export_document(doc_id)
            elif action == backup_action:
                self.backup_document(doc_id)
            elif action == delete_action:
                self.delete_document(doc_id)
                
        except Exception as e:
            logger.error(f"Error showing gallery context menu: {str(e)}")

    def add_new_tag(self, doc_id):
        """Add a new tag to the document"""
        tag, ok = QInputDialog.getText(
            self, 
            "New Tag", 
            "Enter tag name:",
            QLineEdit.Normal
        )
        
        if ok and tag:
            self.db.add_tag(doc_id, tag)

    def open_document_pdf(self, doc_id):
        """Open the document's PDF in default application"""
        try:
            doc_info = self.db.get_document_info(doc_id)
            if doc_info and os.path.exists(doc_info[1]):
                QDesktopServices.openUrl(QUrl.fromLocalFile(doc_info[1]))
            else:
                QMessageBox.warning(self, "File not found", "PDF file could not be found.")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error opening PDF: {str(e)}")

    def preview_document(self, doc_id):
        """Preview document in a dialog"""
        try:
            doc_info = self.db.get_document_info(doc_id)
            if doc_info and os.path.exists(doc_info[1]):
                preview = PreviewDialog(doc_info[1], self)
                preview.exec_()
            else:
                QMessageBox.warning(self, "File not found", "PDF file could not be found.")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error previewing document: {str(e)}")

    def export_document(self, doc_id):
        """Export single document to Excel"""
        try:
            file_path, _ = QFileDialog.getSaveFileName(
                self, "Save Document", "", "Excel Files (*.xlsx)"
            )
            if file_path:
                if not file_path.endswith('.xlsx'):
                    file_path += '.xlsx'
                
                # Export just this document
                self.db.export_to_excel(file_path, {'id': doc_id})
                QMessageBox.information(self, "Success", f"Document exported to {file_path}")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Export failed: {str(e)}")

    def backup_document(self, doc_id):
        """Create a backup of document PDF"""
        try:
            doc_info = self.db.get_document_info(doc_id)
            if doc_info and os.path.exists(doc_info[1]):
                backup_path = self.db.create_backup(doc_id, doc_info[1], "Manual backup")
                if backup_path:
                    QMessageBox.information(
                        self, 
                        "Backup Created", 
                        f"Backup created at:\n{backup_path}"
                    )
            else:
                QMessageBox.warning(self, "File not found", "PDF file could not be found.")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Backup failed: {str(e)}")

    def delete_document(self, doc_id):
        """Delete a document after confirmation"""
        try:
            reply = QMessageBox.question(
                self, 'Confirm Delete',
                'Bạn có chắc chắn muốn xóa tài liệu này không?\nThao tác này sẽ xóa cả file PDF.',
                QMessageBox.Yes | QMessageBox.No | QMessageBox.Cancel,
                QMessageBox.No
            )
            
            if reply == QMessageBox.Yes:
                self.db.delete_document(doc_id)
                self.load_documents()
                if self.current_doc_id == doc_id:
                    self.clear_current_view()
                    self.current_doc_id = None
                QMessageBox.information(self, "Success", "Document deleted successfully")
                
            elif reply == QMessageBox.No:
                # Ask if user wants to keep the file
                reply2 = QMessageBox.question(
                    self, 'Keep File?',
                    'Bạn có muốn giữ lại file PDF trên ổ đĩa không?',
                    QMessageBox.Yes | QMessageBox.No,
                    QMessageBox.Yes
                )
                
                keep_file = (reply2 == QMessageBox.Yes)
                self.db.delete_document(doc_id, keep_file=keep_file)
                
                self.load_documents()
                if self.current_doc_id == doc_id:
                    self.clear_current_view()
                    self.current_doc_id = None
                QMessageBox.information(self, "Success", "Document deleted from database")
                
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error deleting document: {str(e)}")

    def delete_current_document(self):
        """Delete currently selected document"""
        if self.current_doc_id:
            self.delete_document(self.current_doc_id)
        else:
            QMessageBox.information(self, "No selection", "Please select a document first")

    def add_file(self):
        """Add a single document file (PDF, DOC, DOCX)"""
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Select Document File", "", 
            "Document Files (*.pdf *.doc *.docx);;PDF Files (*.pdf);;Word Files (*.doc *.docx)"
        )
        if file_path:
            # Check if file is Word document
            if file_path.lower().endswith(('.doc', '.docx')):
                pdf_path = self.convert_word_to_pdf(file_path)
                if not pdf_path:
                    return
            else:
                pdf_path = file_path
                
            preview = PreviewDialog(pdf_path, self)
            if preview.exec_() == QDialog.Accepted:
                self.process_file(pdf_path)

    def add_files(self):
        """Add multiple document files"""
        file_paths, _ = QFileDialog.getOpenFileNames(
            self, "Select Document Files", "", 
            "Document Files (*.pdf *.doc *.docx);;PDF Files (*.pdf);;Word Files (*.doc *.docx)"
        )
        if file_paths:
            # Convert Word documents to PDF first
            pdf_paths = []
            for file_path in file_paths:
                if file_path.lower().endswith(('.doc', '.docx')):
                    pdf_path = self.convert_word_to_pdf(file_path)
                    if pdf_path:
                        pdf_paths.append(pdf_path)
                else:
                    pdf_paths.append(file_path)
                    
            if pdf_paths:
                self.process_files(pdf_paths)

    def is_converted_pdf(self, pdf_path):
        """Check if a PDF was converted from Word (exists in temp folder)"""
        # Thêm điều kiện kiểm tra kỹ hơn
        return str(TEMP_DIR) in str(pdf_path) and "temp" in str(pdf_path).lower()

    def process_file(self, file_path):
        """Process a single PDF file"""
        progress = ProgressDialog(100, "Processing PDF", self)
        progress.setWindowModality(Qt.ApplicationModal)
        
        self.ocr_worker = OCRWorker(self.ocr_system, file_path)
        self.ocr_worker.progress.connect(progress.update_progress)
        self.ocr_worker.finished.connect(lambda results, detections: 
                                      self.ocr_completed(file_path, results, detections))
        self.ocr_worker.error.connect(self.show_error)
        
        # Connect cancel signal
        progress.canceled.connect(self.cancel_ocr)
        
        self.ocr_worker.start()
        progress.exec_()

    def process_files(self, file_paths):
        """Process multiple PDF files"""
        progress = ProgressDialog(len(file_paths), "Processing PDFs", self)
        progress.setWindowModality(Qt.ApplicationModal)
        
        self.batch_worker = BatchProcessWorker(self.ocr_system, file_paths)
        self.batch_worker.progress.connect(progress.update_progress)
        self.batch_worker.finished.connect(self.batch_completed)
        self.batch_worker.error.connect(self.show_error)
        
        # Connect cancel signal
        progress.canceled.connect(self.cancel_batch)
        
        self.batch_worker.start()
        progress.exec_()

    def cancel_ocr(self):
        """Cancel OCR process"""
        if self.ocr_worker and self.ocr_worker.isRunning():
            self.ocr_worker.cancel()
            self.ocr_worker.wait(1000)  # Wait up to 1 second
            if self.ocr_worker.isRunning():
                self.ocr_worker.terminate()

    def cancel_batch(self):
        """Cancel batch process"""
        if self.batch_worker and self.batch_worker.isRunning():
            self.batch_worker.cancel()
            self.batch_worker.wait(1000)  # Wait up to 1 second
            if self.batch_worker.isRunning():
                self.batch_worker.terminate()

    def ocr_completed(self, file_path, results, detections):
        """Handle completion of OCR process for a single file"""
        try:
            # Kiểm tra xem file có phải là file tạm
            is_temp_file = self.is_converted_pdf(file_path)
            
            # Nếu là file tạm, tạo bản sao vĩnh viễn
            permanent_file_path = file_path
            if is_temp_file:
                perm_filename = f"{Path(file_path).stem}_perm_{int(time.time())}.pdf"
                permanent_file_path = str(OUTPUT_DIR / perm_filename)
                shutil.copy2(file_path, permanent_file_path)
                logger.info(f"Created permanent copy: {permanent_file_path}")
            
            # Add document to database với đường dẫn vĩnh viễn
            doc_id = self.db.add_document(permanent_file_path, results)
            
            # Add detections for each page
            for page_num, page_detections in detections:
                if page_detections:
                    self.db.add_page_detections(doc_id, page_num, page_detections)
            
            # Refresh document list and show the new document
            self.load_documents()
            self.show_document(doc_id)
            
            # Add to recent files
            self.settings.setValue("last_directory", str(Path(permanent_file_path).parent))
            
            # Clean up temporary converted PDF if needed
            if is_temp_file and os.path.exists(file_path):
                try:
                    # Kiểm tra lại lần cuối để đảm bảo file vĩnh viễn tồn tại
                    if os.path.exists(permanent_file_path):
                        os.remove(file_path)
                        logger.info(f"Removed temporary PDF: {file_path}")
                    else:
                        logger.warning(f"Permanent file not created, keeping temp file: {file_path}")
                except Exception as e:
                    logger.warning(f"Could not remove temporary PDF: {str(e)}")
            
            QMessageBox.information(self, "Success", "PDF processing completed successfully!")
            
        except Exception as e:
            logger.error(f"Error in OCR completion: {str(e)}")
            QMessageBox.critical(self, "Error", f"Error saving results: {str(e)}")

    def batch_completed(self, results):
        """Handle completion of batch OCR process"""
        try:
            num_processed = 0
            for file_path, result, detections in results:
                try:
                    # Kiểm tra xem file có phải là file tạm
                    is_temp_file = self.is_converted_pdf(file_path)
                    
                    # Nếu là file tạm, tạo bản sao vĩnh viễn
                    if is_temp_file:
                        perm_filename = f"{Path(file_path).stem}_perm_{int(time.time())}.pdf"
                        permanent_file_path = str(OUTPUT_DIR / perm_filename)
                        shutil.copy2(file_path, permanent_file_path)
                        logger.info(f"Created permanent copy: {permanent_file_path}")
                        file_path = permanent_file_path
                    
                    # Add document to database
                    doc_id = self.db.add_document(file_path, result)
                    
                    # Add detections for each page
                    for page_num, page_detections in detections:
                        if page_detections:
                            self.db.add_page_detections(doc_id, page_num, page_detections)
                    
                    num_processed += 1
                    
                    # Clean up temporary converted PDF if needed
                    if is_temp_file and os.path.exists(file_path):
                        try:
                            os.remove(file_path)
                            logger.info(f"Removed temporary PDF: {file_path}")
                        except Exception as e:
                            logger.warning(f"Could not remove temporary PDF: {str(e)}")
                    
                except Exception as e:
                    logger.error(f"Error processing {file_path}: {str(e)}")
            
            # Refresh document list
            self.load_documents()
            
            if results:
                # Remember last directory
                self.settings.setValue("last_directory", str(Path(results[0][0]).parent))
            
            QMessageBox.information(
                self, 
                "Batch Processing Complete", 
                f"Successfully processed {num_processed} out of {len(results)} files."
            )
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error in batch processing: {str(e)}")

    def show_error(self, error_msg):
        """Display error message"""
        QMessageBox.critical(self, "Error", f"An error occurred: {error_msg}")

    def load_documents(self):
        """Load all documents from database"""
        try:
            documents = self.db.get_all_documents()
            self.update_document_table(documents)
            self.update_document_gallery(documents)
            
            # Update document count
            doc_count = len(documents)
            self.doc_count_label.setText(f"Documents: {doc_count}")
            
        except Exception as e:
            logger.error(f"Error loading documents: {str(e)}")
            QMessageBox.critical(self, "Error", f"Error loading documents: {str(e)}")

    def update_document_table(self, documents):
        """Update document table with data"""
        self.doc_table.setSortingEnabled(False)
        self.doc_table.setRowCount(len(documents))
        
        for row, doc in enumerate(documents):
            # ID
            id_item = QTableWidgetItem(str(doc[0]))
            id_item.setData(Qt.UserRole, doc[0])  # Store ID for reference
            self.doc_table.setItem(row, 0, id_item)
            
            # Filename
            filename_item = QTableWidgetItem(doc[2])
            filename_item.setToolTip(doc[1])  # Full path as tooltip
            self.doc_table.setItem(row, 1, filename_item)
            
            # Date created
            date_item = QTableWidgetItem(str(doc[3]))
            self.doc_table.setItem(row, 2, date_item)
            
            # Version count
            version_item = QTableWidgetItem(str(doc[5]))
            version_item.setTextAlignment(Qt.AlignCenter)
            self.doc_table.setItem(row, 3, version_item)
            
            # Document number
            number_item = QTableWidgetItem(str(doc[8] or ""))
            self.doc_table.setItem(row, 4, number_item)
            
            # Document urgency
            urgency_text = doc[10] or "Không"
            urgency_item = QTableWidgetItem(urgency_text)
            urgency_item.setTextAlignment(Qt.AlignCenter)
            
            # Color code urgency
            if urgency_text == "Độ Mật":
                urgency_item.setBackground(QColor(255, 200, 200))  # Light red
            elif urgency_text == "Hỏa Tốc":
                urgency_item.setBackground(QColor(255, 140, 0))  # Orange
                urgency_item.setForeground(QColor(255, 255, 255))  # White text
                
            self.doc_table.setItem(row, 5, urgency_item)
            
            # Page count
            page_count_item = QTableWidgetItem(str(doc[4] or ""))
            page_count_item.setTextAlignment(Qt.AlignCenter)
            self.doc_table.setItem(row, 6, page_count_item)
            
        self.doc_table.setSortingEnabled(True)
    def update_document_gallery(self, documents):
        """Update document gallery with thumbnails"""
        self.doc_gallery.clear()
        
        # Calculate number of rows needed (with 4 columns)
        num_rows = (len(documents) + 3) // 4
        self.doc_gallery.setRowCount(num_rows)
        
        # Add documents to gallery
        for i, doc in enumerate(documents):
            # Calculate row and column
            row = i // 4
            col = i % 4
            
            # Create thumbnail item
            item = QTableWidgetItem()
            item.setData(Qt.UserRole, doc[0])  # Store document ID
            
            # Set item text (document info)
            doc_info = f"{doc[2]}\nID: {doc[0]}"
            if doc[8]:  # So ky hieu
                doc_info += f"\n{doc[8]}"
            if doc[10]:  # Do khan
                doc_info += f"\n{doc[10]}"
                
            item.setText(doc_info)
            item.setTextAlignment(Qt.AlignCenter)
            
            # Load thumbnail if available or generate
            self.load_thumbnail_for_item(item, doc[1], doc[0])
            
            # Set item size
            self.doc_gallery.setItem(row, col, item)
            self.doc_gallery.setRowHeight(row, 200)
            
    def load_thumbnail_for_item(self, item, file_path, doc_id):
        """Load or generate thumbnail for gallery item"""
        # Check for cached thumbnail
        thumb_path = IMAGES_DIR / f"thumb_{doc_id}.jpg"
        
        if os.path.exists(thumb_path):
            # Load existing thumbnail
            pixmap = QPixmap(str(thumb_path))
            item.setIcon(QIcon(pixmap))
            return
            
        # Generate thumbnail if file exists
        if os.path.exists(file_path):
            try:
                # Open first page of PDF
                with fitz.open(file_path) as pdf:
                    if len(pdf) > 0:
                        page = pdf[0]
                        # Render page to pixmap
                        pix = page.get_pixmap(matrix=fitz.Matrix(0.2, 0.2))
                        
                        # Convert to QImage and save thumbnail
                        img = QImage(pix.samples, pix.width, pix.height, pix.stride, QImage.Format_RGB888)
                        img.save(str(thumb_path), "JPG")
                        
                        # Set item icon
                        item.setIcon(QIcon(QPixmap.fromImage(img)))
                        return
            except Exception as e:
                logger.error(f"Error generating thumbnail: {str(e)}")
                
        # Use default icon if thumbnail generation failed
        item.setIcon(QIcon.fromTheme("application-pdf"))

    def gallery_item_selected(self, item):
        """Handle selection in gallery view"""
        doc_id = item.data(Qt.UserRole)
        if doc_id:
            self.show_document(doc_id)

    def show_document(self, doc_id):
        """Display a document and its data"""
        try:
            # Get document data
            doc_version = self.db.get_latest_version(doc_id)
            doc_info = self.db.get_document_info(doc_id)
            
            if not doc_version or not doc_info:
                QMessageBox.warning(self, "Warning", "Document data not found")
                return
                    
            # Store current document ID
            self.current_doc_id = doc_id  # Đảm bảo lưu đúng ID hiện tại
            
            # Load document content in editor FIRST
            self.ocr_editor.load_data(doc_version)
            self.ocr_editor.doc_id = doc_id  # Đặt lại ID một cách rõ ràng
            self.ocr_editor.update_suggestions(self.db)
            
            # Load document versions
            versions = self.db.get_document_versions(doc_id)
            self.update_version_list(versions)
            
            # Update metadata display
            self.update_metadata_display(doc_info, doc_version)
            
            # Load PDF file
            pdf_path = doc_info[1]
            if os.path.exists(pdf_path):
                if self.pdf_viewer.load_pdf(pdf_path):
                    self.load_page_detections(doc_id, self.pdf_viewer.current_page)
                else:
                    self.detections_by_page = {}
            else:
                QMessageBox.warning(self, "Warning", f"PDF file not found: {pdf_path}")
                self.pdf_viewer.clear()
                self.detections_by_page = {}
                    
            # Update status bar
            self.update_status_info(doc_id, doc_info)
                    
        except Exception as e:
            logger.error(f"Error showing document: {str(e)}")
            traceback.print_exc()
            QMessageBox.warning(self, "Error", f"Error loading document: {str(e)}")

    def load_page_detections(self, doc_id, page_num):
        """Load detections for a specific page"""
        if not doc_id or page_num is None or not hasattr(self.pdf_viewer, 'set_detection_boxes'):
            return
            
        try:
            detections = self.db.get_document_detections(doc_id, page_num)
            
            if isinstance(detections, list):  # Ensure detections is a list
                boxes = []
                for d in detections:
                    if isinstance(d, dict) and 'box' in d:
                        box = d.get('box', [])
                        if len(box) >= 4:
                            boxes.append(box)
                
                self.pdf_viewer.set_detection_boxes(boxes)
            else:
                self.pdf_viewer.set_detection_boxes([])
                
        except Exception as e:
            print(f"Error loading page detections: {str(e)}")  
            self.pdf_viewer.set_detection_boxes([])

    def update_metadata_display(self, doc_info, doc_version):
        """Update metadata information display"""
        try:
            if not doc_info or not doc_version:
                for field in self.metadata_fields.values():
                    field.clear()
                return
                    
            # Basic info
            self.metadata_fields['id'].setText(str(doc_info[0]))
            self.metadata_fields['filename'].setText(doc_info[2])
            self.metadata_fields['created'].setText(str(doc_info[3]))
            self.metadata_fields['modified'].setText(str(doc_info[4]))
            
            # Fix: Hiển thị số trang đúng
            # page_count ở vị trí index 7 trong bảng documents
            if len(doc_info) > 7 and doc_info[7] is not None:
                self.metadata_fields['pages'].setText(str(doc_info[7]))
            else:
                # Nếu không có thông tin số trang trong DB, thử đọc từ file PDF
                file_path = doc_info[1]
                if os.path.exists(file_path) and file_path.lower().endswith('.pdf'):
                    try:
                        with fitz.open(file_path) as pdf:
                            page_count = len(pdf)
                            self.metadata_fields['pages'].setText(str(page_count))
                            
                            # Cập nhật page_count vào DB
                            if self.current_doc_id:
                                try:
                                    conn = self.db.conn_pool.get_connection()
                                    cursor = conn.cursor()
                                    cursor.execute(
                                        'UPDATE documents SET page_count = ? WHERE id = ?',
                                        (page_count, self.current_doc_id)
                                    )
                                    conn.commit()
                                except Exception as e:
                                    logger.error(f"Error updating page count: {str(e)}")
                    except:
                        self.metadata_fields['pages'].setText("1")
                else:
                    self.metadata_fields['pages'].setText("N/A")
            
            # File size
            file_path = doc_info[1]
            if os.path.exists(file_path):
                size_bytes = os.path.getsize(file_path)
                if size_bytes < 1024:
                    self.metadata_fields['filesize'].setText(f"{size_bytes} B")
                elif size_bytes < 1024 * 1024:
                    self.metadata_fields['filesize'].setText(f"{size_bytes/1024:.1f} KB")
                else:
                    self.metadata_fields['filesize'].setText(f"{size_bytes/(1024*1024):.1f} MB")
            else:
                self.metadata_fields['filesize'].setText("File not found")
                    
            # Version count
            versions = self.db.get_document_versions(doc_info[0])
            self.metadata_fields['versions'].setText(str(len(versions)))
            
            # Document content fields
            self.metadata_fields['cqbh'].setText(f"{doc_version[3] or ''}\n{doc_version[4] or ''}")
            self.metadata_fields['so_kh'].setText(doc_version[5] or '')
            self.metadata_fields['loai_vb'].setText(doc_version[6] or '')
            
            # Độ khẩn with color coding
            do_khan = doc_version[12] or 'Không'
            self.metadata_fields['do_khan'].setText(do_khan)
            
            if do_khan == 'Độ Mật':
                self.metadata_fields['do_khan'].setStyleSheet("color: darkred; font-weight: bold;")
            elif do_khan == 'Hỏa Tốc':
                self.metadata_fields['do_khan'].setStyleSheet("color: orangered; font-weight: bold;")
            else:
                self.metadata_fields['do_khan'].setStyleSheet("")
                    
        except Exception as e:
            logger.error(f"Error updating metadata: {str(e)}")
            for field in self.metadata_fields.values():
                field.clear()

    def update_status_info(self, doc_id, doc_info):
        """Update status bar with document info"""
        if doc_info and os.path.exists(doc_info[1]):
            file_info = f"ID: {doc_id} | {os.path.basename(doc_info[1])}"
            self.doc_info_label.setText(file_info)
        else:
            self.doc_info_label.setText("")

    def document_selected(self, item):
        """Handle document selection in table"""
        try:
            row = item.row()
            doc_id = int(self.doc_table.item(row, 0).text())
            self.show_document(doc_id)
        except Exception as e:
            logger.error(f"Error selecting document: {str(e)}")

    def version_selected(self, item):
        """Handle version selection in table"""
        try:
            row = item.row()
            version_num = int(self.version_list.item(row, 0).text())
            
            if not self.current_doc_id:
                return
                
            # Load the specific version
            version_data = self.db.get_document_version(self.current_doc_id, version_num)
            if version_data:
                self.ocr_editor.load_data(version_data)
                self.update_metadata_display(self.db.get_document_info(self.current_doc_id), version_data)
                
        except Exception as e:
            logger.error(f"Error loading version: {str(e)}")

    def update_version_list(self, versions):
        """Update version list with document versions"""
        self.version_list.setRowCount(len(versions))
        for row, version in enumerate(versions):
            # Version number
            version_item = QTableWidgetItem(str(version[2]))
            self.version_list.setItem(row, 0, version_item)
            
            # Modified by
            modified_by_item = QTableWidgetItem(version[13])
            self.version_list.setItem(row, 1, modified_by_item)
            
            # Date
            date_item = QTableWidgetItem(str(version[14]))
            self.version_list.setItem(row, 2, date_item)

    def update_document(self, doc_id, updates):
        """Update document with new data"""
        try:
            # Verify the document ID matches current selection
            if doc_id != self.current_doc_id:
                raise ValueError(f"Document ID mismatch: trying to update {doc_id} but current document is {self.current_doc_id}")
                
            # Create new version
            new_version = self.db.create_new_version(doc_id, updates)
            
            if new_version:
                # Refresh document list
                self.load_documents()
                
                # Reload the same document
                self.show_document(doc_id)
                
                # Update suggestions
                self.ocr_editor.update_suggestions(self.db)
                
                self.statusBar().showMessage(f"Document {doc_id} updated successfully", 5000)
            else:
                raise Exception("Failed to create new version")
                
        except Exception as e:
            logger.error(f"Error updating document: {str(e)}")
            QMessageBox.critical(self, "Error", f"Error updating document: {str(e)}")

    def field_value_changed(self, field_name, new_value):
        """Handle field value change for real-time updates"""
        # This can be used for real-time validation or UI updates
        # without waiting for save
        pass

    def search_documents(self):
        """Search documents based on query and search type"""
        query = self.search_input.text()
        search_type = self.search_type.currentText().lower()
        
        # Map UI search type to database search type
        type_mapping = {
            "tất cả": "all",
            "tên file": "file_name",
            "nội dung": "content",
            "số ký hiệu": "so_ki_hieu",
            "độ khẩn": "do_khan"
        }
        
        db_search_type = type_mapping.get(search_type, "all")
        
        try:
            documents = self.db.search_documents(query, db_search_type)
            self.update_document_table(documents)
            self.update_document_gallery(documents)
            
            self.statusBar().showMessage(
                f"Tìm thấy {len(documents)} tài liệu phù hợp với '{query}'", 5000
            )
            
        except Exception as e:
            logger.error(f"Search error: {str(e)}")
            self.statusBar().showMessage(f"Lỗi tìm kiếm: {str(e)}", 5000)

    def show_advanced_search(self):
        """Show advanced search dialog"""
        # This would implement a more complex search dialog
        # with multiple criteria
        pass

    def export_excel(self):
        """Export all documents to Excel with optional filtering"""
        try:
            # Get export file path
            file_path, _ = QFileDialog.getSaveFileName(
                self, "Export to Excel", "", "Excel Files (*.xlsx)"
            )
            
            if file_path:
                if not file_path.endswith('.xlsx'):
                    file_path += '.xlsx'
                
                # Get current search/filter criteria
                query = self.search_input.text()
                search_type = self.search_type.currentText().lower()
                
                filter_criteria = {}
                if query:
                    if search_type == "tên file":
                        filter_criteria['file_name'] = query
                    elif search_type == "số ký hiệu":
                        filter_criteria['so_ki_hieu'] = query
                    elif search_type == "độ khẩn":
                        filter_criteria['do_khan'] = query
                    else:
                        filter_criteria['text'] = query
                
                # Export with filters
                self.db.export_to_excel(file_path, filter_criteria)
                QMessageBox.information(self, "Success", f"Data exported to {file_path}")
            
        except Exception as e:
            logger.error(f"Export error: {str(e)}")
            QMessageBox.critical(self, "Error", f"Export failed: {str(e)}")

    def backup_database(self):
        """Create backup of the database"""
        try:
            backup_path, _ = QFileDialog.getSaveFileName(
                self, 
                "Backup Database", 
                str(BACKUP_DIR / f"documents_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.db"),
                "Database Files (*.db)"
            )
            
            if backup_path:
                if not backup_path.endswith('.db'):
                    backup_path += '.db'
                
                # Close database connections
                self.db.close()
                
                # Copy database file
                shutil.copy2(self.db.db_path, backup_path)
                
                # Reconnect
                self.db = DocumentDatabase()
                
                QMessageBox.information(
                    self, 
                    "Backup Complete", 
                    f"Database backed up to:\n{backup_path}"
                )
                
        except Exception as e:
            logger.error(f"Database backup error: {str(e)}")
            QMessageBox.critical(self, "Backup Error", f"Database backup failed: {str(e)}")

    def toggle_preview(self):
        """Toggle PDF preview for currently selected document"""
        if not self.current_doc_id:
            QMessageBox.information(self, "Information", "Please select a document first")
            return
            
        # Get document info
        doc_info = self.db.get_document_info(self.current_doc_id)
        if not doc_info:
            QMessageBox.warning(self, "Warning", "Document information not found")
            return
            
        # Check if file exists
        pdf_path = doc_info[1]
        if not os.path.exists(pdf_path):
            QMessageBox.warning(self, "File not found", f"PDF file not found at: {pdf_path}")
            return
            
        # Show preview dialog
        preview = PreviewDialog(pdf_path, self)
        preview.exec_()

    def toggle_dark_mode(self):
        """Toggle dark mode"""
        is_dark = self.theme_manager.toggle_theme(QApplication.instance())
        
        # Update menu checkmark
        for action in self.menuBar().actions():
            if action.text() == "&View":
                for subaction in action.menu().actions():
                    if subaction.text() == "Dark Mode":
                        subaction.setChecked(is_dark)
                        break
                break

    def switch_view(self, view_mode):
        """Switch between different view modes"""
        if view_mode == self.current_view_mode:
            return
            
        self.current_view_mode = view_mode
        
        if view_mode == "details":
            self.doc_tabs.setCurrentIndex(0)
        elif view_mode == "gallery":
            self.doc_tabs.setCurrentIndex(1)

    def page_changed(self, page_num):
        """Handle PDF page change"""
        if self.current_doc_id:
            self.load_page_detections(self.current_doc_id, page_num)

    def show_settings(self):
        """Show settings dialog"""
        # This would open a settings dialog where user can
        # configure application settings
        pass

    def show_statistics(self):
        """Show document statistics"""
        # This would show statistics about documents in the system
        pass

    def show_about(self):
        """Show about dialog"""
        QMessageBox.about(
            self,
            "About OCR Document Manager",
            f"<h2>OCR Document Manager</h2>"
            f"<p>Version: {APP_VERSION}</p>"
            f"<p>A document management system with OCR capabilities.</p>"
            f"<p>© 2025 All Rights Reserved</p>"
        )

    def show_help(self):
        """Show help documentation"""
        help_file = BASE_DIR / "help.html"
        if os.path.exists(help_file):
            QDesktopServices.openUrl(QUrl.fromLocalFile(str(help_file)))
        else:
            QMessageBox.information(
                self,
                "Help",
                "Hướng dẫn sử dụng:\n\n"
                "1. Thêm tài liệu: Sử dụng nút 'Thêm File' hoặc menu File > Thêm File\n"
                "2. Xử lý OCR: Hệ thống sẽ tự động trích xuất văn bản\n"
                "3. Chỉnh sửa kết quả: Sử dụng panel bên phải để chỉnh sửa\n"
                "4. Tìm kiếm: Nhập từ khóa vào ô tìm kiếm\n"
                "5. Xuất dữ liệu: Sử dụng chức năng Xuất Excel\n"
            )

    def refresh_view(self):
        """Refresh current view and reload data"""
        self.load_documents()
        
        # Get currently selected document
        if self.current_doc_id:
            self.show_document(self.current_doc_id)
        else:
            self.clear_current_view()
            
        self.statusBar().showMessage("Data refreshed", 3000)

    def clear_selection(self):
        """Clear current selection"""
        self.doc_table.clearSelection()
        self.version_list.clearSelection()
        self.current_doc_id = None
        self.clear_current_view()
        
    def clear_current_view(self):
        """Clear all views and editors"""
        self.pdf_viewer.clear()
        self.ocr_editor.clear()
        
        # Clear metadata
        for field in self.metadata_fields.values():
            field.clear()
            
        self.version_list.setRowCount(0)
        self.detections_by_page = {}
        self.doc_info_label.setText("")

    def save_current(self):
        """Save current document changes"""
        if self.current_doc_id:
            self.ocr_editor.save_changes()
        else:
            QMessageBox.information(self, "No selection", "Please select a document first")

    def auto_save(self):
        """Auto-save current document if changes detected"""
        if self.current_doc_id and self.ocr_editor.save_btn.isEnabled():
            self.ocr_editor.save_changes()
            self.statusBar().showMessage("Auto-saved", 2000)

    def closeEvent(self, event):
        """Handle application close"""
        # Save window settings
        self.save_window_settings()
        
        # Save theme settings
        self.theme_manager.save_theme_settings()
        
        # Check for unsaved changes
        if self.current_doc_id and self.ocr_editor.save_btn.isEnabled():
            reply = QMessageBox.question(
                self, 'Unsaved Changes',
                'You have unsaved changes. Do you want to save before exiting?',
                QMessageBox.Save | QMessageBox.Discard | QMessageBox.Cancel,
                QMessageBox.Save
            )
            
            if reply == QMessageBox.Save:
                self.ocr_editor.save_changes()
                event.accept()
            elif reply == QMessageBox.Discard:
                event.accept()
            else:
                event.ignore()
                return
        
        # Close database connections
        self.db.close()
        
        # Stop any running workers
        if self.ocr_worker and self.ocr_worker.isRunning():
            self.ocr_worker.cancel()
            self.ocr_worker.wait(1000)
            
        if self.batch_worker and self.batch_worker.isRunning():
            self.batch_worker.cancel()
            self.batch_worker.wait(1000)
            
        if self.repair_worker and self.repair_worker.isRunning():
            self.repair_worker.cancel()
            self.repair_worker.wait(1000)
            
        # Accept close event
        event.accept()

    def process_custom_box(self, rect, class_id):
        """Xử lý OCR cho box tùy chỉnh vừa được tạo"""
        try:
            from PIL import Image  # Import sớm để tránh lỗi tham chiếu trước khi gán
            import numpy as np
            
            # Hiển thị thông báo đang xử lý
            self.statusBar().showMessage("Đang xử lý OCR cho vùng được chọn...", 2000)
            QApplication.processEvents()  # Cho phép UI cập nhật
            
            # Lấy thông tin trang hiện tại
            if not self.pdf_viewer.pages or self.pdf_viewer.current_page >= len(self.pdf_viewer.pages):
                return
                
            current_pil_image = self.pdf_viewer.pages[self.pdf_viewer.current_page]
            
            # Áp dụng xoay nếu cần
            if self.pdf_viewer.rotation != 0:
                current_pil_image = current_pil_image.rotate(-self.pdf_viewer.rotation, expand=True)
            
            # Lấy kích thước ảnh gốc
            img_width, img_height = current_pil_image.size
            
            # Lưu thông tin debug tọa độ - lưu rect gốc trước khi điều chỉnh
            orig_rect = QRect(rect)
            
            # Đảm bảo tọa độ không vượt quá kích thước ảnh
            x = max(0, min(rect.x(), img_width - 1))
            y = max(0, min(rect.y(), img_height - 1))
            w = min(rect.width(), img_width - x)
            h = min(rect.height(), img_height - y)
            
            # Tạo rect đã điều chỉnh để lưu debug
            adjusted_rect = QRect(x, y, w, h)
            
            # Lưu thông tin tọa độ để debug
            self.debug_coords_to_file(orig_rect, adjusted_rect, (img_width, img_height), class_id)
            
            # Kiểm tra kích thước tối thiểu
            if w < 10 or h < 10:
                QMessageBox.warning(self, "Cảnh báo", "Vùng được chọn quá nhỏ để OCR. Vui lòng vẽ lại vùng lớn hơn.")
                return
            
            # Tính toán margin dựa vào loại nội dung
            margin_top = 5
            margin_bottom = 5
            margin_left = 5
            margin_right = 5
            
            # Điều chỉnh margin theo từng loại văn bản
            if class_id == 7:  # Noi_Nhan
                # Thường là danh sách nhiều dòng, cần margin lớn hơn
                margin_top = 10
                margin_bottom = 10
                margin_left = 15
                margin_right = 15
            elif class_id == 5:  # ND_Chinh
                # Nội dung chính cũng cần margin lớn
                margin_top = 10
                margin_bottom = 10
                margin_left = 10
                margin_right = 10
            
            # Thêm thông tin debug
            print(f"Vùng cắt gốc: x={x}, y={y}, w={w}, h={h}, img_size={img_width}x{img_height}")
            
            # Tạo thư mục debug
            debug_dir = TEMP_DIR / "debug"
            debug_dir.mkdir(exist_ok=True)
            timestamp = int(time.time())
            
            # Lưu ảnh gốc để debug
            debug_orig_path = debug_dir / f"original_full_page_{timestamp}.png"
            current_pil_image.save(str(debug_orig_path))
            print(f"Đã lưu ảnh trang gốc tại: {debug_orig_path}")
            
            # Vẽ rect lên ảnh gốc để kiểm tra
            debug_img = current_pil_image.copy()
            draw = ImageDraw.Draw(debug_img)
            draw.rectangle((x, y, x + w, y + h), outline="red", width=3)
            debug_rect_path = debug_dir / f"original_with_rect_{class_id}_{timestamp}.png"
            debug_img.save(str(debug_rect_path))
            print(f"Đã lưu ảnh có khung tại: {debug_rect_path}")
            
            # Cắt vùng ảnh với margin để đảm bảo không cắt mất chữ
            x_with_margin = max(0, x - margin_left)
            y_with_margin = max(0, y - margin_top)
            w_with_margin = min(w + margin_left + margin_right, img_width - x_with_margin)
            h_with_margin = min(h + margin_top + margin_bottom, img_height - y_with_margin)
            
            # Lưu tọa độ cắt để debug
            crop_coords = (x_with_margin, y_with_margin, x_with_margin + w_with_margin, y_with_margin + h_with_margin)
            print(f"Tọa độ cắt sau khi thêm margin: {crop_coords}")
            
            # Vẽ rect với margin lên ảnh gốc để kiểm tra
            debug_img_margin = current_pil_image.copy()
            draw_margin = ImageDraw.Draw(debug_img_margin)
            draw_margin.rectangle(crop_coords, outline="blue", width=3)
            debug_rect_margin_path = debug_dir / f"original_with_margin_rect_{class_id}_{timestamp}.png"
            debug_img_margin.save(str(debug_rect_margin_path))
            print(f"Đã lưu ảnh có khung margin tại: {debug_rect_margin_path}")
            
            # Cắt vùng ảnh
            region = current_pil_image.crop(crop_coords)
            
            # Lưu ảnh vùng cắt để debug
            debug_dir = TEMP_DIR / "debug"
            debug_dir.mkdir(exist_ok=True)
            timestamp = int(time.time())
            debug_path = debug_dir / f"region_class_{class_id}_{timestamp}.png"
            region.save(str(debug_path))
            print(f"Đã lưu ảnh debug tại: {debug_path}")
            
            # Kiểm tra kích thước ảnh cắt
            if region.size[0] < 10 or region.size[1] < 10:
                QMessageBox.warning(self, "Cảnh báo", "Vùng được chọn quá nhỏ để OCR. Vui lòng vẽ lại vùng lớn hơn.")
                return
            
            # Thực hiện OCR với class_id đã chọn
            class_names = {
                0: 'CQBH', 1: 'Chu_Ky', 2: 'Chuc_Vu', 3: 'Do_Khan',
                4: 'Loai_VB', 5: 'ND_Chinh', 6: 'Ngay_BH', 7: 'Noi_Nhan', 8: 'So_Ki_Hieu'
            }
            class_name = class_names.get(class_id, '')
            
            # Thử cả hai phương pháp OCR, ưu tiên kết quả tốt hơn
            final_text = ""
            
            # 1. Thử với EasyOCR trước
            easyocr_success = False
            try:
                # Tiền xử lý ảnh tối ưu cho loại class
                processed_region = self.ocr_system.preprocess_image_for_document(region, class_id)
                
                # Tạo PIL Image từ mảng đã xử lý nếu cần
                if not isinstance(processed_region, Image.Image):
                    processed_region = Image.fromarray(processed_region)
                
                # Lưu ảnh đã xử lý để debug
                processed_path = debug_dir / f"processed_class_{class_id}_{timestamp}.png"
                processed_region.save(str(processed_path))
                print(f"Đã lưu ảnh đã xử lý tại: {processed_path}")
                
                # Khởi tạo EasyOCR nếu chưa có
                if not hasattr(self.ocr_system, 'ocr_reader') or self.ocr_system.ocr_reader is None:
                    import easyocr
                    self.ocr_system.ocr_reader = easyocr.Reader(['vi'], gpu=False)
                
                # Chuẩn bị ngôn ngữ OCR
                languages = ['vi']
                if class_id in [7, 8, 4]:  # Noi_Nhan, So_Ki_Hieu, Loai_VB
                    languages = ['vi', 'en']
                
                # Thực hiện OCR
                img_array = np.array(processed_region)
                
                # Xử lý đặc biệt cho Noi_Nhan
                if class_id == 7:  # Noi_Nhan
                    # Thử với paragraph mode trước
                    results = self.ocr_system.ocr_reader.readtext(img_array, paragraph=True)
                    
                    # Nếu không có kết quả hoặc kết quả quá ngắn, thử lại với paragraph=False
                    if not results or len(results) == 0 or len("\n".join([r[1] for r in results])) < 10:
                        results = self.ocr_system.ocr_reader.readtext(img_array, paragraph=False)
                else:
                    results = self.ocr_system.ocr_reader.readtext(img_array)
                
                # Xử lý kết quả
                if results:
                    # Trích xuất text từ kết quả
                    if class_id == 7:  # Noi_Nhan - xử lý đặc biệt
                        # Sắp xếp kết quả theo tọa độ y (từ trên xuống dưới)
                        sorted_results = sorted(results, key=lambda r: (r[0][0][1] + r[0][2][1])/2)
                        easy_ocr_text = "\n".join([r[1] for r in sorted_results])
                    else:
                        easy_ocr_text = "\n".join([r[1] for r in results])
                    
                    final_text = easy_ocr_text
                    easyocr_success = len(final_text) > 0
            except Exception as e:
                print(f"Lỗi EasyOCR: {str(e)}")
            
            # 2. Thử với Tesseract nếu EasyOCR không cho kết quả tốt
            tesseract_text = ""
            if not easyocr_success or len(final_text) < 10:
                tesseract_installed = True
                try:
                    import pytesseract
                    pytesseract.get_tesseract_version()
                except (ImportError, Exception):
                    tesseract_installed = False
                    print("Tesseract không được cài đặt hoặc không tìm thấy trong PATH")
                
                if tesseract_installed:
                    try:
                        # Xử lý ảnh đặc biệt cho Tesseract
                        from PIL import ImageEnhance
                        
                        # Chuyển sang ảnh xám
                        gray_img = region.convert('L')
                        
                        # Tăng độ tương phản
                        enhancer = ImageEnhance.Contrast(gray_img)
                        enhanced_img = enhancer.enhance(2.0)
                        
                        # Lưu ảnh tăng cường
                        enhanced_path = debug_dir / f"tesseract_enhanced_{class_id}_{timestamp}.png"
                        enhanced_img.save(str(enhanced_path))
                        
                        img_array = np.array(enhanced_img)
                        
                        # Thử các PSM khác nhau dựa trên loại vùng
                        if class_id == 7:  # Noi_Nhan
                            custom_config = '--oem 1 --psm 6 -l vie'  # PSM 6 cho block text
                        elif class_id == 5:  # ND_Chinh
                            custom_config = '--oem 1 --psm 6 -l vie'  # PSM 6 cho block text
                        else:
                            custom_config = '--oem 1 --psm 4 -l vie'  # PSM 4 cho single column
                        
                        # Thực hiện OCR
                        tesseract_text = pytesseract.image_to_string(img_array, config=custom_config)
                        
                        # Nếu kết quả tesseract dài hơn, dùng kết quả của tesseract
                        if not final_text or len(tesseract_text.strip()) > len(final_text.strip()):
                            final_text = tesseract_text
                    except Exception as e:
                        print(f"Lỗi Tesseract: {str(e)}")
            
            # Nếu không có Tesseract và EasyOCR không thành công, thử phương pháp cuối cùng
            if not final_text.strip():
                # Thử một lần cuối với chế độ đặc biệt của EasyOCR
                try:
                    # Xử lý ảnh đặc biệt
                    gray_img = region.convert('L')
                    # Tăng độ tương phản
                    from PIL import ImageEnhance
                    enhancer = ImageEnhance.Contrast(gray_img)
                    enhanced_img = enhancer.enhance(2.0)  # Tăng độ tương phản
                    # Scale lên để dễ nhận dạng hơn
                    scaled_img = enhanced_img.resize((enhanced_img.width*2, enhanced_img.height*2), Image.LANCZOS)
                    
                    # Lưu ảnh tăng cường để debug
                    enhanced_path = debug_dir / f"enhanced_class_{class_id}_{timestamp}.png"
                    scaled_img.save(str(enhanced_path))
                    
                    # Thử OCR với ảnh tăng cường
                    results = self.ocr_system.ocr_reader.readtext(np.array(scaled_img), detail=0)
                    if results:
                        final_text = "\n".join(results)
                except Exception as e:
                    print(f"Lỗi xử lý ảnh nâng cao: {str(e)}")
                    
            # Nếu vẫn không thành công, thông báo lỗi
            if not final_text.strip():
                QMessageBox.warning(self, "OCR Error", "Không thể trích xuất text từ vùng đã chọn")
                return
                
            # Xử lý kết quả dựa vào loại class
            if class_id == 7:  # Noi_Nhan
                # Xử lý đặc biệt cho Noi nhan
                lines = final_text.split('\n')
                final_text = "\n".join([line.strip() for line in lines if line.strip()])
                
                # Chuẩn hóa các lỗi thường gặp
                final_text = final_text.replace("- ", "- ")
                final_text = final_text.replace("..", ".")
            
            elif class_id == 0:  # CQBH
                # Chuẩn hóa tên cơ quan
                common_typos = {
                    "UBND TINH": "UBND TỈNH",
                    "UBND THANH PHO": "UBND THÀNH PHỐ",
                    "CONG HOA": "CỘNG HÒA",
                    "DOC LAP": "ĐỘC LẬP"
                }
                for typo, correct in common_typos.items():
                    final_text = final_text.replace(typo, correct)
            
            elif class_id == 6:  # Ngay_BH
                # Chuẩn hóa định dạng ngày tháng
                final_text = self.standardize_date(final_text)
            
            # Cập nhật UI với kết quả OCR
            self.update_ocr_field(class_name, final_text)
            self.statusBar().showMessage(f"Đã trích xuất thành công {class_name}", 3000)
            
        except Exception as e:
            logger.error(f"Error processing custom box: {str(e)}")
            traceback.print_exc()
            QMessageBox.warning(self, "Error", f"Lỗi xử lý OCR: {str(e)}")
    
    def standardize_date(self, date_text):
        """Chuẩn hóa định dạng ngày tháng"""
        try:
            # Loại bỏ các ký tự không cần thiết
            date_text = date_text.strip()
            # Chuẩn hóa dấu ngày tháng
            date_text = date_text.replace('/', '-').replace('.', '-')
            # Sửa các số hay nhận nhầm
            date_text = date_text.replace('l', '1').replace('O', '0').replace('o', '0')
            
            # Tìm mẫu ngày tháng DD-MM-YYYY hoặc D-M-YYYY
            date_pattern = r'(\d{1,2})[-./](\d{1,2})[-./](\d{2,4})'
            import re
            match = re.search(date_pattern, date_text)
            if match:
                day, month, year = match.groups()
                # Đảm bảo định dạng DD-MM-YYYY
                if len(year) == 2:
                    year = '20' + year  # Giả sử năm hiện tại là thế kỷ 21
                return f"ngày {int(day)} tháng {int(month)} năm {year}"
            else:
                # Tìm kiểu "ngày X tháng Y năm Z"
                vn_date_pattern = r'ngày\s+(\d{1,2})\s+tháng\s+(\d{1,2})\s+năm\s+(\d{2,4})'
                match = re.search(vn_date_pattern, date_text, re.IGNORECASE)
                if match:
                    day, month, year = match.groups()
                    # Đảm bảo định dạng
                    if len(year) == 2:
                        year = '20' + year
                    return f"ngày {int(day)} tháng {int(month)} năm {year}"
            
            return date_text
        except Exception as e:
            print(f"Lỗi chuẩn hóa ngày: {e}")
            return date_text
    
    def update_ocr_field(self, class_name, text):
        """Cập nhật trường dữ liệu tương ứng với class_name"""
        if not text.strip():
            return
            
        # Ánh xạ class_name đến field trong OCRResultEditor
        field_mapping = {
            'CQBH': 'cqbh',
            'Chu_Ky': 'chu_ky',
            'Chuc_Vu': 'chuc_vu',
            'Do_Khan': 'do_khan',
            'Loai_VB': 'loai_vb',
            'ND_Chinh': 'nd_chinh',
            'Ngay_BH': 'ngay_bh',
            'Noi_Nhan': 'noi_nhan',
            'So_Ki_Hieu': 'so_ki_hieu'
        }
        
        field_id = field_mapping.get(class_name)
        
        if not field_id or field_id not in self.ocr_editor.fields:
            return
            
        editor = self.ocr_editor.fields[field_id]
        
        # Xử lý đặc biệt cho CQBH (có thể là trên hoặc dưới)
        if field_id == 'cqbh':
            current_selection = self.ocr_editor.cqbh_selector.currentText()
            current_text = editor.text()
            
            if current_selection == "Trên":
                # Sửa lỗi f-string
                second_part = ''
                if '\n' in current_text:
                    second_part = current_text.split('\n')[-1]
                self.ocr_editor.full_cqbh_text = f"{text}\n{second_part}"
                editor.setText(text)
            elif current_selection == "Dưới":
                # Sửa lỗi f-string
                first_part = ''
                if '\n' in current_text:
                    first_part = current_text.split('\n')[0]
                self.ocr_editor.full_cqbh_text = f"{first_part}\n{text}"
                editor.setText(text)
            else:  # Tất cả
                self.ocr_editor.full_cqbh_text = text
                editor.setText(text)
                
        elif field_id == 'do_khan':
            # Tìm index gần nhất trong combobox
            index = -1
            for i in range(editor.count()):
                if text.lower() in editor.itemText(i).lower():
                    index = i
                    break
            
            if index >= 0:
                editor.setCurrentIndex(index)
                
        elif isinstance(editor, QTextEdit):
            editor.setPlainText(text)
            
        else:
            editor.setText(text)
            
        # Báo hiệu có thay đổi
        self.ocr_editor.content_changed()
        
        # Thông báo thành công
        self.statusBar().showMessage(f"Đã cập nhật trường {class_name}", 3000)

    def debug_coords_to_file(self, orig_rect, adjusted_rect, img_dims, class_id):
        """Lưu thông tin tọa độ để debug"""
        try:
            debug_dir = TEMP_DIR / "debug"
            debug_dir.mkdir(exist_ok=True)
            timestamp = int(time.time())
            debug_file = debug_dir / f"coords_debug_{class_id}_{timestamp}.txt"
            
            with open(debug_file, 'w', encoding='utf-8') as f:
                f.write(f"===== DEBUG TỌA ĐỘ BOX: CLASS_ID={class_id} =====\n")
                f.write(f"Zoom level: {self.pdf_viewer.zoom_level}%\n")
                f.write(f"Kích thước ảnh gốc: {img_dims[0]}x{img_dims[1]}\n\n")
                
                f.write("Tọa độ ban đầu (viewport):\n")
                f.write(f"  x: {orig_rect.x()}, y: {orig_rect.y()}, w: {orig_rect.width()}, h: {orig_rect.height()}\n\n")
                
                f.write("Tọa độ đã điều chỉnh (ảnh gốc):\n")
                f.write(f"  x: {adjusted_rect.x()}, y: {adjusted_rect.y()}, w: {adjusted_rect.width()}, h: {adjusted_rect.height()}\n\n")
                
                # Thêm thông tin cuộn trang
                f.write("Thông tin cuộn trang:\n")
                h_scroll = self.pdf_viewer.scroll_area.horizontalScrollBar().value()
                v_scroll = self.pdf_viewer.scroll_area.verticalScrollBar().value()
                f.write(f"  h_scroll: {h_scroll}, v_scroll: {v_scroll}\n\n")
                
                # Thêm thông tin offset
                display_size = self.pdf_viewer.image_label.size()
                pixmap_size = self.pdf_viewer.pixmap.size() if hasattr(self.pdf_viewer, 'pixmap') else QSize(0, 0)
                offset_x = max(0, (display_size.width() - pixmap_size.width()) / 2)
                offset_y = max(0, (display_size.height() - pixmap_size.height()) / 2)
                f.write(f"Offset: offset_x={offset_x}, offset_y={offset_y}\n")
                f.write(f"Display size: {display_size.width()}x{display_size.height()}\n")
                f.write(f"Pixmap size: {pixmap_size.width()}x{pixmap_size.height()}\n")
            
            print(f"Đã lưu thông tin debug tọa độ tại: {debug_file}")
            return str(debug_file)
        except Exception as e:
            print(f"Lỗi khi lưu debug tọa độ: {str(e)}")
            return None
def main():
    # Must set these attributes before creating QApplication
    QApplication.setAttribute(Qt.AA_EnableHighDpiScaling, True)
    QApplication.setAttribute(Qt.AA_UseHighDpiPixmaps, True)

    app = QApplication(sys.argv)
    
    # Set application info
    app.setApplicationName("OCR Document Manager")
    app.setApplicationVersion(APP_VERSION)
    app.setOrganizationName("OCRApp")
    
    # Set application style
    app.setStyle(QStyleFactory.create('Fusion'))
    
    # Register custom fonts if available
    font_dir = BASE_DIR / 'fonts'
    if font_dir.exists():
        for font_file in font_dir.glob('*.ttf'):
            QFontDatabase.addApplicationFont(str(font_file))
    
    # Create and show main window
    window = MainWindow()
    window.show()
    
    # Start event loop
    sys.exit(app.exec_())

if __name__ == "__main__":
    main()